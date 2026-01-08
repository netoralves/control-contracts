from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponse, JsonResponse, FileResponse
from django.contrib import messages
from django.db.models import Q, Sum, Value, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group, Permission, User
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.utils.encoding import smart_str
from django.views.decorators.http import require_GET
from django.core.serializers.json import DjangoJSONEncoder
from django.views.decorators.http import require_http_methods
from django.utils.timezone import now
from django.utils.timezone import is_aware
from pandas._libs.tslibs.nattype import NaTType
from datetime import datetime, timedelta
from django.utils import timezone
from django import forms
from decimal import Decimal


import csv, os, json, io
import pandas as pd

from .models import (
    Cliente,
    Contrato,
    ItemContrato,
    ItemFornecedor,
    OrdemFornecimento,
    OrdemServico,
    ImportExportLog,
    Colaborador,
    SLA,
    Tarefa,
    LancamentoHora,
    CentroCusto,
    Projeto,
    Backlog,
    Sprint,
    FeedbackSprintOS,
)
from .models import (
    TermoAditivo,
    RegimeLegal,
    TipoTermoAditivo,
)
from .services import ContratoService
from .forms import (
    ClienteForm,
    ContratoForm,
    ItemContratoForm,
    ItemFornecedorForm,
    OrdemFornecimentoForm,
    OrdemServicoForm,
    LancamentoHoraForm,
    ContratoPublicoForm,
    TermoAditivoForm,
    AnaliseContratoForm,
    PlanoTrabalhoForm,
    SLAForm,
    FeedbackSprintOSForm,
    CriarTicketContatoForm,
)
from .models import AnaliseContrato, DocumentoContrato, PlanoTrabalho, SLAImportante, ClausulaCritica, MatrizRACI, QuadroPenalizacao
from .utils import map_tipo_item_contrato_para_fornecedor
from decimal import Decimal


CLIENTE_LIST_COLUMNS = [
    {"id": "id", "label": "ID"},
    {"id": "razao", "label": "Razão Social"},
    {"id": "fantasia", "label": "Nome Fantasia"},
    {"id": "cidade", "label": "Cidade"},
    {"id": "estado", "label": "Estado"},
    {"id": "documento", "label": "CNPJ/CPF"},
    {"id": "telefone", "label": "Telefone"},
    {"id": "status", "label": "Status"},
]

CONTRATO_LIST_COLUMNS = [
    {"id": "numero", "label": "Nº Contrato"},
    {"id": "cliente", "label": "Cliente"},
    {"id": "fornecedor", "label": "Fornecedores"},
    {"id": "assinatura", "label": "Data Assinatura"},
    {"id": "fim", "label": "Data Fim"},
    {"id": "valor", "label": "Valor Total (R$)"},
    {"id": "valor_faturado", "label": "Valor Total Faturado (R$)"},
    {"id": "valor_nao_faturado", "label": "Valor Total Não Faturado (R$)"},
    {"id": "situacao", "label": "Situação"},
]

ITEM_CONTRATO_LIST_COLUMNS = [
    {"id": "lote", "label": "Lote"},
    {"id": "numero", "label": "Nº Item"},
    {"id": "contrato", "label": "Contrato"},
    {"id": "descricao", "label": "Descrição"},
    {"id": "unidade", "label": "Unidade"},
    {"id": "quantidade", "label": "Qtd"},
    {"id": "consumida", "label": "Qtd. Consumida"},
    {"id": "saldo", "label": "Qtd. Saldo"},
]

ITEM_FORNECEDOR_LIST_COLUMNS = [
    {"id": "fornecedor", "label": "Fornecedor"},
    {"id": "tipo", "label": "Tipo"},
    {"id": "sku", "label": "SKU"},
    {"id": "descricao", "label": "Descrição"},
    {"id": "unidade", "label": "Unidade"},
    {"id": "valor_unitario", "label": "Valor Unitário"},
]


# Controle de Login
def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        login(request, form.get_user())
        return redirect("dashboard")
    if request.method == "POST":
        messages.error(request, "Usuário ou senha inválidos.")

    return render(request, "contracts/login.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect("login")


# Dashboard
@login_required
def dashboard(request):
    hoje = timezone.now().date()
    
    # ========== INDICADORES BÁSICOS ==========
    total_clientes = Cliente.objects.count()
    total_contratos = Contrato.objects.count()
    contratos_ativos = Contrato.objects.filter(situacao="Ativo").count()
    contratos_inativos = Contrato.objects.filter(situacao="Inativo").count()

    # ========== INDICADORES FINANCEIROS ==========
    # Valor total dos contratos (soma de todos os itens)
    valor_total_contratos = (
        ItemContrato.objects.aggregate(
            total=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F("quantidade") * F("valor_unitario"),
                        output_field=DecimalField(max_digits=20, decimal_places=2),
                    )
                ),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=2)),
            )
        )["total"]
        or 0
    )

    # Valor faturado (OS e OF com status "faturada")
    valor_faturado_os = (
        OrdemServico.objects.filter(status="faturada").aggregate(
            total=Coalesce(
                Sum("valor_total"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=2)),
            )
        )["total"]
        or 0
    )

    valor_faturado_of = (
        OrdemFornecimento.objects.filter(status="faturada").aggregate(
            total=Coalesce(
                Sum("valor_total"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=2)),
            )
        )["total"]
        or 0
    )

    valor_total_faturado = valor_faturado_os + valor_faturado_of
    valor_nao_faturado = valor_total_contratos - valor_total_faturado
    
    # Taxa de utilização (% do valor faturado vs total)
    taxa_utilizacao = (
        (valor_total_faturado / valor_total_contratos * 100)
        if valor_total_contratos > 0
        else 0
    )
    
    # Valor médio por contrato
    valor_medio_contrato = (
        valor_total_contratos / total_contratos if total_contratos > 0 else 0
    )

    # ========== INDICADORES DE VENCIMENTO ==========
    # Contratos que vencem nos próximos 30, 60 e 90 dias
    data_30_dias = hoje + timedelta(days=30)
    data_60_dias = hoje + timedelta(days=60)
    data_90_dias = hoje + timedelta(days=90)
    
    contratos_vencendo_30 = Contrato.objects.filter(
        situacao="Ativo",
        data_fim__gte=hoje,
        data_fim__lte=data_30_dias
    ).count()
    
    contratos_vencendo_60 = Contrato.objects.filter(
        situacao="Ativo",
        data_fim__gte=hoje,
        data_fim__lte=data_60_dias
    ).count()
    
    contratos_vencendo_90 = Contrato.objects.filter(
        situacao="Ativo",
        data_fim__gte=hoje,
        data_fim__lte=data_90_dias
    ).count()
    
    contratos_vencidos = Contrato.objects.filter(
        situacao="Ativo",
        data_fim__lt=hoje
    ).count()

    # ========== INDICADORES DE EXECUÇÃO (OS/OF) ==========
    # OS/OF por status
    os_abertas = OrdemServico.objects.filter(status="aberta").count()
    os_execucao = OrdemServico.objects.filter(status="execucao").count()
    os_finalizadas = OrdemServico.objects.filter(status="finalizada").count()
    os_faturadas = OrdemServico.objects.filter(status="faturada").count()
    total_os = OrdemServico.objects.count()
    
    of_abertas = OrdemFornecimento.objects.filter(status="aberta").count()
    of_execucao = OrdemFornecimento.objects.filter(status="execucao").count()
    of_finalizadas = OrdemFornecimento.objects.filter(status="finalizada").count()
    of_faturadas = OrdemFornecimento.objects.filter(status="faturada").count()
    total_of = OrdemFornecimento.objects.count()
    
    # Taxa de execução (finalizadas + faturadas / total)
    taxa_execucao_os = (
        ((os_finalizadas + os_faturadas) / total_os * 100) if total_os > 0 else 0
    )
    taxa_execucao_of = (
        ((of_finalizadas + of_faturadas) / total_of * 100) if total_of > 0 else 0
    )

    # ========== INDICADORES DE CONSUMO ==========
    # Total de itens e consumo
    total_itens = ItemContrato.objects.count()
    itens_consumidos = ItemContrato.objects.filter(
        Q(ordemservico__status="faturada") | Q(ordemfornecimento__status="faturada")
    ).distinct().count()
    
    # Quantidade total vs consumida
    quantidade_total = ItemContrato.objects.aggregate(
        total=Coalesce(Sum("quantidade"), Value(0), output_field=DecimalField())
    )["total"] or 0
    
    quantidade_consumida_os = OrdemServico.objects.filter(
        status="faturada"
    ).aggregate(
        total=Coalesce(Sum("quantidade"), Value(0), output_field=DecimalField())
    )["total"] or 0
    
    quantidade_consumida_of = OrdemFornecimento.objects.filter(
        status="faturada"
    ).aggregate(
        total=Coalesce(Sum("quantidade"), Value(0), output_field=DecimalField())
    )["total"] or 0
    
    quantidade_total_consumida = quantidade_consumida_os + quantidade_consumida_of
    taxa_consumo = (
        (quantidade_total_consumida / quantidade_total * 100)
        if quantidade_total > 0
        else 0
    )

    # ========== TOP CLIENTES POR VALOR ==========
    top_clientes = (
        Contrato.objects.values("cliente__nome_fantasia", "cliente__nome_razao_social")
        .annotate(
            valor_total=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F("itens__quantidade") * F("itens__valor_unitario"),
                        output_field=DecimalField(max_digits=20, decimal_places=2),
                    )
                ),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=2)),
            )
        )
        .order_by("-valor_total")[:5]
    )
    
    top_clientes_list = [
        {
            "nome": c["cliente__nome_fantasia"] or c["cliente__nome_razao_social"],
            "valor": c["valor_total"] or 0,
        }
        for c in top_clientes
    ]

    # ========== DISTRIBUIÇÃO POR FORNECEDOR ==========
    # Contagem de contratos por fornecedor (primeiro fornecedor de cada contrato)
    contratos_por_fornecedor = {}
    for contrato in Contrato.objects.filter(situacao="Ativo"):
        fornecedores = contrato.fornecedores or []
        if fornecedores:
            fornecedor_principal = fornecedores[0]
            contratos_por_fornecedor[fornecedor_principal] = (
                contratos_por_fornecedor.get(fornecedor_principal, 0) + 1
            )

    # ========== NOVOS INDICADORES DE DESEMPENHO ==========
    # Saldo Disponível Total (valor não faturado)
    saldo_disponivel_total = valor_nao_faturado
    
    # Contratos com baixa utilização (< 50% faturado)
    contratos_baixa_utilizacao = []
    for contrato in Contrato.objects.filter(situacao="Ativo"):
        valor_total_contrato = contrato.get_valor_total_itens()
        valor_faturado_contrato = contrato.get_valor_total_faturado()
        if valor_total_contrato > 0:
            taxa_util_contrato = (valor_faturado_contrato / valor_total_contrato) * 100
            if taxa_util_contrato < 50:
                contratos_baixa_utilizacao.append({
                    "numero": contrato.numero_contrato,
                    "cliente": str(contrato.cliente),
                    "valor_total": float(valor_total_contrato),
                    "valor_faturado": float(valor_faturado_contrato),
                    "taxa_utilizacao": round(taxa_util_contrato, 2),
                })
    
    # Ordenar por taxa de utilização (menor primeiro)
    contratos_baixa_utilizacao.sort(key=lambda x: x["taxa_utilizacao"])
    contratos_baixa_utilizacao_count = len(contratos_baixa_utilizacao)
    
    # Itens com saldo crítico (< 10% do saldo inicial)
    itens_saldo_critico = []
    for item in ItemContrato.objects.all():
        if item.saldo_quantidade_inicial > 0:
            percentual_saldo = (item.saldo_quantidade_atual / item.saldo_quantidade_inicial) * 100
            if percentual_saldo < 10 and item.saldo_quantidade_atual > 0:
                itens_saldo_critico.append({
                    "numero": item.numero_item,
                    "descricao": item.descricao[:50],
                    "contrato": str(item.contrato.numero_contrato),
                    "saldo_atual": float(item.saldo_quantidade_atual),
                    "saldo_inicial": float(item.saldo_quantidade_inicial),
                    "percentual": round(percentual_saldo, 2),
                })
    
    # Top 5 itens mais consumidos
    top_itens_consumidos = (
        ItemContrato.objects.annotate(
            quantidade_consumida_total=ExpressionWrapper(
                Coalesce(
                    Sum("ordemservico__quantidade", filter=Q(ordemservico__status="faturada")),
                    Value(0, output_field=DecimalField())
                ) + Coalesce(
                    Sum("ordemfornecimento__quantidade", filter=Q(ordemfornecimento__status="faturada")),
                    Value(0, output_field=DecimalField())
                ),
                output_field=DecimalField()
            )
        )
        .filter(quantidade_consumida_total__gt=0)
        .order_by("-quantidade_consumida_total")[:5]
    )
    
    top_itens_list = []
    for item in top_itens_consumidos:
        top_itens_list.append({
            "descricao": item.descricao[:40] if len(item.descricao) > 40 else item.descricao,
            "contrato": item.contrato.numero_contrato,
            "quantidade_consumida": float(item.quantidade_consumida_total) if item.quantidade_consumida_total else 0.0,
        })
    
    # Top 5 fornecedores por valor faturado
    fornecedores_faturamento = {}
    for contrato in Contrato.objects.filter(situacao="Ativo"):
        fornecedores = contrato.fornecedores or []
        valor_faturado_contrato = contrato.get_valor_total_faturado()
        for fornecedor in fornecedores:
            if fornecedor not in fornecedores_faturamento:
                fornecedores_faturamento[fornecedor] = Decimal(0)
            fornecedores_faturamento[fornecedor] += Decimal(str(valor_faturado_contrato)) / len(fornecedores) if fornecedores else Decimal(0)
    
    top_fornecedores = sorted(
        fornecedores_faturamento.items(),
        key=lambda x: x[1],
        reverse=True
    )[:5]
    
    top_fornecedores_list = [
        {
            "nome": fornecedor,
            "valor_faturado": float(valor),
        }
        for fornecedor, valor in top_fornecedores
    ]
    
    # Tempo médio de execução OS (em dias)
    os_com_tempo = OrdemServico.objects.filter(
        status__in=["finalizada", "faturada"],
        data_inicio__isnull=False,
        data_termino__isnull=False,
    )
    
    tempos_execucao_os = []
    for os_item in os_com_tempo:
        if os_item.data_inicio and os_item.data_termino:
            dias = (os_item.data_termino - os_item.data_inicio).days
            if dias > 0:
                tempos_execucao_os.append(dias)
    
    tempo_medio_execucao_os = sum(tempos_execucao_os) / len(tempos_execucao_os) if tempos_execucao_os else 0
    
    # Tempo médio de execução OF (em dias)
    of_com_tempo = OrdemFornecimento.objects.filter(
        status__in=["finalizada", "faturada"],
        data_ativacao__isnull=False,
    )
    
    tempos_execucao_of = []
    for of_item in of_com_tempo:
        if of_item.data_ativacao:
            # Usar data de criação como início
            # data_ativacao já é um DateField (date), criado_em é DateTimeField (datetime)
            dias = (of_item.data_ativacao - of_item.criado_em.date()).days
            if dias > 0:
                tempos_execucao_of.append(dias)
    
    tempo_medio_execucao_of = sum(tempos_execucao_of) / len(tempos_execucao_of) if tempos_execucao_of else 0
    
    # ========== GRÁFICOS ==========
    # Gráfico de faturamento mensal (últimos 6 meses)
    meses_labels = []
    meses_valores = []
    for i in range(5, -1, -1):
        data_mes = hoje - timedelta(days=30 * i)
        mes_label = data_mes.strftime("%b/%Y")
        meses_labels.append(mes_label)
        
        # Faturamento do mês (OS e OF faturadas no mês)
        valor_mes_os = (
            OrdemServico.objects.filter(
                status="faturada",
                data_faturamento__year=data_mes.year,
                data_faturamento__month=data_mes.month,
            ).aggregate(
                total=Coalesce(
                    Sum("valor_total"),
                    Value(0, output_field=DecimalField()),
                )
            )["total"]
            or 0
        )
        
        valor_mes_of = (
            OrdemFornecimento.objects.filter(
                status="faturada",
                data_faturamento__year=data_mes.year,
                data_faturamento__month=data_mes.month,
            ).aggregate(
                total=Coalesce(
                    Sum("valor_total"),
                    Value(0, output_field=DecimalField()),
                )
            )["total"]
            or 0
        )
        
        meses_valores.append(float(valor_mes_os + valor_mes_of))

    # Garantir que todos os valores sejam float
    meses_valores_float = [float(v) if v else 0.0 for v in meses_valores]
    
    # Passar objetos Python diretamente (json_script fará a serialização)
    grafico_faturamento = {
        "labels": meses_labels,
            "datasets": [
                {
                    "label": "Faturamento",
                "data": meses_valores_float,
                "backgroundColor": "rgba(59, 130, 246, 0.6)",
            }
        ],
    }

    # Gráfico de contratos por cliente (top 5)
    if top_clientes_list:
        grafico_contratos_por_cliente = {
            "labels": [str(c.get("nome", ""))[:20] for c in top_clientes_list if c.get("nome")],
            "datasets": [
                {
                    "label": "Valor (R$)",
                    "data": [float(c.get("valor", 0) or 0) for c in top_clientes_list],
                    "backgroundColor": [
                        "rgba(59, 130, 246, 0.6)",
                        "rgba(16, 185, 129, 0.6)",
                        "rgba(245, 158, 11, 0.6)",
                        "rgba(239, 68, 68, 0.6)",
                        "rgba(139, 92, 246, 0.6)",
            ],
                }
            ],
        }
    else:
        grafico_contratos_por_cliente = {
            "labels": ["Sem Dados"],
            "datasets": [
                {
                    "label": "Contratos",
                    "data": [1],
                    "backgroundColor": ["rgba(255, 99, 132, 0.6)"],
                }
            ],
        }

    # Gráfico de status de contratos
    grafico_status_contratos = {
        "labels": ["Ativos", "Inativos"],
        "datasets": [
            {
                "label": "Contratos",
                "data": [contratos_ativos, contratos_inativos],
                "backgroundColor": [
                    "rgba(16, 185, 129, 0.6)",
                    "rgba(239, 68, 68, 0.6)",
                ],
            }
        ],
    }
    
    # Gráfico de status de OS/OF
    grafico_status_os_of = {
        "labels": ["Abertas", "Em Execução", "Finalizadas", "Faturadas"],
        "datasets": [
            {
                "label": "OS",
                "data": [int(os_abertas or 0), int(os_execucao or 0), int(os_finalizadas or 0), int(os_faturadas or 0)],
                "backgroundColor": "rgba(59, 130, 246, 0.6)",
            },
            {
                "label": "OF",
                "data": [int(of_abertas or 0), int(of_execucao or 0), int(of_finalizadas or 0), int(of_faturadas or 0)],
                "backgroundColor": "rgba(16, 185, 129, 0.6)",
            },
        ],
    }
    
    # Gráfico de distribuição por fornecedor (valor faturado)
    if top_fornecedores_list:
        grafico_fornecedores = {
            "labels": [str(f.get("nome", ""))[:20] for f in top_fornecedores_list if f.get("nome")],
            "datasets": [
                {
                    "label": "Valor Faturado (R$)",
                    "data": [float(f.get("valor_faturado", 0) or 0) for f in top_fornecedores_list],
                    "backgroundColor": [
                        "rgba(59, 130, 246, 0.6)",
                        "rgba(16, 185, 129, 0.6)",
                        "rgba(245, 158, 11, 0.6)",
                        "rgba(239, 68, 68, 0.6)",
                        "rgba(139, 92, 246, 0.6)",
                    ],
                }
            ],
        }
    else:
        grafico_fornecedores = {
            "labels": ["Sem Dados"],
            "datasets": [
                {
                    "label": "Fornecedores",
                    "data": [1],
                    "backgroundColor": ["rgba(255, 99, 132, 0.6)"],
                }
            ],
        }

    # ========== DADOS ADICIONAIS PARA O TEMPLATE ==========
    # Contratos recentes (últimos 5)
    contratos_recentes = (
        Contrato.objects.select_related("cliente")
        .order_by("-data_assinatura")[:5]
    )
    
    # Próximos vencimentos (próximos 30 dias)
    proximos_vencimentos_list = []
    for contrato in Contrato.objects.filter(
        situacao="Ativo",
        data_fim__gte=hoje,
        data_fim__lte=hoje + timedelta(days=30)
    ).select_related("cliente")[:5]:
        dias_restantes = (contrato.data_fim - hoje).days
        proximos_vencimentos_list.append({
            "cliente": str(contrato.cliente),
            "numero": contrato.numero_contrato,
            "dias_restantes": dias_restantes,
        })
    
    # Atividades recentes (placeholder - pode ser expandido com um modelo de log)
    atividades_recentes = []
    
    # Valores financeiros adicionais (placeholders - podem ser calculados se houver modelo de pagamentos)
    a_receber = valor_nao_faturado  # Valor não faturado pode ser considerado "a receber"
    em_atraso = Decimal(0)  # Placeholder - requer modelo de pagamentos
    recebido = valor_total_faturado  # Valor faturado pode ser considerado "recebido"
    
    # Garantir que os gráficos não sejam None
    if grafico_faturamento is None:
        grafico_faturamento = {"labels": ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun"], "datasets": [{"label": "Faturamento", "data": [0, 0, 0, 0, 0, 0]}]}
    if grafico_contratos_por_cliente is None:
        grafico_contratos_por_cliente = {"labels": ["Sem Dados"], "datasets": [{"label": "Valor (R$)", "data": [1]}]}
    if grafico_status_os_of is None:
        grafico_status_os_of = {"labels": ["Abertas", "Em Execução", "Finalizadas", "Faturadas"], "datasets": [{"label": "OS", "data": [0, 0, 0, 0]}, {"label": "OF", "data": [0, 0, 0, 0]}]}
    if grafico_fornecedores is None:
        grafico_fornecedores = {"labels": ["Sem Dados"], "datasets": [{"label": "Valor Faturado (R$)", "data": [1]}]}
    
    # ========== CONTEXT ==========
    context = {
        # Básicos
        "total_clientes": total_clientes,
        "total_contratos": total_contratos,
        "contratos_ativos": contratos_ativos,
        "contratos_inativos": contratos_inativos,
        # Financeiros
        "valor_total_contratos": valor_total_contratos,
        "valor_total_faturado": valor_total_faturado,
        "valor_nao_faturado": valor_nao_faturado,
        "taxa_utilizacao": round(taxa_utilizacao, 2),
        "valor_medio_contrato": valor_medio_contrato,
        "saldo_disponivel_total": saldo_disponivel_total,
        "a_receber": a_receber,
        "em_atraso": em_atraso,
        "recebido": recebido,
        # Vencimento
        "contratos_vencendo_30": contratos_vencendo_30,
        "contratos_vencendo_60": contratos_vencendo_60,
        "contratos_vencendo_90": contratos_vencendo_90,
        "contratos_vencidos": contratos_vencidos,
        "proximos_vencimentos": proximos_vencimentos_list,
        # Execução
        "os_abertas": os_abertas,
        "os_execucao": os_execucao,
        "os_finalizadas": os_finalizadas,
        "os_faturadas": os_faturadas,
        "total_os": total_os,
        "of_abertas": of_abertas,
        "of_execucao": of_execucao,
        "of_finalizadas": of_finalizadas,
        "of_faturadas": of_faturadas,
        "total_of": total_of,
        "taxa_execucao_os": round(taxa_execucao_os, 2),
        "taxa_execucao_of": round(taxa_execucao_of, 2),
        "tempo_medio_execucao_os": round(tempo_medio_execucao_os, 1),
        "tempo_medio_execucao_of": round(tempo_medio_execucao_of, 1),
        # Consumo
        "total_itens": total_itens,
        "quantidade_total": quantidade_total,
        "quantidade_total_consumida": quantidade_total_consumida,
        "taxa_consumo": round(taxa_consumo, 2),
        # Novos indicadores
        "contratos_baixa_utilizacao": contratos_baixa_utilizacao[:5],  # Top 5
        "contratos_baixa_utilizacao_count": contratos_baixa_utilizacao_count,
        "itens_saldo_critico": itens_saldo_critico[:5],  # Top 5
        "top_itens_consumidos": top_itens_list,
        "top_fornecedores": top_fornecedores_list,
        # Top clientes
        "top_clientes": top_clientes_list,
        # Dados adicionais
        "contratos_recentes": contratos_recentes,
        "atividades_recentes": atividades_recentes,
        # Gráficos
        "grafico_faturamento": grafico_faturamento,
        "grafico_contratos_por_cliente": grafico_contratos_por_cliente,
        "grafico_status_contratos": grafico_status_contratos,
        "grafico_status_os_of": grafico_status_os_of,
        "grafico_fornecedores": grafico_fornecedores,
    }
    return render(request, "contracts/dashboard.html", context)


# Controle de Permissões
def group_required(*group_names):
    def in_groups(user):
        return user.is_authenticated and (
            bool(user.groups.filter(name__in=group_names)) or user.is_superuser
        )

    return user_passes_test(in_groups)


# APIs para carregamento dinâmico
@require_GET
def api_contratos_por_cliente(request):
    cliente_id = request.GET.get("cliente_id")
    if not cliente_id:
        return JsonResponse({"error": "Cliente não informado."}, status=400)

    contratos = Contrato.objects.filter(cliente_id=cliente_id)
    data = [{"id": c.id, "numero": c.numero_contrato} for c in contratos]
    return JsonResponse({"contratos": data})


@require_GET
def api_itens_contrato_por_contrato(request):
    contrato_id = request.GET.get("contrato_id")
    tipos = request.GET.get("tipo", "").split(",")
    itens = []

    if contrato_id:
        queryset = ItemContrato.objects.filter(contrato_id=contrato_id, tipo__in=tipos)

        for item in queryset:
            itens.append(
                {
                    "id": item.id,
                    "descricao": f"{item.numero_item} - {item.descricao}",
                    "valor_unitario": float(item.valor_unitario),
                }
            )

    return JsonResponse({"itens_contrato": itens})


@require_GET
def api_itens_fornecedor_por_item_contrato(request):
    contrato_id = request.GET.get("contrato")
    item_contrato_id = request.GET.get("item_contrato")

    if not contrato_id or not item_contrato_id:
        return JsonResponse([], safe=False)

    try:
        contrato = Contrato.objects.get(id=contrato_id)
        item_contrato = ItemContrato.objects.get(id=item_contrato_id)
    except (Contrato.DoesNotExist, ItemContrato.DoesNotExist):
        return JsonResponse([], safe=False)

    tipo_item = map_tipo_item_contrato_para_fornecedor(item_contrato.tipo)
    if not tipo_item:
        return JsonResponse([], safe=False)

    fornecedores = [f.upper() for f in contrato.fornecedores]
    itens = ItemFornecedor.objects.filter(
        tipo=tipo_item, fornecedor__in=fornecedores
    ).values("id", "descricao")
    return JsonResponse(list(itens), safe=False)


@require_GET
def api_itens_fornecedor_por_contrato(request):
    contrato_id = request.GET.get("contrato_id")

    if not contrato_id:
        return JsonResponse({"itens_contrato": []})

    try:
        contrato = Contrato.objects.get(id=contrato_id)
    except Contrato.DoesNotExist:
        return JsonResponse({"itens_contrato": []})

    itens = contrato.itens.filter(tipo__in=["servico", "treinamento"]).values(
        "id", "descricao"
    )
    return JsonResponse({"itens_contrato": list(itens)})


@require_GET
def api_item_contrato_saldo(request, item_id):
    try:
        item = ItemContrato.objects.get(id=item_id)
    except ItemContrato.DoesNotExist:
        return JsonResponse({"error": "Item não encontrado."}, status=404)

    saldo = item.saldo_quantidade_atual or 0
    return JsonResponse({
        "saldo": float(saldo),
        "unidade": item.unidade or ""
    })


@require_GET
def api_item_contrato_valor(request, item_id):
    """API para retornar o valor unitário de um item de contrato"""
    try:
        item = ItemContrato.objects.get(id=item_id)
        return JsonResponse({"valor_unitario": float(item.valor_unitario or 0)})
    except ItemContrato.DoesNotExist:
        return JsonResponse({"error": "Item não encontrado."}, status=404)


@require_GET
def api_item_fornecedor_valor(request, item_id):
    """API para retornar o valor unitário de um item de fornecedor"""
    try:
        item = ItemFornecedor.objects.get(id=item_id)
        return JsonResponse({"valor_unitario": float(item.valor_unitario or 0)})
    except ItemFornecedor.DoesNotExist:
        return JsonResponse({"error": "Item não encontrado."}, status=404)


@require_GET
def api_itens_fornecedor_servico_por_contrato(request):
    """API para retornar itens de fornecedor do tipo serviço vinculados a um contrato"""
    contrato_id = request.GET.get("contrato_id")
    
    if not contrato_id:
        return JsonResponse({"itens": []})
    
    try:
        contrato = Contrato.objects.get(id=contrato_id)
        fornecedores_contrato = [f.upper() for f in contrato.fornecedores] if contrato.fornecedores else []
        
        # Filtrar itens de serviço vinculados aos fornecedores do contrato
        itens = ItemFornecedor.objects.filter(tipo="servico")
        if fornecedores_contrato:
            itens = itens.filter(
                Q(fornecedor__in=fornecedores_contrato) |
                Q(outro_fornecedor__in=fornecedores_contrato)
            )
        
        itens_data = [
            {
                "id": item.id,
                "descricao": item.descricao,
                "valor_unitario": float(item.valor_unitario),
                "fornecedor": item.nome_fornecedor
            }
            for item in itens
        ]
        
        return JsonResponse({"itens": itens_data})
    except Contrato.DoesNotExist:
        return JsonResponse({"error": "Contrato não encontrado."}, status=404)


# Cliente - Listagem com filtros e paginação
@group_required("Admin", "Gerente", "Leitor")
def cliente_list(request):
    clientes = Cliente.objects.all()

    nome = request.GET.get("nome")
    cidade = request.GET.get("cidade")
    estado = request.GET.get("estado")
    ativo = request.GET.get("ativo")

    if nome:
        clientes = clientes.filter(nome_razao_social__icontains=nome)
    if cidade:
        clientes = clientes.filter(cidade__icontains=cidade)
    if estado:
        clientes = clientes.filter(estado__icontains=estado)
    if ativo == "ativo":
        clientes = clientes.filter(ativo=True)
    elif ativo == "inativo":
        clientes = clientes.filter(ativo=False)

    # Estatísticas
    total_clientes = Cliente.objects.count()
    clientes_ativos = Cliente.objects.filter(ativo=True).count()
    clientes_inativos = Cliente.objects.filter(ativo=False).count()
    clientes_com_contratos = Cliente.objects.filter(contratos__isnull=False).distinct().count()
    clientes_sem_contratos = total_clientes - clientes_com_contratos

    paginator = Paginator(clientes, 10)
    page = request.GET.get("page")
    clientes_page = paginator.get_page(page)

    return render(
        request,
        "cliente/list.html",
        {
            "clientes": clientes_page,
            "colunas": CLIENTE_LIST_COLUMNS,
            "total_clientes": total_clientes,
            "clientes_ativos": clientes_ativos,
            "clientes_inativos": clientes_inativos,
            "clientes_com_contratos": clientes_com_contratos,
            "clientes_sem_contratos": clientes_sem_contratos,
            "nome_filter": nome or "",
            "cidade_filter": cidade or "",
            "estado_filter": estado or "",
            "ativo_filter": ativo or "",
        },
    )


# Cliente - Criar
@group_required("Admin", "Gerente")
def cliente_create(request):
    from .forms import ClienteForm, ContatoClienteFormSet
    
    if request.method == "POST":
        form = ClienteForm(request.POST)
        formset = ContatoClienteFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            cliente = form.save()
            formset.instance = cliente
            formset.save()
            messages.success(request, "Cliente criado com sucesso!")
            return redirect("cliente_list")
    else:
        form = ClienteForm()
        formset = ContatoClienteFormSet()
    
    return render(request, "cliente/form.html", {
        "form": form,
        "formset": formset
    })


# Cliente - Editar
@group_required("Admin", "Gerente")
def cliente_update(request, pk):
    from .forms import ClienteForm, ContatoClienteFormSet
    
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == "POST":
        form = ClienteForm(request.POST, instance=cliente)
        formset = ContatoClienteFormSet(request.POST, instance=cliente)
        
        if form.is_valid() and formset.is_valid():
            cliente = form.save()
            formset.save()
            messages.success(request, "Cliente atualizado com sucesso!")
            return redirect("cliente_detail", pk=cliente.pk)
    else:
        form = ClienteForm(instance=cliente)
        formset = ContatoClienteFormSet(instance=cliente)
    
    return render(request, "cliente/form.html", {
        "form": form,
        "formset": formset,
        "cliente": cliente
    })


# Cliente - Detalhar (com abas para contratos, leads, oportunidades e reclamações)
@group_required("Admin", "Gerente", "Leitor")
def cliente_detail(request, pk):
    cliente = get_object_or_404(
        Cliente.objects.select_related('gerente_comercial', 'gerente_sucessos').prefetch_related('contatos'),
        pk=pk
    )
    
    # Contratos do cliente
    contratos = cliente.contratos.all().order_by('-data_assinatura')
    
    # Tickets de Customer Success do cliente
    from .models import FeedbackSprintOS
    tickets_cs = FeedbackSprintOS.objects.filter(cliente=cliente).select_related(
        'contrato', 'projeto', 'sprint', 'ordem_servico', 'gerente_sucessos'
    ).order_by('-criado_em')
    
    # Estatísticas
    total_contratos = contratos.count()
    contratos_ativos = contratos.filter(situacao="Ativo").count()
    total_tickets = tickets_cs.count()
    
    context = {
        "cliente": cliente,
        "contratos": contratos[:10],  # Últimos 10
        "tickets_cs": tickets_cs[:20],  # Últimos 20 tickets
        "total_contratos": total_contratos,
        "contratos_ativos": contratos_ativos,
        "total_tickets": total_tickets,
        "leads_ativos": 0,
        "oportunidades_ativas": 0,
        "reclamacoes_abertas": 0,
    }
    
    return render(request, "cliente/detail.html", context)


# Cliente - Deletar
@group_required("Admin", "Gerente")
def cliente_delete(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == "POST":
        cliente.delete()
        return redirect("cliente_list")
    return render(request, "cliente/confirm_delete.html", {"cliente": cliente})


# Cliente - Exportação CSV
@group_required("Admin", "Gerente")
def export_clientes_csv(request):
    clientes = Cliente.objects.all()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="clientes.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "Razão Social",
            "Nome Fantasia",
            "Tipo Cliente",
            "Tipo Pessoa",
            "CNPJ/CPF",
            "Natureza Jurídica",
            "Inscrição Estadual",
            "Inscrição Municipal",
            "Endereço",
            "Número",
            "Complemento",
            "Bairro",
            "Cidade",
            "Estado",
            "CEP",
            "País",
            "Nome Responsável",
            "Cargo Responsável",
            "Telefone Contato",
            "Email Contato",
            "Ativo",
        ]
    )

    for cliente in clientes:
        writer.writerow(
            [
                cliente.id,
                cliente.nome_razao_social,
                cliente.nome_fantasia,
                cliente.tipo_cliente,
                cliente.tipo_pessoa,
                cliente.cnpj_cpf,
                cliente.natureza_juridica,
                cliente.inscricao_estadual,
                cliente.inscricao_municipal,
                cliente.endereco,
                cliente.numero,
                cliente.complemento,
                cliente.bairro,
                cliente.cidade,
                cliente.estado,
                cliente.cep,
                cliente.pais,
                cliente.nome_responsavel,
                cliente.cargo_responsavel,
                cliente.telefone_contato,
                cliente.email_contato,
                "Ativo" if cliente.ativo else "Inativo",
            ]
        )

    return response


# Função get_contratos_queryset removida - usar gestao_contratos_list ao invés
# Views legadas de contrato removidas - usar gestao_contratos_* ao invés


# Função de exportação CSV legada removida - usar gestao_contratos ao invés


# Item de Contrato - Editar
@group_required("Admin", "Gerente")
def itemcontrato_update(request, pk):
    item = get_object_or_404(ItemContrato, pk=pk)
    if request.method == "POST":
        form = ItemContratoForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Item de contrato atualizado com sucesso!")
            return redirect("gestao_contratos_detail", pk=item.contrato.pk)
    else:
        form = ItemContratoForm(instance=item)
    return render(request, "item_contrato/form.html", {"form": form, "item": item})


# Item de Contrato - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def itemcontrato_detail(request, pk):
    item = get_object_or_404(ItemContrato, pk=pk)
    return render(request, "item_contrato/detail.html", {"item": item})


# Item de Contrato - Deletar
@group_required("Admin", "Gerente")
def itemcontrato_delete(request, pk):
    item = get_object_or_404(ItemContrato, pk=pk)
    contrato_pk = item.contrato.pk
    if request.method == "POST":
        item.delete()
        messages.success(request, "Item de contrato excluído com sucesso!")
        return redirect("gestao_contratos_detail", pk=contrato_pk)
    return render(request, "item_contrato/confirm_delete.html", {"object": item, "contrato": item.contrato})




# Item de Fornecedor - Listar com filtros e paginação
@group_required("Admin", "Gerente", "Leitor")
def itemfornecedor_list(request):
    itens = ItemFornecedor.objects.all()

    fornecedor = request.GET.get("fornecedor")
    tipo = request.GET.get("tipo")

    if fornecedor:
        itens = itens.filter(fornecedor=fornecedor)

    if tipo:
        itens = itens.filter(tipo=tipo)

    # Estatísticas
    total_itens = ItemFornecedor.objects.count()
    itens_produto = ItemFornecedor.objects.filter(tipo="produto").count()
    itens_servico = ItemFornecedor.objects.filter(tipo="servico").count()
    itens_treinamento = ItemFornecedor.objects.filter(tipo="treinamento").count()
    
    # Contar fornecedores únicos
    # Fornecedores padrão (excluindo "Outro Fornecedor")
    fornecedores_padrao = ItemFornecedor.objects.exclude(fornecedor="Outro Fornecedor").values('fornecedor').distinct().count()
    # Fornecedores customizados (outro_fornecedor)
    fornecedores_custom = ItemFornecedor.objects.filter(fornecedor="Outro Fornecedor").exclude(outro_fornecedor__isnull=True).exclude(outro_fornecedor__exact='').values('outro_fornecedor').distinct().count()
    fornecedores_unicos = fornecedores_padrao + fornecedores_custom

    paginator = Paginator(itens, 10)
    page = request.GET.get("page")
    itens_page = paginator.get_page(page)

    filtro_form = ItemFornecedorForm()
    context = {
        "itens": itens_page,
        "fornecedor": fornecedor,
        "tipo": tipo,
        "form": filtro_form,
        "colunas": ITEM_FORNECEDOR_LIST_COLUMNS,
        "total_itens": total_itens,
        "itens_produto": itens_produto,
        "itens_servico": itens_servico,
        "itens_treinamento": itens_treinamento,
        "fornecedores_unicos": fornecedores_unicos,
    }
    return render(request, "item_fornecedor/list.html", context)


# Item de Fornecedor - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def itemfornecedor_detail(request, pk):
    item = get_object_or_404(ItemFornecedor, pk=pk)
    return render(request, "item_fornecedor/detail.html", {"item": item})


# Item de Fornecedor - Criar
@group_required("Admin", "Gerente")
def itemfornecedor_create(request):
    if request.method == "POST":
        form = ItemFornecedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("item_fornecedor_list")
    else:
        form = ItemFornecedorForm()
    return render(request, "item_fornecedor/form.html", {"form": form})


# Item de Fornecedor - Editar
@group_required("Admin", "Gerente")
def itemfornecedor_update(request, pk):
    item = get_object_or_404(ItemFornecedor, pk=pk)
    if request.method == "POST":
        form = ItemFornecedorForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect("item_fornecedor_list")
    else:
        form = ItemFornecedorForm(instance=item)
    return render(request, "item_fornecedor/form.html", {"form": form})


# Item de Fornecedor - Deletar
@group_required("Admin", "Gerente")
def itemfornecedor_delete(request, pk):
    item = get_object_or_404(ItemFornecedor, pk=pk)
    if request.method == "POST":
        item.delete()
        return redirect("item_fornecedor_list")
    return render(request, "item_fornecedor/confirm_delete.html", {"object": item})


# Item de Fornecedor - Exportação CSV
@group_required("Admin", "Gerente")
def export_item_fornecedor_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="itens_fornecedor.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "Fornecedor",
            "Outro Fornecedor",
            "Tipo",
            "SKU",
            "Descrição",
            "Unidade",
            "Valor Unitário",
            "Observações",
        ]
    )

    itens = ItemFornecedor.objects.all()
    for item in itens:
        writer.writerow(
            [
                item.id,
                item.fornecedor,
                item.outro_fornecedor if item.fornecedor == "Outro Fornecedor" else "",
                item.tipo,
                item.sku,
                item.descricao,
                item.unidade,
                f"{item.valor_unitario:.2f}".replace(".", ","),
                item.observacoes or "",
            ]
        )

    return response


# Todas as colunas disponíveis
colunas = [
    {"id": "numero", "label": "Número"},
    {"id": "cliente", "label": "Cliente"},
    {"id": "contrato", "label": "Contrato"},
    {"id": "itemcontrato", "label": "Item Contrato"},
    {"id": "itemfornecedor", "label": "Item Fornecedor"},
    {"id": "unidade", "label": "Unidade"},
    {"id": "quantidade", "label": "Quantidade"},
    {"id": "vigencia_produto", "label": "Vigência do Produto (meses)"},
    {"id": "valor_unitario", "label": "Valor Unitário"},
    {"id": "valor_total", "label": "Valor Total"},
    {"id": "data_ativacao", "label": "Data Ativação"},
    {"id": "data_faturamento", "label": "Data Faturamento"},
    {"id": "observacoes", "label": "Observações"},
    {"id": "status", "label": "Status"},
]


# Ordem de Fornecimento - Listar
@group_required("Admin", "Gerente", "Leitor")
def ordemfornecimento_list(request):
    ordens = OrdemFornecimento.objects.all()

    # Filtros
    numero = request.GET.get("numero")
    cliente = request.GET.get("cliente")
    status = request.GET.get("status")

    if numero:
        ordens = ordens.filter(numero_of__icontains=numero)
    if cliente:
        ordens = ordens.filter(cliente__nome_fantasia__icontains=cliente)
    if status:
        ordens = ordens.filter(status=status)

    # Estatísticas
    total_ordens = OrdemFornecimento.objects.count()
    ordens_abertas = OrdemFornecimento.objects.filter(status="aberta").count()
    ordens_execucao = OrdemFornecimento.objects.filter(status="execucao").count()
    ordens_finalizadas = OrdemFornecimento.objects.filter(status="finalizada").count()
    ordens_faturadas = OrdemFornecimento.objects.filter(status="faturada").count()
    
    # Valor total das ordens faturadas
    valor_total_faturado = OrdemFornecimento.objects.filter(status="faturada").aggregate(
        total=Sum('valor_total')
    )['total'] or 0

    # Paginação
    paginator = Paginator(ordens, 10)
    page = request.GET.get("page")
    ordens_page = paginator.get_page(page)

    context = {
        "ordens": ordens_page,
        "colunas": colunas,
        "numero": numero,
        "cliente": cliente,
        "status": status,
        "total_ordens": total_ordens,
        "ordens_abertas": ordens_abertas,
        "ordens_execucao": ordens_execucao,
        "ordens_finalizadas": ordens_finalizadas,
        "ordens_faturadas": ordens_faturadas,
        "valor_total_faturado": valor_total_faturado,
    }
    return render(request, "ordem_fornecimento/list.html", context)


# Ordem de Fornecimento - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def ordemfornecimento_detail(request, pk):
    ordem = get_object_or_404(OrdemFornecimento, pk=pk)
    return render(request, "ordem_fornecimento/detail.html", {"ordem": ordem})


# Ordem de Fornecimento - Criar
@group_required("Admin", "Gerente")
def ordemfornecimento_create(request):
    if request.method == "POST":
        form = OrdemFornecimentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ordem de Fornecimento criada com sucesso.")
            return redirect("ordem_fornecimento_list")
    else:
        form = OrdemFornecimentoForm()
    return render(request, "ordem_fornecimento/form.html", {"form": form})


# Ordem de Fornecimento - Editar
@group_required("Admin", "Gerente")
def ordemfornecimento_update(request, pk):
    ordem = get_object_or_404(OrdemFornecimento, pk=pk)
    if request.method == "POST":
        form = OrdemFornecimentoForm(request.POST, instance=ordem)
        if form.is_valid():
            form.save()
            messages.success(request, "Ordem de Fornecimento atualizada com sucesso.")
            return redirect("ordem_fornecimento_list")
    else:
        form = OrdemFornecimentoForm(instance=ordem)
    return render(request, "ordem_fornecimento/form.html", {"form": form})


# Ordem de Fornecimento - Deletar
@group_required("Admin", "Gerente")
def ordemfornecimento_delete(request, pk):
    ordem = get_object_or_404(OrdemFornecimento, pk=pk)
    if request.method == "POST":
        ordem.delete()
        messages.success(request, "Ordem de Fornecimento excluída com sucesso.")
        return redirect("ordem_fornecimento_list")
    return render(request, "ordem_fornecimento/confirm_delete.html", {"ordem": ordem})


# Ordem de Fornecimento - Exportação CSV
@group_required("Admin", "Gerente", "Leitor")
def export_ordemfornecimento_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="ordens_fornecimento.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "Número OF",
            "Número OF Cliente",
            "Cliente",
            "Contrato",
            "Item do Contrato",
            "Item do Fornecedor",
            "Unidade",
            "Quantidade",
            "Vigência Produto",
            "Valor Unitário",
            "Valor Total",
            "Status",
            "Data Ativação",
            "Data Faturamento",
            "Observações",
            "Criado em",
            "Atualizado em",
        ]
    )

    ordens = OrdemFornecimento.objects.all()

    for ordem in ordens:
        writer.writerow(
            [
                ordem.id,
                smart_str(ordem.numero_of),
                smart_str(ordem.numero_of_cliente),
                smart_str(ordem.cliente.nome_fantasia),
                smart_str(ordem.contrato.numero),
                smart_str(ordem.item_contrato.descricao),
                smart_str(', '.join([item.item_fornecedor.descricao for item in ordem.itens_fornecedor.all()]) or 'N/A'),
                smart_str(ordem.unidade),
                ordem.quantidade,
                ordem.vigencia_produto,
                ordem.valor_unitario,
                ordem.valor_total,
                ordem.get_status_display(),
                ordem.data_ativacao.strftime("%d/%m/%Y") if ordem.data_ativacao else "",
                (
                    ordem.data_faturamento.strftime("%d/%m/%Y")
                    if ordem.data_faturamento
                    else ""
                ),
                smart_str(ordem.observacoes),
                (
                    ordem.criado_em.strftime("%d/%m/%Y %H:%M:%S")
                    if ordem.criado_em
                    else ""
                ),
                (
                    ordem.atualizado_em.strftime("%d/%m/%Y %H:%M:%S")
                    if ordem.atualizado_em
                    else ""
                ),
            ]
        )

    return response


# Ordem de Serviço - Listar com filtros e paginação
@group_required("Admin", "Gerente", "Técnico")
def ordemservico_list(request):
    ordens = OrdemServico.objects.all()

    numero = request.GET.get("numero")
    cliente = request.GET.get("cliente")
    status = request.GET.get("status")

    if numero:
        ordens = ordens.filter(numero_os__icontains=numero)

    if cliente:
        ordens = ordens.filter(cliente__nome_fantasia__icontains=cliente)

    if status:
        ordens = ordens.filter(status=status)

    paginator = Paginator(ordens.order_by("-data_inicio"), 10)
    page = request.GET.get("page")
    ordens_page = paginator.get_page(page)

    colunas = [
        {"id": "numero", "label": "Número OS"},
        {"id": "numero_cliente", "label": "Número OS Cliente"},
        {"id": "cliente", "label": "Cliente"},
        {"id": "contrato", "label": "Contrato"},
        {"id": "itemcontrato", "label": "Item do Contrato"},
        {"id": "itemfornecedor", "label": "Item do Fornecedor"},
        {"id": "unidade", "label": "Unidade"},
        {"id": "quantidade", "label": "Quantidade"},
        {"id": "valor_unitario", "label": "Valor Unitário"},
        {"id": "valor_total", "label": "Valor Total"},
        {"id": "gerente_projetos", "label": "Gerente de Projetos"},
        {"id": "consultor_tecnico", "label": "Consultor Técnico"},
        {"id": "data_inicio", "label": "Data Início"},
        {"id": "hora_inicio", "label": "Hora Início"},
        {"id": "data_termino", "label": "Data Término"},
        {"id": "hora_termino", "label": "Hora Término"},
        {"id": "status", "label": "Status"},
        {"id": "data_emissao_trd", "label": "Data Emissão TRD"},
        {"id": "data_faturamento", "label": "Data Faturamento"},
        {"id": "horas_consultor", "label": "Horas Consultor"},
        {"id": "horas_gerente", "label": "Horas Gerente"},
        {"id": "horas_totais", "label": "Horas Totais"},
    ]

    context = {
        "ordens": ordens_page,
        "colunas": colunas,
        "filtro_numero": numero or "",
        "filtro_cliente": cliente or "",
        "filtro_status": status or "",
    }
    return render(request, "ordem_servico/list.html", context)


# Ordem de Serviço - Detalhar (com tarefas e planejado x realizado)
@group_required("Admin", "Gerente", "Técnico")
def ordemservico_detail(request, pk):
    ordem = get_object_or_404(
        OrdemServico.objects.select_related(
            'cliente', 'contrato', 'item_contrato', 
            'item_fornecedor_consultor', 'item_fornecedor_gerente', 'projeto'
        ),
        pk=pk
    )
    
    # Tarefas vinculadas à OS através do projeto
    # Buscar tarefas do projeto vinculado à OS
    tarefas = Tarefa.objects.none()
    if hasattr(ordem, 'projeto') and ordem.projeto:
        tarefas = Tarefa.objects.filter(
            projeto=ordem.projeto
        ).select_related('responsavel', 'projeto').prefetch_related('lancamentos_horas__colaborador').order_by('-criado_em')
    
    # Estatísticas de tarefas
    total_tarefas = tarefas.count()
    tarefas_concluidas = tarefas.filter(status='concluida').count()
    tarefas_em_andamento = tarefas.filter(status='em_andamento').count()
    tarefas_pendentes = tarefas.filter(status='pendente').count()
    
    # Calcular horas planejadas e realizadas
    ordem.calcular_horas_tarefas()
    
    # Estatísticas de horas
    horas_planejadas_total = ordem.horas_planejadas or Decimal('0.00')
    horas_realizadas_total = ordem.horas_realizadas or Decimal('0.00')
    diferenca_horas = horas_realizadas_total - horas_planejadas_total
    percentual_execucao = ordem.percentual_execucao if ordem.horas_planejadas > 0 else Decimal('0.00')
    
    context = {
        "ordem": ordem,
        "tarefas": tarefas,
        "total_tarefas": total_tarefas,
        "tarefas_concluidas": tarefas_concluidas,
        "tarefas_em_andamento": tarefas_em_andamento,
        "tarefas_pendentes": tarefas_pendentes,
        "horas_planejadas_total": horas_planejadas_total,
        "horas_realizadas_total": horas_realizadas_total,
        "diferenca_horas": diferenca_horas,
        "percentual_execucao": percentual_execucao,
    }
    
    return render(request, "ordem_servico/detail.html", context)


# Ordem de Serviço - Criar
@group_required("Admin", "Gerente")
def ordemservico_create(request):
    contrato_id = request.GET.get("contrato")
    if request.method == "POST":
        form = OrdemServicoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("ordem_servico_list")
    else:
        form = OrdemServicoForm(initial={"contrato": contrato_id})
    return render(request, "ordem_servico/form.html", {"form": form})


# Ordem de Serviço - Editar
@group_required("Admin", "Gerente")
def ordemservico_update(request, pk):
    ordem = get_object_or_404(OrdemServico, pk=pk)

    if request.method == "POST":
        form = OrdemServicoForm(request.POST, instance=ordem)
        if form.is_valid():
            form.save()
            messages.success(request, "OS atualizada com sucesso!")
            return redirect("ordem_servico_detail", pk=pk)
        else:
            messages.error(request, "Erro ao atualizar a OS. Verifique os campos.")
            # Exibir erros específicos
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = OrdemServicoForm(instance=ordem)

    return render(request, "ordem_servico/form.html", {"form": form, "ordem": ordem})


# Ordem de Serviço - Deletar
@group_required("Admin", "Gerente")
def ordemservico_delete(request, pk):
    ordem = get_object_or_404(OrdemServico, pk=pk)
    if request.method == "POST":
        ordem.delete()
        return redirect("ordem_servico_list")
    return render(request, "ordem_servico/confirm_delete.html", {"object": ordem})


# Ordem de Serviço - Exportação CSV
@group_required("Admin", "Gerente", "Técnico")
def export_ordemservico_csv(request):
    ordens = OrdemServico.objects.all()

    contrato = request.GET.get("contrato")
    cliente = request.GET.get("cliente")
    status = request.GET.get("status")

    if contrato:
        ordens = ordens.filter(contrato__numero_contrato__icontains=contrato)

    if cliente:
        ordens = ordens.filter(cliente__nome_fantasia__icontains=cliente)

    if status:
        ordens = ordens.filter(status=status)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="ordens_servico.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "Número OS",
            "Número OS Cliente",
            "Cliente",
            "Contrato",
            "Item Contrato",
            "Item Fornecedor",
            "Quantidade",
            "Valor Unitário (R$)",
            "Valor Total (R$)",
            "Status",
            "Data Início",
            "Hora Início",
            "Data Término",
            "Hora Término",
            "Data Emissão TRD",
            "Data Faturamento",
        ]
    )

    for ordem in ordens:
        writer.writerow(
            [
                ordem.id,
                ordem.numero_os,
                ordem.numero_os_cliente,
                ordem.cliente.nome_fantasia,
                ordem.contrato.numero_contrato,
                ordem.item_contrato.descricao,
                ', '.join([item.item_fornecedor.descricao for item in ordem.itens_fornecedor.all()]) or 'N/A',
                ordem.quantidade,
                float(ordem.valor_unitario),
                float(ordem.valor_total),
                ordem.get_status_display(),
                ordem.data_inicio,
                ordem.hora_inicio,
                ordem.data_termino or "",
                ordem.hora_termino or "",
                ordem.data_emissao_trd or "",
                ordem.data_faturamento or "",
            ]
        )

    return response


@require_http_methods(["GET", "POST"])
def import_export_view(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        try:
            excel_file = pd.ExcelFile(file)

            # Clientes
            Cliente.objects.all().delete()
            df_clientes = pd.read_excel(excel_file, sheet_name="Clientes")
            for _, row in df_clientes.iterrows():
                Cliente.objects.create(**row.to_dict())

            # Contratos
            Contrato.objects.all().delete()
            df_contratos = pd.read_excel(excel_file, sheet_name="Contratos")
            for _, row in df_contratos.iterrows():
                cliente = Cliente.objects.get(id=row["cliente_id"])
                fornecedores = (
                    eval(row["fornecedores"])
                    if isinstance(row["fornecedores"], str)
                    else []
                )
                Contrato.objects.create(
                    cliente=cliente,
                    numero=row["numero"],
                    data_assinatura=row["data_assinatura"],
                    data_fim=row["data_fim"],
                    valor_total=row["valor_total"],
                    situacao=row["situacao"],
                    fornecedores=fornecedores,
                )

            # Itens de Contrato
            ItemContrato.objects.all().delete()
            df_itens_contrato = pd.read_excel(excel_file, sheet_name="ItensContrato")
            for _, row in df_itens_contrato.iterrows():
                contrato = Contrato.objects.get(id=row["contrato_id"])
                ItemContrato.objects.create(
                    contrato=contrato, **row.drop("contrato_id").to_dict()
                )

            # Itens de Fornecedor
            ItemFornecedor.objects.all().delete()
            df_itens_fornecedor = pd.read_excel(
                excel_file, sheet_name="ItensFornecedor"
            )
            for _, row in df_itens_fornecedor.iterrows():
                ItemFornecedor.objects.create(**row.to_dict())

            # Ordens de Serviço
            OrdemServico.objects.all().delete()
            df_os = pd.read_excel(excel_file, sheet_name="OrdensServico")
            for _, row in df_os.iterrows():
                OrdemServico.objects.create(**row.to_dict())

            # Ordens de Fornecimento
            OrdemFornecimento.objects.all().delete()
            df_of = pd.read_excel(excel_file, sheet_name="OrdensFornecimento")
            for _, row in df_of.iterrows():
                OrdemFornecimento.objects.create(**row.to_dict())

            messages.success(request, "Dados importados com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro na importação: {str(e)}")

        return redirect("import_export_view")

    return render(request, "import_export/import_export.html")


def export_excel_view(request):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine="openpyxl")

    models_and_sheets = [
        (Cliente, "Clientes"),
        (Contrato, "Contratos"),
        (ItemContrato, "ItensContrato"),
        (ItemFornecedor, "ItensFornecedor"),
        (OrdemFornecimento, "OrdensFornecimento"),
        (OrdemServico, "OrdensServico"),
    ]

    for model, sheet_name in models_and_sheets:
        queryset = model.objects.all().values()
        df = pd.DataFrame.from_records(queryset)
        df = clean_datetimes(df)
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        "attachment; filename=ControleContratos_Exportado.xlsx"
    )
    return response


def clean_datetimes(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].apply(
                lambda x: (
                    x.replace(tzinfo=None)
                    if isinstance(x, datetime) and is_aware(x)
                    else x
                )
            )
        elif df[col].dtype == "object":
            # Tenta converter strings para datetime e remover fuso se necessário
            try:
                df[col] = pd.to_datetime(df[col], errors="coerce")
                df[col] = df[col].apply(
                    lambda x: (
                        x.replace(tzinfo=None)
                        if isinstance(x, datetime) and is_aware(x)
                        else x
                    )
                )
            except Exception:
                pass
    return df


# Visualizar Logs de Importação e Exportação
@group_required("Admin", "Gerente")
def logs_import_export(request):
    logs = ImportExportLog.objects.order_by("-data")
    return render(request, "import_export/import_export_logs.html", {"logs": logs})


# ========== GESTÃO DE CLIENTES - NOVAS VIEWS ==========

# Colaborador - Listagem (Unificada com usuários do Django)
@group_required("Admin", "Gerente", "Leitor")
def colaborador_list(request):
    # Buscar todos os usuários do sistema (unificando visão)
    usuarios = User.objects.all().select_related('colaborador').order_by('username')
    
    nome = request.GET.get("nome")
    cargo = request.GET.get("cargo")
    ativo = request.GET.get("ativo")
    tem_colaborador = request.GET.get("tem_colaborador")
    
    # Filtrar por nome (no usuário ou no colaborador)
    if nome:
        usuarios = usuarios.filter(
            Q(username__icontains=nome) |
            Q(first_name__icontains=nome) |
            Q(last_name__icontains=nome) |
            Q(email__icontains=nome) |
            Q(colaborador__nome_completo__icontains=nome)
        )
    
    # Filtrar por cargo
    if cargo:
        usuarios = usuarios.filter(colaborador__cargo__icontains=cargo)
    
    # Filtrar por status ativo/inativo
    if ativo == "ativo":
        usuarios = usuarios.filter(colaborador__ativo=True)
    elif ativo == "inativo":
        usuarios = usuarios.filter(colaborador__ativo=False)
    
    # Filtrar por ter ou não colaborador
    if tem_colaborador == "sim":
        usuarios = usuarios.filter(colaborador__isnull=False)
    elif tem_colaborador == "nao":
        usuarios = usuarios.filter(colaborador__isnull=True)
    
    # Estatísticas
    total_usuarios = User.objects.count()
    total_colaboradores = Colaborador.objects.count()
    usuarios_sem_colaborador = User.objects.filter(colaborador__isnull=True).count()
    colaboradores_ativos = Colaborador.objects.filter(ativo=True).count()
    colaboradores_inativos = Colaborador.objects.filter(ativo=False).count()
    
    paginator = Paginator(usuarios, 10)
    page = request.GET.get("page")
    usuarios_page = paginator.get_page(page)
    
    context = {
        "usuarios": usuarios_page,
        "total_usuarios": total_usuarios,
        "total_colaboradores": total_colaboradores,
        "usuarios_sem_colaborador": usuarios_sem_colaborador,
        "colaboradores_ativos": colaboradores_ativos,
        "colaboradores_inativos": colaboradores_inativos,
    }
    
    return render(request, "colaborador/list.html", context)


# Colaborador - Criar
@group_required("Admin", "Gerente")
def colaborador_create(request):
    from .forms import ColaboradorForm
    user_id = request.GET.get('user_id')
    
    if request.method == "POST":
        form = ColaboradorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Colaborador criado com sucesso!")
            return redirect("colaborador_list")
    else:
        form = ColaboradorForm()
        # Se foi passado user_id, pré-selecionar o usuário
        if user_id:
            try:
                user = User.objects.get(pk=user_id)
                form.fields['user'].initial = user
                form.fields['user'].queryset = User.objects.filter(pk=user_id)
                form.fields['email'].initial = user.email
                form.fields['nome_completo'].initial = user.get_full_name() or user.username
                form.fields['criar_usuario'].initial = False
            except User.DoesNotExist:
                pass
    
    return render(request, "colaborador/form.html", {"form": form})


# Colaborador - Editar
@group_required("Admin", "Gerente")
def colaborador_update(request, pk):
    from .forms import ColaboradorForm
    colaborador = get_object_or_404(Colaborador, pk=pk)
    if request.method == "POST":
        form = ColaboradorForm(request.POST, instance=colaborador)
        if form.is_valid():
            form.save()
            messages.success(request, "Colaborador atualizado com sucesso!")
            return redirect("colaborador_list")
    else:
        form = ColaboradorForm(instance=colaborador)
    return render(request, "colaborador/form.html", {"form": form})


# Colaborador - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def colaborador_detail(request, pk):
    colaborador = get_object_or_404(
        Colaborador.objects.select_related('user').prefetch_related('user__groups'),
        pk=pk
    )
    
    # Estatísticas do colaborador
    clientes_comercial = colaborador.clientes_comercial.count() if colaborador.cargo and "comercial" in colaborador.cargo.lower() else 0
    clientes_sucessos = colaborador.clientes_sucessos.count() if colaborador.cargo and "sucessos" in colaborador.cargo.lower() else 0
    contratos_gerenciados = colaborador.contratos_gerenciados.count() if colaborador.cargo and "gerente" in colaborador.cargo.lower() else 0
    tarefas = colaborador.tarefas.count()
    lancamentos_horas = colaborador.lancamentos_horas.count()
    
    context = {
        "colaborador": colaborador,
        "clientes_comercial": clientes_comercial,
        "clientes_sucessos": clientes_sucessos,
        "contratos_gerenciados": contratos_gerenciados,
        "tarefas": tarefas,
        "lancamentos_horas": lancamentos_horas,
    }
    
    return render(request, "colaborador/detail.html", context)


# Colaborador - Alterar Senha
@group_required("Admin", "Gerente")
def colaborador_password_change(request, pk):
    """Altera a senha do colaborador"""
    colaborador = get_object_or_404(
        Colaborador.objects.select_related('user'),
        pk=pk
    )
    
    if not colaborador.user:
        messages.error(request, "Este colaborador não possui um usuário associado.")
        return redirect("colaborador_detail", pk=colaborador.pk)
    
    from .forms import ColaboradorPasswordChangeForm
    
    if request.method == "POST":
        form = ColaboradorPasswordChangeForm(colaborador.user, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"Senha do colaborador {colaborador.nome_completo} alterada com sucesso!")
            return redirect("colaborador_detail", pk=colaborador.pk)
        else:
            messages.error(request, "Erro ao alterar a senha. Verifique os campos.")
    else:
        form = ColaboradorPasswordChangeForm(colaborador.user)
    
    context = {
        "form": form,
        "colaborador": colaborador,
    }
    
    return render(request, "colaborador/password_change.html", context)


# ========== GESTÃO DE GRUPOS ==========

# Grupo - Listar
@group_required("Admin")
def grupo_list(request):
    """Lista todos os grupos"""
    grupos = Group.objects.prefetch_related('permissions', 'user_set').all()
    
    search = request.GET.get("search")
    if search:
        grupos = grupos.filter(name__icontains=search)
    
    # Estatísticas
    total_grupos = Group.objects.count()
    total_permissoes = Permission.objects.count()
    total_usuarios = User.objects.filter(groups__isnull=False).distinct().count()
    
    paginator = Paginator(grupos, 20)
    page = request.GET.get("page")
    grupos_page = paginator.get_page(page)
    
    context = {
        "grupos": grupos_page,
        "search": search,
        "total_grupos": total_grupos,
        "total_permissoes": total_permissoes,
        "total_usuarios": total_usuarios,
    }
    return render(request, "grupo/list.html", context)


# Grupo - Criar
@group_required("Admin")
def grupo_create(request):
    """Cria um novo grupo"""
    from .forms import GrupoForm
    
    if request.method == "POST":
        form = GrupoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Grupo criado com sucesso!")
            return redirect("grupo_list")
        else:
            messages.error(request, "Erro ao criar o grupo. Verifique os campos.")
    else:
        form = GrupoForm()
    
    return render(request, "grupo/form.html", {"form": form})


# Grupo - Editar
@group_required("Admin")
def grupo_update(request, pk):
    """Edita um grupo"""
    from .forms import GrupoForm
    
    grupo = get_object_or_404(Group, pk=pk)
    
    if request.method == "POST":
        form = GrupoForm(request.POST, instance=grupo)
        if form.is_valid():
            form.save()
            messages.success(request, "Grupo atualizado com sucesso!")
            return redirect("grupo_detail", pk=grupo.pk)
        else:
            messages.error(request, "Erro ao atualizar o grupo. Verifique os campos.")
    else:
        form = GrupoForm(instance=grupo)
    
    return render(request, "grupo/form.html", {"form": form, "grupo": grupo})


# Grupo - Detalhar
@group_required("Admin", "Gerente")
def grupo_detail(request, pk):
    """Detalhes do grupo com permissões e usuários"""
    grupo = get_object_or_404(
        Group.objects.prefetch_related('permissions', 'user_set'),
        pk=pk
    )
    
    # Organizar permissões por app
    permissoes_por_app = {}
    for perm in grupo.permissions.all().select_related('content_type'):
        app_label = perm.content_type.app_label
        if app_label not in permissoes_por_app:
            permissoes_por_app[app_label] = []
        permissoes_por_app[app_label].append(perm)
    
    # Usuários do grupo
    usuarios = grupo.user_set.select_related('colaborador').all()
    
    context = {
        "grupo": grupo,
        "permissoes_por_app": permissoes_por_app,
        "usuarios": usuarios,
        "total_permissoes": grupo.permissions.count(),
        "total_usuarios": usuarios.count(),
    }
    
    return render(request, "grupo/detail.html", context)


# Grupo - Deletar
@group_required("Admin")
def grupo_delete(request, pk):
    """Exclui um grupo"""
    grupo = get_object_or_404(Group, pk=pk)
    
    if request.method == "POST":
        nome_grupo = grupo.name
        total_usuarios = grupo.user_set.count()
        grupo.delete()
        messages.success(
            request,
            f"Grupo '{nome_grupo}' excluído com sucesso! ({total_usuarios} usuário(s) foram removidos do grupo.)"
        )
        return redirect("grupo_list")
    
    context = {
        "grupo": grupo,
        "total_usuarios": grupo.user_set.count(),
    }
    return render(request, "grupo/confirm_delete.html", context)


# ========== GESTÃO DE CONTRATOS E OF - NOVAS VIEWS ==========

# SLA - Listagem
@group_required("Admin", "Gerente", "Leitor")
def sla_list(request):
    slas = SLA.objects.select_related('contrato').all()
    
    contrato_id = request.GET.get("contrato")
    tipo = request.GET.get("tipo")
    ativo = request.GET.get("ativo")
    
    if contrato_id:
        slas = slas.filter(contrato_id=contrato_id)
    if tipo:
        slas = slas.filter(tipo=tipo)
    if ativo == "ativo":
        slas = slas.filter(ativo=True)
    elif ativo == "inativo":
        slas = slas.filter(ativo=False)
    
    paginator = Paginator(slas, 10)
    page = request.GET.get("page")
    slas_page = paginator.get_page(page)
    
    return render(request, "sla/list.html", {"slas": slas_page})


# SLA - Criar
@group_required("Admin", "Gerente")
def sla_create(request):
    from .forms import SLAForm
    if request.method == "POST":
        form = SLAForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "SLA criado com sucesso!")
            return redirect("sla_list")
    else:
        form = SLAForm()
        # Preencher contrato se vier na URL
        contrato_id = request.GET.get("contrato")
        if contrato_id:
            form.fields['contrato'].initial = contrato_id
    return render(request, "sla/form.html", {"form": form})


# SLA - Editar
@group_required("Admin", "Gerente")
def sla_update(request, pk):
    from .forms import SLAForm
    sla = get_object_or_404(SLA, pk=pk)
    if request.method == "POST":
        form = SLAForm(request.POST, instance=sla)
        if form.is_valid():
            form.save()
            messages.success(request, "SLA atualizado com sucesso!")
            return redirect("sla_list")
    else:
        form = SLAForm(instance=sla)
    return render(request, "sla/form.html", {"form": form})


# SLA - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def sla_detail(request, pk):
    sla = get_object_or_404(SLA.objects.select_related('contrato'), pk=pk)
    return render(request, "sla/detail.html", {"sla": sla})


# Fila de Faturamento - Listagem
@group_required("Admin", "Gerente", "Leitor")
def fila_faturamento(request):
    """Fila de OS e OF pendentes de faturamento"""
    # OS finalizadas mas não faturadas
    os_pendentes = OrdemServico.objects.filter(
        status="finalizada"
    ).select_related('cliente', 'contrato', 'item_contrato').order_by('-data_emissao_trd')
    
    # OF finalizadas mas não faturadas
    of_pendentes = OrdemFornecimento.objects.filter(
        status="finalizada"
    ).select_related('cliente', 'contrato', 'item_contrato').prefetch_related('itens_fornecedor').order_by('-data_ativacao')
    
    # Filtros
    tipo = request.GET.get("tipo")
    cliente_id = request.GET.get("cliente")
    contrato_id = request.GET.get("contrato")
    
    if tipo == "os":
        of_pendentes = OrdemFornecimento.objects.none()
    elif tipo == "of":
        os_pendentes = OrdemServico.objects.none()
    
    if cliente_id:
        os_pendentes = os_pendentes.filter(cliente_id=cliente_id)
        of_pendentes = of_pendentes.filter(cliente_id=cliente_id)
    
    if contrato_id:
        os_pendentes = os_pendentes.filter(contrato_id=contrato_id)
        of_pendentes = of_pendentes.filter(contrato_id=contrato_id)
    
    # Estatísticas
    total_os_pendentes = os_pendentes.count()
    total_of_pendentes = of_pendentes.count()
    valor_total_os = sum(os.receita_prevista for os in os_pendentes if os.receita_prevista)
    valor_total_of = sum(of.valor_total for of in of_pendentes if of.valor_total)
    
    context = {
        "os_pendentes": os_pendentes[:50],  # Limitar a 50 para performance
        "of_pendentes": of_pendentes[:50],
        "total_os_pendentes": total_os_pendentes,
        "total_of_pendentes": total_of_pendentes,
        "valor_total_os": valor_total_os,
        "valor_total_of": valor_total_of,
        "valor_total_geral": valor_total_os + valor_total_of,
    }
    
    return render(request, "fila_faturamento/list.html", context)


# Fila de Faturamento - Marcar como Faturado (OS)
@group_required("Admin", "Gerente")
def marcar_os_faturada(request, pk):
    """Marca uma OS como faturada"""
    os_item = get_object_or_404(OrdemServico, pk=pk)
    if os_item.status == "finalizada":
        os_item.status = "faturada"
        if not os_item.data_faturamento:
            os_item.data_faturamento = timezone.now().date()
        os_item.save()
        messages.success(request, f"OS {os_item.numero_os} marcada como faturada!")
    else:
        messages.error(request, "Apenas OS finalizadas podem ser marcadas como faturadas.")
    return redirect("fila_faturamento")


# Fila de Faturamento - Marcar como Faturado (OF)
@group_required("Admin", "Gerente")
def marcar_of_faturada(request, pk):
    """Marca uma OF como faturada"""
    of_item = get_object_or_404(OrdemFornecimento, pk=pk)
    if of_item.status == "finalizada":
        of_item.status = "faturada"
        if not of_item.data_faturamento:
            of_item.data_faturamento = timezone.now().date()
        of_item.save()
        messages.success(request, f"OF {of_item.numero_of} marcada como faturada!")
    else:
        messages.error(request, "Apenas OF finalizadas podem ser marcadas como faturadas.")
    return redirect("fila_faturamento")


# ========== GESTÃO DE OS COM TAREFAS ==========

# Tarefa - Criar vinculada a OS
@group_required("Admin", "Gerente")
def tarefa_os_create(request, os_id):
    """Cria uma tarefa vinculada a uma OS"""
    from .forms import TarefaForm
    ordem_servico = get_object_or_404(OrdemServico, pk=os_id)
    
    if request.method == "POST":
        form = TarefaForm(request.POST)
        if form.is_valid():
            tarefa = form.save(commit=False)
            tarefa.ordem_servico = ordem_servico
            tarefa.save()
            messages.success(request, "Tarefa criada com sucesso!")
            return redirect("ordem_servico_detail", pk=os_id)
    else:
        form = TarefaForm()
        # Preencher OS se vier na URL
        form.fields['ordem_servico'].initial = ordem_servico
        form.fields['ordem_servico'].widget = forms.HiddenInput()
    
    return render(request, "tarefa/form_os.html", {"form": form, "ordem_servico": ordem_servico})


# Tarefa - Editar vinculada a OS
@group_required("Admin", "Gerente")
def tarefa_os_update(request, os_id, tarefa_id):
    """Edita uma tarefa vinculada a uma OS"""
    from .forms import TarefaForm
    ordem_servico = get_object_or_404(OrdemServico, pk=os_id)
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id, ordem_servico=ordem_servico)
    
    if request.method == "POST":
        form = TarefaForm(request.POST, instance=tarefa)
        if form.is_valid():
            form.save()
            messages.success(request, "Tarefa atualizada com sucesso!")
            return redirect("ordem_servico_detail", pk=os_id)
    else:
        form = TarefaForm(instance=tarefa)
        form.fields['ordem_servico'].widget = forms.HiddenInput()
    
    return render(request, "tarefa/form_os.html", {"form": form, "ordem_servico": ordem_servico, "tarefa": tarefa})


# Lançamento de Hora - Criar para tarefa
@group_required("Admin", "Gerente", "Tecnico")
def lancamento_hora_create(request, tarefa_id):
    """Cria um lançamento de hora para uma tarefa"""
    from .forms import LancamentoHoraForm
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id)
    
    if request.method == "POST":
        form = LancamentoHoraForm(request.POST)
        if form.is_valid():
            lancamento = form.save(commit=False)
            lancamento.tarefa = tarefa
            lancamento.save()
            messages.success(request, "Lançamento de hora criado com sucesso!")
            if tarefa.ordem_servico:
                return redirect("ordem_servico_detail", pk=tarefa.ordem_servico.pk)
            return redirect("tarefa_list")
    else:
        form = LancamentoHoraForm()
        form.fields['tarefa'].initial = tarefa
        form.fields['tarefa'].widget = forms.HiddenInput()
    
    return render(request, "lancamento_hora/form.html", {"form": form, "tarefa": tarefa})


# Lançamento de Hora - Editar
@group_required("Admin", "Gerente", "Tecnico")
def lancamento_hora_update(request, lancamento_id):
    """Edita um lançamento de hora"""
    from .forms import LancamentoHoraForm
    lancamento = get_object_or_404(LancamentoHora, pk=lancamento_id)
    tarefa = lancamento.tarefa
    
    if request.method == "POST":
        form = LancamentoHoraForm(request.POST, instance=lancamento)
        if form.is_valid():
            form.save()
            messages.success(request, "Lançamento de hora atualizado com sucesso!")
            if tarefa.ordem_servico:
                return redirect("ordem_servico_detail", pk=tarefa.ordem_servico.pk)
            return redirect("tarefa_list")
    else:
        form = LancamentoHoraForm(instance=lancamento)
        form.fields['tarefa'].widget = forms.HiddenInput()
    
    return render(request, "lancamento_hora/form.html", {"form": form, "tarefa": tarefa})


# Lançamento de Hora - Deletar
@group_required("Admin", "Gerente")
def lancamento_hora_delete(request, lancamento_id):
    """Deleta um lançamento de hora"""
    lancamento = get_object_or_404(LancamentoHora, pk=lancamento_id)
    tarefa = lancamento.tarefa
    
    if request.method == "POST":
        lancamento.delete()
        messages.success(request, "Lançamento de hora excluído com sucesso!")
        if tarefa.ordem_servico:
            return redirect("ordem_servico_detail", pk=tarefa.ordem_servico.pk)
        return redirect("tarefa_list")
    
    return render(request, "lancamento_hora/confirm_delete.html", {"lancamento": lancamento})


# ========== GESTÃO ÁGIL DE PROJETOS ==========

# Projeto - Listar
@group_required("Admin", "Gerente", "Leitor")
def projeto_list(request):
    """Lista todos os projetos"""
    projetos = Projeto.objects.select_related('contrato', 'contrato__cliente', 'gerente_projeto').all()
    
    # Filtros
    contrato_id = request.GET.get("contrato")
    status = request.GET.get("status")
    search = request.GET.get("search")
    
    if contrato_id:
        projetos = projetos.filter(contrato_id=contrato_id)
    if status:
        projetos = projetos.filter(status=status)
    if search:
        projetos = projetos.filter(
            Q(nome__icontains=search) |
            Q(descricao__icontains=search) |
            Q(contrato__numero_contrato__icontains=search) |
            Q(contrato__cliente__nome_razao_social__icontains=search)
        )
    
    # Paginação
    paginator = Paginator(projetos, 20)
    page = request.GET.get("page")
    projetos = paginator.get_page(page)
    
    # Contratos para filtro
    contratos = Contrato.objects.all().order_by("numero_contrato")
    
    context = {
        "projetos": projetos,
        "contratos": contratos,
        "status_choices": Projeto.STATUS_CHOICES,
    }
    
    return render(request, "projeto/list.html", context)


# Projeto - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def projeto_detail(request, pk):
    """Detalhes do projeto com sprints e plano de trabalho"""
    projeto = get_object_or_404(
        Projeto.objects.select_related('contrato', 'gerente_projeto'),
        pk=pk
    )
    
    # Tarefas do projeto (agora todas as tarefas têm projeto obrigatório)
    tarefas_projeto = Tarefa.objects.filter(
        projeto=projeto
    ).select_related('responsavel', 'sprint', 'projeto').order_by('-prioridade', '-criado_em')
    
    # Todas as tarefas do projeto (para aba Tarefas)
    todas_tarefas = Tarefa.objects.filter(
        projeto=projeto
    ).select_related('responsavel', 'backlog', 'sprint', 'projeto').order_by('-criado_em')
    
    # Sprints
    sprints = projeto.sprints.all().order_by('-data_inicio')
    
    # Dados para Canvas (Kanban)
    # Tarefas do backlog do projeto (sem sprint)
    tarefas_backlog = Tarefa.objects.filter(
        projeto=projeto,
        sprint__isnull=True
    ).select_related('responsavel').order_by('-prioridade', '-criado_em')
    
    # Sprints do projeto ordenadas por data (para canvas)
    sprints_canvas = projeto.sprints.all().order_by('data_inicio', 'data_fim')
    
    # Para cada sprint, buscar suas tarefas (para canvas)
    sprints_com_tarefas = []
    for sprint in sprints_canvas:
        tarefas_sprint = Tarefa.objects.filter(
            projeto=projeto,
            sprint=sprint
        ).select_related('responsavel').order_by('ordem_sprint', '-criado_em')
        
        sprints_com_tarefas.append({
            'sprint': sprint,
            'tarefas': tarefas_sprint,
            'total_tarefas': tarefas_sprint.count(),
            'tarefas_nao_iniciadas': tarefas_sprint.filter(status_sprint='nao_iniciada').count(),
            'tarefas_em_execucao': tarefas_sprint.filter(status_sprint='em_execucao').count(),
            'tarefas_finalizadas': tarefas_sprint.filter(status_sprint='finalizada').count(),
        })
    
    # Plano de trabalho do projeto
    plano_trabalho = getattr(projeto, 'plano_trabalho', None)
    
    # Estatísticas
    total_tarefas_projeto = tarefas_projeto.count()
    total_sprints = sprints.count()
    sprints_pendentes = sprints.filter(status="aberta").count()
    sprints_execucao = sprints.filter(status="execucao").count()
    sprints_finalizadas = sprints.filter(status="finalizada").count()
    sprints_faturadas = sprints.filter(status="faturada").count()
    total_tarefas_backlog = tarefas_backlog.count()
    
    # Calcular dias restantes para o vencimento do contrato
    from datetime import date
    dias_restantes_contrato = None
    meses_restantes = None
    if projeto.data_fim_prevista:
        hoje = date.today()
        diferenca = projeto.data_fim_prevista - hoje
        dias_restantes_contrato = diferenca.days
        meses_restantes = dias_restantes_contrato / 30  # Aproximação
    
    context = {
        "projeto": projeto,
        "tarefas_projeto": tarefas_projeto,
        "todas_tarefas": todas_tarefas,
        "sprints": sprints,
        "plano_trabalho": plano_trabalho,
        "total_tarefas_projeto": total_tarefas_projeto,
        "total_sprints": total_sprints,
        "sprints_pendentes": sprints_pendentes,
        "sprints_execucao": sprints_execucao,
        "sprints_finalizadas": sprints_finalizadas,
        "sprints_faturadas": sprints_faturadas,
        "dias_restantes_contrato": dias_restantes_contrato,
        "meses_restantes": meses_restantes,
        # Dados para Canvas
        "tarefas_backlog": tarefas_backlog,
        "sprints_com_tarefas": sprints_com_tarefas,
        "total_tarefas_backlog": total_tarefas_backlog,
    }
    
    return render(request, "projeto/detail.html", context)


# Projeto - Canvas (Visualização Kanban)
@login_required
def projeto_canvas(request, pk):
    """Visualização de canvas (Kanban) do projeto com backlog e sprints"""
    projeto = get_object_or_404(
        Projeto.objects.select_related('contrato', 'gerente_projeto'),
        pk=pk
    )
    
    # Tarefas do backlog do projeto (sem sprint)
    tarefas_backlog = Tarefa.objects.filter(
        projeto=projeto,
        sprint__isnull=True
    ).select_related('responsavel').order_by('-prioridade', '-criado_em')
    
    # Sprints do projeto ordenadas por data
    sprints = projeto.sprints.all().order_by('data_inicio', 'data_fim')
    
    # Para cada sprint, buscar suas tarefas
    sprints_com_tarefas = []
    for sprint in sprints:
        tarefas_sprint = Tarefa.objects.filter(
            projeto=projeto,
            sprint=sprint
        ).select_related('responsavel').order_by('-prioridade', '-criado_em')
        
        sprints_com_tarefas.append({
            'sprint': sprint,
            'tarefas': tarefas_sprint,
            'total_tarefas': tarefas_sprint.count(),
            'tarefas_nao_iniciadas': tarefas_sprint.filter(status_sprint='nao_iniciada').count(),
            'tarefas_em_execucao': tarefas_sprint.filter(status_sprint='em_execucao').count(),
            'tarefas_finalizadas': tarefas_sprint.filter(status_sprint='finalizada').count(),
        })
    
    # Estatísticas
    total_tarefas_backlog = tarefas_backlog.count()
    total_sprints = sprints.count()
    total_tarefas_projeto = Tarefa.objects.filter(projeto=projeto).count()
    
    context = {
        "projeto": projeto,
        "tarefas_backlog": tarefas_backlog,
        "sprints_com_tarefas": sprints_com_tarefas,
        "total_tarefas_backlog": total_tarefas_backlog,
        "total_sprints": total_sprints,
        "total_tarefas_projeto": total_tarefas_projeto,
    }
    
    return render(request, "projeto/canvas.html", context)


# Projeto - Criar
@group_required("Admin", "Gerente")
def projeto_create(request):
    """Cria um novo projeto (1 projeto por contrato)"""
    from .forms import ProjetoForm
    contrato_id = request.GET.get("contrato")
    
    if request.method == "POST":
        form = ProjetoForm(request.POST)
        if form.is_valid():
            contrato = form.cleaned_data['contrato']
            # Verificar se já existe projeto para este contrato
            if Projeto.objects.filter(contrato=contrato).exists():
                messages.error(request, f"Já existe um projeto para o contrato {contrato.numero_contrato}!")
                return render(request, "projeto/form.html", {"form": form})
            
            projeto = form.save(commit=False)
            # Sincronizar datas com o contrato
            if contrato.data_assinatura:
                projeto.data_inicio = contrato.data_assinatura
            if contrato.data_fim:
                projeto.data_fim_prevista = contrato.data_fim
            projeto.save()
            
            # Não criar backlog automaticamente - backlogs são criados manualmente ou via conversão
            messages.success(request, "Projeto criado com sucesso!")
            return redirect("projeto_detail", pk=projeto.pk)
    else:
        form = ProjetoForm()
        if contrato_id:
            form.fields['contrato'].initial = contrato_id
    
    return render(request, "projeto/form.html", {"form": form})


# Projeto - Editar
@group_required("Admin", "Gerente")
def projeto_update(request, pk):
    """Edita um projeto"""
    from .forms import ProjetoForm
    projeto = get_object_or_404(Projeto, pk=pk)
    
    if request.method == "POST":
        form = ProjetoForm(request.POST, instance=projeto)
        if form.is_valid():
            form.save()
            messages.success(request, "Projeto atualizado com sucesso!")
            return redirect("projeto_detail", pk=projeto.pk)
    else:
        form = ProjetoForm(instance=projeto)
    
    return render(request, "projeto/form.html", {"form": form, "projeto": projeto})


# Projeto - Deletar
@group_required("Admin", "Gerente")
def projeto_delete(request, pk):
    """Exclui um projeto e todas as sprints e tarefas vinculadas"""
    projeto = get_object_or_404(Projeto, pk=pk)
    
    if request.method == "POST":
        # Contar itens que serão excluídos
        total_sprints = projeto.sprints.count()
        total_tarefas = 0
        if hasattr(projeto, 'backlog'):
            total_tarefas = projeto.backlog.tarefas.count()
        
        # Excluir projeto (cascade excluirá sprints, backlog e tarefas)
        projeto.delete()
        
        messages.success(request, f"Projeto excluído com sucesso! ({total_sprints} sprint(s) e {total_tarefas} tarefa(s) também foram excluídos.)")
        return redirect("projeto_list")
    
    # Contar itens para exibição na confirmação
    total_sprints = projeto.sprints.count()
    total_tarefas = 0
    if hasattr(projeto, 'backlog'):
        total_tarefas = projeto.backlog.tarefas.count()
    
    context = {
        "projeto": projeto,
        "total_sprints": total_sprints,
        "total_tarefas": total_tarefas,
    }
    
    return render(request, "projeto/confirm_delete.html", context)


# Backlog - Criar via AJAX
@group_required("Admin", "Gerente")
def backlog_create_ajax(request, contrato_id):
    """Cria um novo backlog via AJAX"""
    from django.http import JsonResponse
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descricao = request.POST.get('descricao', '')
        prioridade = request.POST.get('prioridade', 'media')
        
        if not titulo:
            return JsonResponse({'success': False, 'error': 'Título é obrigatório.'}, status=400)
        
        try:
            backlog = Backlog.objects.create(
                contrato=contrato,
                titulo=titulo,
                descricao=descricao,
                prioridade=prioridade,
                status='pendente'
            )
            # Retornar dados do backlog criado para atualizar a interface
            return JsonResponse({
                'success': True,
                'message': 'Backlog criado com sucesso!',
                'backlog': {
                    'id': backlog.id,
                    'titulo': backlog.titulo or 'Backlog sem título',
                    'descricao': backlog.descricao or '',
                    'prioridade': backlog.prioridade,
                    'prioridade_display': backlog.get_prioridade_display(),
                    'contrato_id': contrato.id,
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método não permitido.'}, status=405)


# Backlog - Excluir via AJAX
@group_required("Admin", "Gerente")
def backlog_delete_ajax(request, backlog_id):
    """Exclui um backlog via AJAX"""
    from django.http import JsonResponse
    backlog = get_object_or_404(Backlog, pk=backlog_id)
    
    if request.method == 'POST':
        try:
            backlog_id = backlog.id
            backlog.delete()
            return JsonResponse({
                'success': True,
                'message': 'Backlog excluído com sucesso!',
                'backlog_id': backlog_id
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método não permitido.'}, status=405)


# Backlog - Converter em Projeto via AJAX
@group_required("Admin", "Gerente")
def backlog_converter_projeto_ajax(request, backlog_id):
    """Converte um backlog em projeto via AJAX"""
    from django.http import JsonResponse
    backlog = get_object_or_404(Backlog, pk=backlog_id)
    
    if request.method == 'POST':
        nome_projeto = request.POST.get('nome_projeto')
        descricao_projeto = request.POST.get('descricao_projeto', '')
        gerente_id = request.POST.get('gerente_projeto')
        
        if not nome_projeto:
            return JsonResponse({'success': False, 'error': 'Nome do projeto é obrigatório.'}, status=400)
        
        try:
            gerente = None
            if gerente_id:
                from .models import Colaborador
                gerente = get_object_or_404(Colaborador, pk=gerente_id)
            
            projeto = backlog.converter_para_projeto(
                nome_projeto=nome_projeto,
                descricao_projeto=descricao_projeto,
                gerente_projeto=gerente  # Colaborador ou None
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Backlog convertido em projeto "{projeto.nome}" com sucesso!',
                'projeto_id': projeto.id,
                'redirect_url': reverse('projeto_detail', kwargs={'pk': projeto.pk})
            })
        except Exception as e:
            import traceback
            print(f"Erro ao converter backlog: {traceback.format_exc()}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método não permitido.'}, status=405)


# Backlog - Converter em Projeto (view antiga mantida para compatibilidade)
@group_required("Admin", "Gerente")
def backlog_converter_projeto(request, backlog_id):
    """Converte um backlog em projeto"""
    backlog = get_object_or_404(Backlog, pk=backlog_id)
    
    if request.method == 'POST':
        nome_projeto = request.POST.get('nome_projeto')
        descricao_projeto = request.POST.get('descricao_projeto', '')
        gerente_id = request.POST.get('gerente_projeto')
        
        if not nome_projeto:
            messages.error(request, 'Nome do projeto é obrigatório.')
            return redirect('gestao_contratos_detail', pk=backlog.contrato.pk)
        
        try:
            gerente = None
            if gerente_id:
                from .models import Colaborador
                gerente = get_object_or_404(Colaborador, pk=gerente_id)
            
            projeto = backlog.converter_para_projeto(
                nome_projeto=nome_projeto,
                descricao_projeto=descricao_projeto,
                gerente_projeto=gerente  # Colaborador ou None
            )
            
            messages.success(request, f'Backlog convertido em projeto "{projeto.nome}" com sucesso!')
            return redirect('projeto_detail', pk=projeto.pk)
        except Exception as e:
            import traceback
            print(f"Erro ao converter backlog: {traceback.format_exc()}")
            messages.error(request, f'Erro ao converter backlog: {str(e)}')
            return redirect('gestao_contratos_detail', pk=backlog.contrato.pk)
    
    # GET - mostrar formulário
    from .models import Colaborador
    gerentes = Colaborador.objects.filter(ativo=True, cargo__icontains='gerente')
    
    context = {
        'backlog': backlog,
        'gerentes': gerentes,
    }
    return render(request, 'backlog/convert_projeto.html', context)


# Backlog - Gerenciar Tarefas (agora vinculado ao contrato)
@group_required("Admin", "Gerente")
def backlog_gerenciar(request, contrato_id):
    """Gerencia tarefas do backlog do contrato"""
    from .forms import TarefaForm
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    # Buscar o primeiro backlog pendente do contrato ou criar um novo se não existir
    backlog = Backlog.objects.filter(contrato=contrato, status='pendente').first()
    if not backlog:
        backlog = Backlog.objects.create(
            contrato=contrato,
            titulo=f"Backlog - {contrato.numero_contrato}",
            status='pendente'
        )
    
    if request.method == "POST":
        # Criar nova tarefa no backlog
        form = TarefaForm(request.POST)
        if form.is_valid():
            tarefa = form.save(commit=False)
            tarefa.backlog = backlog
            tarefa.save()
            messages.success(request, "Tarefa adicionada ao backlog!")
            return redirect("backlog_gerenciar", contrato_id=contrato_id)
    else:
        form = TarefaForm()
        form.fields['backlog'].initial = backlog
        form.fields['backlog'].widget = forms.HiddenInput()
        # Ocultar sprint e ordem_servico
        if 'sprint' in form.fields:
            form.fields['sprint'].widget = forms.HiddenInput()
        if 'ordem_servico' in form.fields:
            form.fields['ordem_servico'].widget = forms.HiddenInput()
    
    # Apenas tarefas que não estão em sprint (no backlog)
    tarefas = backlog.tarefas.filter(sprint__isnull=True).select_related('responsavel').order_by('-prioridade', '-criado_em')
    
    # Projetos gerados a partir deste backlog
    projetos_gerados = backlog.projetos_gerados.all()
    
    context = {
        "contrato": contrato,
        "backlog": backlog,
        "form": form,
        "tarefas": tarefas,
        "projetos_gerados": projetos_gerados,
    }
    
    return render(request, "backlog/gerenciar.html", context)


# Sprint - Criar
@group_required("Admin", "Gerente")
def sprint_create(request, projeto_id):
    """Cria uma nova sprint"""
    from .forms import SprintForm
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    
    if request.method == "POST":
        form = SprintForm(request.POST, projeto_id=projeto_id)
        if form.is_valid():
            sprint = form.save(commit=False)
            # Garantir que o projeto está vinculado (caso o campo oculto não tenha funcionado)
            sprint.projeto = projeto
            sprint.save()
            messages.success(request, "Sprint criada com sucesso!")
            return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint.pk)
    else:
        form = SprintForm(projeto_id=projeto_id)
        # Filtrar OSs disponíveis (sem sprint vinculada) do contrato do projeto
        if "ordem_servico" in form.fields:
            form.fields["ordem_servico"].queryset = OrdemServico.objects.filter(
                contrato=projeto.contrato,
                sprint__isnull=True
            )
            form.fields["ordem_servico"].empty_label = "Criar OS automaticamente"
    
    return render(request, "sprint/form.html", {"form": form, "projeto": projeto})


# Sprint - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def sprint_detail(request, projeto_id, sprint_id):
    """Detalhes da sprint com tarefas"""
    from datetime import datetime, date
    
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    sprint = get_object_or_404(Sprint, pk=sprint_id, projeto=projeto)
    
    # Atualizar status da sprint e OS com base nas datas
    hoje = date.today()
    novo_status = None
    
    if sprint.data_inicio and sprint.data_fim:
        if sprint.data_inicio > hoje:
            novo_status = "aberta"  # Pendente/Aberta
        elif sprint.data_inicio <= hoje and sprint.data_fim >= hoje:
            novo_status = "execucao"  # Em Execução
        elif sprint.data_fim < hoje:
            novo_status = "finalizada"  # Finalizada
        
        # Atualizar status se mudou (e não está faturada)
        if novo_status and sprint.status != novo_status and sprint.status != "faturada":
            sprint.status = novo_status
            sprint.save()
    
    # Calcular dias e horas restantes para o término
    dias_restantes = None
    horas_restantes = 0
    if sprint.data_fim:
        agora = datetime.now()
        fim_sprint = datetime.combine(sprint.data_fim, datetime.max.time().replace(hour=19, minute=0, second=0))
        diferenca = fim_sprint - agora
        
        if diferenca.total_seconds() > 0:
            dias_restantes = diferenca.days
            horas_restantes = int((diferenca.seconds // 3600))
        else:
            dias_restantes = -abs(diferenca.days)
            horas_restantes = 0
    
    # Tarefas da sprint
    tarefas_sprint = sprint.tarefas.all().select_related('responsavel').order_by('-prioridade', '-criado_em')
    
    # Separar tarefas de consultor e tarefa de gestão
    tarefas_consultor = tarefas_sprint.exclude(titulo__icontains="gestão").exclude(titulo__icontains="gerente")
    tarefa_gestao = tarefas_sprint.filter(titulo__icontains="gestão").first()
    
    # Calcular total de horas do consultor
    total_horas_consultor = sum(t.horas_planejadas for t in tarefas_consultor if t.horas_planejadas)
    
    # Tarefas disponíveis no backlog do projeto (sem sprint)
    tarefas_backlog = Tarefa.objects.filter(
        projeto=projeto,
        sprint__isnull=True
    ).select_related('responsavel').order_by('-prioridade', '-criado_em')
    
    context = {
        "projeto": projeto,
        "sprint": sprint,
        "tarefas_sprint": tarefas_sprint,
        "tarefas_consultor": tarefas_consultor,
        "tarefa_gestao": tarefa_gestao,
        "total_horas_consultor": total_horas_consultor,
        "tarefas_backlog": tarefas_backlog,
        "dias_restantes": dias_restantes,
        "horas_restantes": horas_restantes,
    }
    
    return render(request, "sprint/detail.html", context)


# Sprint - Editar
@group_required("Admin", "Gerente")
def sprint_update(request, projeto_id, sprint_id):
    """Edita uma sprint"""
    from .forms import SprintForm
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    sprint = get_object_or_404(Sprint, pk=sprint_id, projeto=projeto)
    
    if request.method == "POST":
        form = SprintForm(request.POST, instance=sprint, projeto_id=projeto_id)
        if form.is_valid():
            form.save()
            messages.success(request, "Sprint atualizada com sucesso!")
            return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint.pk)
        else:
            messages.error(request, "Erro ao atualizar a sprint. Verifique os campos.")
    else:
        form = SprintForm(instance=sprint, projeto_id=projeto_id)
        
        # Se a sprint tem uma OS vinculada, atualizar os dados da Sprint com os dados da OS
        if sprint.ordem_servico:
            os = sprint.ordem_servico
            # Sincronizar datas da OS com a Sprint (sempre usar os valores da OS)
            if os.data_inicio:
                sprint.data_inicio = os.data_inicio
                form.initial['data_inicio'] = os.data_inicio.strftime('%Y-%m-%d')
            if os.data_termino:
                sprint.data_fim = os.data_termino
                form.initial['data_fim'] = os.data_termino.strftime('%Y-%m-%d')
            # Se a descrição da sprint estiver vazia, preencher com informações da OS
            if not sprint.descricao and os.numero_os:
                sprint.descricao = f"OS: {os.numero_os}"
                form.initial['descricao'] = f"OS: {os.numero_os}"
            
            # Garantir que a OS vinculada esteja selecionada no campo
            form.initial['ordem_servico'] = sprint.ordem_servico.pk
        else:
            # Se não tem OS, usar as datas da sprint
            if sprint.data_inicio:
                form.initial['data_inicio'] = sprint.data_inicio.strftime('%Y-%m-%d')
            if sprint.data_fim:
                form.initial['data_fim'] = sprint.data_fim.strftime('%Y-%m-%d')
        
        # Filtrar OSs disponíveis (sem sprint vinculada) do contrato do projeto
        # ou a OS atual da sprint
        if "ordem_servico" in form.fields:
            os_disponiveis = OrdemServico.objects.filter(
                contrato=projeto.contrato,
                sprint__isnull=True
            )
            # Se a sprint já tem uma OS, incluir ela também
            if sprint.ordem_servico:
                os_disponiveis = os_disponiveis | OrdemServico.objects.filter(pk=sprint.ordem_servico.pk)
            form.fields["ordem_servico"].queryset = os_disponiveis
            form.fields["ordem_servico"].empty_label = "Criar OS automaticamente"
    
    return render(request, "sprint/form.html", {"form": form, "projeto": projeto, "sprint": sprint})


# Sprint - Deletar
@group_required("Admin", "Gerente")
def sprint_delete(request, projeto_id, sprint_id):
    """Exclui uma sprint e sua OS vinculada"""
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    sprint = get_object_or_404(Sprint, pk=sprint_id, projeto=projeto)
    
    if request.method == "POST":
        # Deletar a OS vinculada antes de deletar a Sprint
        ordem_servico_numero = None
        if sprint.ordem_servico:
            ordem_servico_numero = sprint.ordem_servico.numero_os
            # Deletar a OS primeiro
            sprint.ordem_servico.delete()
        # Agora deletar a Sprint
        sprint.delete()
        if ordem_servico_numero:
            messages.success(request, f"Sprint excluída com sucesso! A OS {ordem_servico_numero} também foi excluída.")
        else:
            messages.success(request, "Sprint excluída com sucesso!")
        return redirect("projeto_detail", pk=projeto_id)
    
    context = {
        "projeto": projeto,
        "sprint": sprint,
    }
    return render(request, "sprint/confirm_delete.html", context)


# Sprint - Fechar (Criar OS)
@group_required("Admin", "Gerente")
def sprint_fechar(request, projeto_id, sprint_id):
    """Fecha a sprint e cria uma OS automaticamente"""
    from .forms import OrdemServicoForm
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    sprint = get_object_or_404(Sprint, pk=sprint_id, projeto=projeto)
    
    if sprint.status == "finalizada":
        messages.error(request, "Esta sprint já foi finalizada!")
        return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)
    
    # Obter tarefas da sprint
    tarefas_sprint = sprint.tarefas.all()
    
    if not tarefas_sprint.exists():
        messages.error(request, "Não é possível fechar uma sprint sem tarefas!")
        return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)
    
    # Calcular totais das tarefas
    total_horas_planejadas = sum(t.horas_planejadas for t in tarefas_sprint if t.horas_planejadas)
    total_horas_consumidas = sum(t.horas_consumidas for t in tarefas_sprint if t.horas_consumidas)
    
    if request.method == "POST":
        # Criar OS baseada na sprint
        form = OrdemServicoForm(request.POST)
        if form.is_valid():
            os_item = form.save(commit=False)
            os_item.cliente = projeto.contrato.cliente
            os_item.contrato = projeto.contrato
            
            # Definir datas baseadas na sprint
            os_item.data_inicio = sprint.data_inicio
            os_item.data_termino = sprint.data_fim
            
            # Calcular quantidade total (horas)
            os_item.quantidade = total_horas_planejadas or Decimal('0.00')
            
            # Salvar OS (gera número automaticamente)
            os_item.save()
            
            # Vincular sprint à OS
            sprint.ordem_servico = os_item
            sprint.status = "finalizada"
            sprint.save()
            
            # Mover tarefas da sprint para a OS e atualizar horas planejadas/realizadas
            for tarefa in tarefas_sprint:
                tarefa.ordem_servico = os_item
                tarefa.save()
            
            # Atualizar horas planejadas e realizadas da OS
            os_item.horas_planejadas = total_horas_planejadas
            os_item.horas_realizadas = total_horas_consumidas
            os_item.save()
            
            messages.success(request, f"Sprint finalizada e OS {os_item.numero_os} criada com sucesso!")
            return redirect("ordem_servico_detail", pk=os_item.pk)
    else:
        # Preencher formulário com dados sugeridos
        form = OrdemServicoForm()
        form.fields['contrato'].initial = projeto.contrato
        form.fields['cliente'].initial = projeto.contrato.cliente
        
        # Tentar sugerir item_contrato do tipo Serviço do contrato
        itens_servico = projeto.contrato.itens.filter(tipo="Serviço").first()
        if itens_servico:
            form.fields['item_contrato'].initial = itens_servico
        
        # Preencher quantidade com total de horas
        form.fields['quantidade'].initial = total_horas_planejadas
        
        # Preencher datas
        form.fields['data_inicio'].initial = sprint.data_inicio
        form.fields['data_termino'].initial = sprint.data_fim
    
    context = {
        "projeto": projeto,
        "sprint": sprint,
        "form": form,
        "tarefas_sprint": tarefas_sprint,
        "total_horas_planejadas": total_horas_planejadas,
        "total_horas_consumidas": total_horas_consumidas,
    }
    
    return render(request, "sprint/fechar.html", context)


# Tarefa - Mover para Sprint
@group_required("Admin", "Gerente")
def tarefa_mover_sprint(request, projeto_id, sprint_id, tarefa_id):
    """Move uma tarefa do backlog para a sprint"""
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    sprint = get_object_or_404(Sprint, pk=sprint_id, projeto=projeto)
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id)
    
    # Validar se a tarefa pertence ao projeto
    if tarefa.projeto == projeto:
        from datetime import datetime, time
        from django.utils import timezone as tz
        from decimal import Decimal
        
        # Validar datas da tarefa em relação à sprint
        if tarefa.data_inicio_prevista:
            # Converter data_inicio da sprint para datetime se necessário
            if isinstance(sprint.data_inicio, datetime):
                data_inicio_sprint = sprint.data_inicio
            else:
                data_inicio_sprint = datetime.combine(sprint.data_inicio, time.min)
                # Tornar timezone-aware se o data_inicio_prevista for aware
                if tz.is_aware(tarefa.data_inicio_prevista):
                    data_inicio_sprint = tz.make_aware(data_inicio_sprint)
            
            if tarefa.data_inicio_prevista < data_inicio_sprint:
                messages.error(request, f"A data de início da tarefa ({tarefa.data_inicio_prevista.strftime('%d/%m/%Y %H:%M')}) não pode ser anterior à data de início da sprint ({sprint.data_inicio.strftime('%d/%m/%Y')})!")
                return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)
        
        if tarefa.data_termino_prevista:
            # Converter data_fim da sprint para datetime se necessário
            if isinstance(sprint.data_fim, datetime):
                data_fim_sprint = sprint.data_fim
            else:
                data_fim_sprint = datetime.combine(sprint.data_fim, time(23, 59, 59))
                # Tornar timezone-aware se o data_termino_prevista for aware
                if tz.is_aware(tarefa.data_termino_prevista):
                    data_fim_sprint = tz.make_aware(data_fim_sprint)
            
            if tarefa.data_termino_prevista > data_fim_sprint:
                messages.error(request, f"A data de término da tarefa ({tarefa.data_termino_prevista.strftime('%d/%m/%Y %H:%M')}) não pode ser maior que a data de fim da sprint ({sprint.data_fim.strftime('%d/%m/%Y')})!")
                return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)
        
        tarefa.sprint = sprint
        tarefa.status_sprint = "nao_iniciada"  # Status inicial ao mover para sprint
        tarefa.save()
        
        # Recalcular total de horas de todas as tarefas de consultor na sprint
        tarefas_consultor_sprint = sprint.tarefas.exclude(titulo__icontains="gestão").exclude(titulo__icontains="gerente")
        total_horas_consultor = sum(t.horas_planejadas for t in tarefas_consultor_sprint if t.horas_planejadas) or Decimal('0.00')
        
        # Se for a primeira tarefa de consultor movida, criar tarefa de gestão automaticamente
        if tarefas_consultor_sprint.count() == 1:  # Primeira tarefa de consultor
            # Verificar se já existe tarefa de gestão
            if not sprint.tarefas.filter(titulo__icontains="gestão").exists():
                # Criar tarefa de gestão (25% das horas do consultor)
                horas_gestao = total_horas_consultor * Decimal('0.25')
                
                # Converter datas da sprint para datetime se necessário
                if isinstance(sprint.data_inicio, datetime):
                    data_inicio_gestao = sprint.data_inicio
                else:
                    data_inicio_gestao = datetime.combine(sprint.data_inicio, time.min)
                    if tz.is_aware(tarefa.data_inicio_prevista) if tarefa.data_inicio_prevista else False:
                        data_inicio_gestao = tz.make_aware(data_inicio_gestao)
                
                if isinstance(sprint.data_fim, datetime):
                    data_termino_gestao = sprint.data_fim
                else:
                    data_termino_gestao = datetime.combine(sprint.data_fim, time(23, 59, 59))
                    if tz.is_aware(tarefa.data_termino_prevista) if tarefa.data_termino_prevista else False:
                        data_termino_gestao = tz.make_aware(data_termino_gestao)
                
                tarefa_gestao = Tarefa.objects.create(
                    projeto=projeto,
                    titulo=f"Gestão de Projetos - {sprint.nome}",
                    descricao="Tarefa de gestão de projetos (25% das horas do consultor técnico)",
                    sprint=sprint,
                    responsavel=projeto.gerente_projeto,
                    prioridade="media",
                    status_sprint="nao_iniciada",
                    data_inicio_prevista=data_inicio_gestao,
                    data_termino_prevista=data_termino_gestao,
                    horas_planejadas=horas_gestao,
                )
                messages.success(request, f"Tarefa movida para a sprint! Tarefa de gestão criada automaticamente ({horas_gestao}h).")
            else:
                messages.success(request, "Tarefa movida para a sprint!")
        else:
            # Atualizar horas da tarefa de gestão se já existir
            tarefa_gestao = sprint.tarefas.filter(titulo__icontains="gestão").first()
            if tarefa_gestao:
                # Recalcular baseado na soma de todas as tarefas de consultor
                horas_gestao = total_horas_consultor * Decimal('0.25')
                tarefa_gestao.horas_planejadas = horas_gestao
                tarefa_gestao.save()
            messages.success(request, "Tarefa movida para a sprint!")
    else:
        messages.error(request, "Tarefa não pertence a este projeto!")
    
    return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)


# Tarefa - Remover da Sprint (voltar ao backlog)
@group_required("Admin", "Gerente")
def tarefa_remover_sprint(request, projeto_id, sprint_id, tarefa_id):
    """Remove uma tarefa da sprint e retorna ao backlog"""
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    sprint = get_object_or_404(Sprint, pk=sprint_id, projeto=projeto)
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id, sprint=sprint)
    
    # Não permitir remover tarefa de gestão
    if "gestão" in tarefa.titulo.lower():
        messages.error(request, "A tarefa de gestão não pode ser removida da sprint!")
        return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)
    
    # Remover tarefa da sprint (volta para o backlog do projeto)
    tarefa.sprint = None
    tarefa.status_sprint = None
    tarefa.save()
    
    # Atualizar horas da tarefa de gestão se ainda houver tarefas de consultor
    tarefas_consultor_sprint = sprint.tarefas.exclude(titulo__icontains="gestão").exclude(titulo__icontains="gerente")
    tarefa_gestao = sprint.tarefas.filter(titulo__icontains="gestão").first()
    
    if tarefa_gestao:
        if tarefas_consultor_sprint.exists():
            # Recalcular baseado na soma de todas as tarefas de consultor na sprint
            total_horas_consultor = sum(t.horas_planejadas for t in tarefas_consultor_sprint if t.horas_planejadas) or Decimal('0.00')
            horas_gestao = total_horas_consultor * Decimal('0.25')
            tarefa_gestao.horas_planejadas = horas_gestao
            tarefa_gestao.save()
        else:
            # Se não há mais tarefas de consultor, remover tarefa de gestão
            tarefa_gestao.delete()
            messages.success(request, "Tarefa removida da sprint e retornada ao backlog! Tarefa de gestão removida (não há mais tarefas de consultor).")
            return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)
    
    messages.success(request, "Tarefa removida da sprint e retornada ao backlog!")
    return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=sprint_id)


# Tarefa - Mover via AJAX (para drag and drop no Canvas)
@group_required("Admin", "Gerente")
@require_http_methods(["POST"])
def tarefa_mover_ajax(request, projeto_id, tarefa_id):
    """Move uma tarefa para uma sprint ou para o backlog via AJAX"""
    import json
    from django.http import JsonResponse
    
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id)
    
    # Validar se a tarefa pertence ao projeto
    if tarefa.projeto != projeto:
        return JsonResponse({'success': False, 'error': 'Tarefa não pertence a este projeto!'}, status=400)
    
    try:
        data = json.loads(request.body)
        nova_sprint_id = data.get('sprint_id')  # None para mover para backlog
        
        from datetime import datetime, time
        from django.utils import timezone as tz
        from decimal import Decimal
        
        if nova_sprint_id:
            # Mover para sprint
            sprint = get_object_or_404(Sprint, pk=nova_sprint_id, projeto=projeto)
            
            # Validar datas da tarefa em relação à sprint
            if tarefa.data_inicio_prevista:
                if isinstance(sprint.data_inicio, datetime):
                    data_inicio_sprint = sprint.data_inicio
                else:
                    data_inicio_sprint = datetime.combine(sprint.data_inicio, time.min)
                    if tz.is_aware(tarefa.data_inicio_prevista):
                        data_inicio_sprint = tz.make_aware(data_inicio_sprint)
                
                if tarefa.data_inicio_prevista < data_inicio_sprint:
                    return JsonResponse({
                        'success': False, 
                        'error': f'A data de início da tarefa não pode ser anterior à data de início da sprint!'
                    }, status=400)
            
            if tarefa.data_termino_prevista:
                if isinstance(sprint.data_fim, datetime):
                    data_fim_sprint = sprint.data_fim
                else:
                    data_fim_sprint = datetime.combine(sprint.data_fim, time(23, 59, 59))
                    if tz.is_aware(tarefa.data_termino_prevista):
                        data_fim_sprint = tz.make_aware(data_fim_sprint)
                
                if tarefa.data_termino_prevista > data_fim_sprint:
                    return JsonResponse({
                        'success': False, 
                        'error': f'A data de término da tarefa não pode ser maior que a data de fim da sprint!'
                    }, status=400)
            
            tarefa.sprint = sprint
            tarefa.status_sprint = "nao_iniciada"  # Status inicial ao mover para sprint
            
            # Se uma posição foi especificada, reordenar as tarefas
            if nova_posicao is not None:
                # Obter todas as tarefas da sprint (exceto a que está sendo movida)
                outras_tarefas = Tarefa.objects.filter(
                    sprint=sprint,
                    projeto=projeto
                ).exclude(pk=tarefa.pk).order_by('ordem_sprint', '-criado_em')
                
                # Reordenar: inserir a tarefa na posição especificada
                tarefas_lista = list(outras_tarefas)
                tarefas_lista.insert(nova_posicao, tarefa)
                
                # Atualizar ordem_sprint de todas as tarefas
                # Primeiro, atualizar todas as outras tarefas
                for idx, t in enumerate(tarefas_lista):
                    if t.pk != tarefa.pk:
                        Tarefa.objects.filter(pk=t.pk).update(ordem_sprint=idx)
                
                # Depois, atualizar a tarefa que está sendo movida
                tarefa.ordem_sprint = nova_posicao
                tarefa.save(update_fields=['sprint', 'status_sprint', 'ordem_sprint'])
            else:
                # Se não especificou posição, colocar no final
                from django.db.models import Max
                ultima_ordem = Tarefa.objects.filter(
                    sprint=sprint,
                    projeto=projeto
                ).exclude(pk=tarefa.pk).aggregate(
                    max_ordem=Max('ordem_sprint')
                )['max_ordem'] or -1
                tarefa.ordem_sprint = ultima_ordem + 1
                tarefa.save(update_fields=['sprint', 'status_sprint', 'ordem_sprint'])
            
            # Recalcular total de horas de todas as tarefas de consultor na sprint
            tarefas_consultor_sprint = sprint.tarefas.exclude(titulo__icontains="gestão").exclude(titulo__icontains="gerente")
            total_horas_consultor = sum(t.horas_planejadas for t in tarefas_consultor_sprint if t.horas_planejadas) or Decimal('0.00')
            
            # Se for a primeira tarefa de consultor movida, criar tarefa de gestão automaticamente
            if tarefas_consultor_sprint.count() == 1:
                if not sprint.tarefas.filter(titulo__icontains="gestão").exists():
                    horas_gestao = total_horas_consultor * Decimal('0.25')
                    
                    if isinstance(sprint.data_inicio, datetime):
                        data_inicio_gestao = sprint.data_inicio
                    else:
                        data_inicio_gestao = datetime.combine(sprint.data_inicio, time.min)
                        if tz.is_aware(tarefa.data_inicio_prevista) if tarefa.data_inicio_prevista else False:
                            data_inicio_gestao = tz.make_aware(data_inicio_gestao)
                    
                    if isinstance(sprint.data_fim, datetime):
                        data_termino_gestao = sprint.data_fim
                    else:
                        data_termino_gestao = datetime.combine(sprint.data_fim, time(23, 59, 59))
                        if tz.is_aware(tarefa.data_termino_prevista) if tarefa.data_termino_prevista else False:
                            data_termino_gestao = tz.make_aware(data_termino_gestao)
                    
                    Tarefa.objects.create(
                        projeto=projeto,
                        titulo=f"Gestão de Projetos - {sprint.nome}",
                        descricao="Tarefa de gestão de projetos (25% das horas do consultor técnico)",
                        sprint=sprint,
                        responsavel=projeto.gerente_projeto,
                        prioridade="media",
                        status_sprint="nao_iniciada",
                        data_inicio_prevista=data_inicio_gestao,
                        data_termino_prevista=data_termino_gestao,
                        horas_planejadas=horas_gestao,
                    )
            else:
                # Atualizar horas da tarefa de gestão se já existir
                tarefa_gestao = sprint.tarefas.filter(titulo__icontains="gestão").first()
                if tarefa_gestao:
                    horas_gestao = total_horas_consultor * Decimal('0.25')
                    tarefa_gestao.horas_planejadas = horas_gestao
                    tarefa_gestao.save()
        else:
            # Mover para backlog (remover da sprint)
            # Não permitir remover tarefa de gestão
            if "gestão" in tarefa.titulo.lower():
                return JsonResponse({
                    'success': False, 
                    'error': 'A tarefa de gestão não pode ser removida da sprint!'
                }, status=400)
            
            sprint_anterior = tarefa.sprint
            tarefa.sprint = None
            tarefa.status_sprint = None
            tarefa.ordem_sprint = 0  # Resetar ordem ao voltar para backlog
            tarefa.save(update_fields=['sprint', 'status_sprint', 'ordem_sprint'])
            
            # Atualizar horas da tarefa de gestão se ainda houver tarefas de consultor
            if sprint_anterior:
                tarefas_consultor_sprint = sprint_anterior.tarefas.exclude(titulo__icontains="gestão").exclude(titulo__icontains="gerente")
                tarefa_gestao = sprint_anterior.tarefas.filter(titulo__icontains="gestão").first()
                
                if tarefa_gestao:
                    if tarefas_consultor_sprint.exists():
                        total_horas_consultor = sum(t.horas_planejadas for t in tarefas_consultor_sprint if t.horas_planejadas) or Decimal('0.00')
                        horas_gestao = total_horas_consultor * Decimal('0.25')
                        tarefa_gestao.horas_planejadas = horas_gestao
                        tarefa_gestao.save()
                    else:
                        # Se não houver mais tarefas de consultor, remover tarefa de gestão
                        tarefa_gestao.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Tarefa movida com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# Tarefa - Criar (no projeto)
@group_required("Admin", "Gerente")
def tarefa_projeto_create(request, projeto_id):
    """Cria uma nova tarefa no projeto"""
    from .forms import TarefaForm
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    if request.method == "POST":
        # Garantir que o projeto seja enviado no POST
        post_data = request.POST.copy()
        post_data['projeto'] = projeto_id
        form = TarefaForm(post_data, projeto_id=projeto_id)
        if form.is_valid():
            tarefa = form.save(commit=False)
            # Definir projeto obrigatoriamente
            tarefa.projeto = projeto
            # Se não foi selecionada sprint, a tarefa fica no backlog do projeto (sprint=None)
            # Se foi selecionada sprint, definir status_sprint inicial
            if tarefa.sprint and not tarefa.status_sprint:
                tarefa.status_sprint = "nao_iniciada"
            tarefa.save()
            messages.success(request, "Tarefa criada com sucesso!")
            return redirect("projeto_detail", pk=projeto_id)
        else:
            messages.error(request, "Erro ao criar a tarefa. Verifique os campos.")
    else:
        form = TarefaForm(projeto_id=projeto_id)
    
    context = {
        "form": form,
        "projeto": projeto,
        "sprints": projeto.sprints.all().order_by('-data_inicio'),
    }
    return render(request, "tarefa/form_projeto_create.html", context)


# Tarefa - Editar (do projeto/sprint)
@group_required("Admin", "Gerente")
def tarefa_projeto_update(request, projeto_id, tarefa_id):
    """Edita uma tarefa do projeto (backlog ou sprint)"""
    from .forms import TarefaForm
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id)
    
    # Verificar se a tarefa pertence ao projeto (agora tarefa tem projeto obrigatório)
    if tarefa.projeto != projeto:
        messages.error(request, "Tarefa não pertence a este projeto!")
        if tarefa.sprint:
            return redirect("sprint_detail", projeto_id=tarefa.projeto.pk, sprint_id=tarefa.sprint.pk)
        return redirect("projeto_detail", pk=tarefa.projeto.pk)
    
    if request.method == "POST":
        form = TarefaForm(request.POST, instance=tarefa, projeto_id=projeto_id)
        if form.is_valid():
            tarefa_editada = form.save(commit=False)
            
            # Validar data fim se estiver em sprint
            if tarefa_editada.sprint and tarefa_editada.data_termino_prevista:
                from datetime import datetime, time
                from django.utils import timezone as tz
                
                # Converter data_fim da sprint para datetime se necessário
                if isinstance(tarefa_editada.sprint.data_fim, datetime):
                    data_fim_sprint = tarefa_editada.sprint.data_fim
                else:
                    data_fim_sprint = datetime.combine(tarefa_editada.sprint.data_fim, time(23, 59, 59))
                    # Tornar timezone-aware se o data_termino_prevista for aware
                    if tz.is_aware(tarefa_editada.data_termino_prevista):
                        data_fim_sprint = tz.make_aware(data_fim_sprint)
                
                if tarefa_editada.data_termino_prevista > data_fim_sprint:
                    messages.error(request, f"A data de término não pode ser maior que a data de fim da sprint ({tarefa_editada.sprint.data_fim.strftime('%d/%m/%Y')})!")
                    return render(request, "tarefa/form_projeto.html", {"form": form, "projeto": projeto, "tarefa": tarefa})
            
            tarefa_editada.save()
            
            # Se a tarefa está em sprint e é de consultor, atualizar tarefa de gestão
            if tarefa_editada.sprint and "gestão" not in tarefa_editada.titulo.lower():
                tarefa_gestao = tarefa_editada.sprint.tarefas.filter(titulo__icontains="gestão").first()
                if tarefa_gestao:
                    # Recalcular baseado na soma de todas as tarefas de consultor na sprint
                    tarefas_consultor = tarefa_editada.sprint.tarefas.exclude(titulo__icontains="gestão").exclude(titulo__icontains="gerente")
                    total_horas_consultor = sum(t.horas_planejadas for t in tarefas_consultor if t.horas_planejadas) or Decimal('0.00')
                    horas_gestao = total_horas_consultor * Decimal('0.25')
                    tarefa_gestao.horas_planejadas = horas_gestao
                    tarefa_gestao.save()
            
            messages.success(request, "Tarefa atualizada com sucesso!")
            if tarefa_editada.sprint:
                return redirect("sprint_detail", projeto_id=projeto_id, sprint_id=tarefa_editada.sprint.pk)
            return redirect("projeto_detail", pk=projeto_id)
    else:
        form = TarefaForm(instance=tarefa, projeto_id=projeto_id)
        
        # Garantir que os valores de data/hora sejam carregados no formato correto para datetime-local
        from django.utils import timezone as tz
        if tarefa.data_inicio_prevista:
            # Converter para timezone local antes de formatar
            data_inicio_local = tz.localtime(tarefa.data_inicio_prevista) if tz.is_aware(tarefa.data_inicio_prevista) else tarefa.data_inicio_prevista
            form.initial['data_inicio_prevista'] = data_inicio_local.strftime('%Y-%m-%dT%H:%M')
        if tarefa.data_termino_prevista:
            # Converter para timezone local antes de formatar
            data_termino_local = tz.localtime(tarefa.data_termino_prevista) if tz.is_aware(tarefa.data_termino_prevista) else tarefa.data_termino_prevista
            form.initial['data_termino_prevista'] = data_termino_local.strftime('%Y-%m-%dT%H:%M')
        
        # Ocultar campos conforme contexto
        if tarefa.sprint:
            if "status_sprint" in form.fields:
                form.fields["status_sprint"].required = False
            if "status" in form.fields:
                form.fields["status"].widget = forms.HiddenInput()
        else:
            if "status_sprint" in form.fields:
                form.fields["status_sprint"].widget = forms.HiddenInput()
    
    return render(request, "tarefa/form_projeto.html", {"form": form, "projeto": projeto, "tarefa": tarefa})


# Tarefa - Ver Detalhes (do projeto)
@group_required("Admin", "Gerente", "Leitor")
def tarefa_projeto_detail(request, projeto_id, tarefa_id):
    """Visualiza detalhes de uma tarefa do projeto (read-only)"""
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id)
    
    # Verificar se a tarefa pertence ao projeto (agora tarefa tem projeto obrigatório)
    if tarefa.projeto != projeto:
        messages.error(request, "Tarefa não pertence a este projeto!")
        return redirect("projeto_detail", pk=tarefa.projeto.pk)
    
    context = {
        "projeto": projeto,
        "tarefa": tarefa,
    }
    return render(request, "tarefa/detail_projeto.html", context)


# Tarefa - Deletar (do projeto)
@group_required("Admin", "Gerente")
def tarefa_projeto_delete(request, projeto_id, tarefa_id):
    """Exclui uma tarefa do projeto"""
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id)
    
    # Verificar se a tarefa pertence ao projeto (agora tarefa tem projeto obrigatório)
    if tarefa.projeto != projeto:
        messages.error(request, "Tarefa não pertence a este projeto!")
        return redirect("projeto_detail", pk=tarefa.projeto.pk)
    
    if request.method == "POST":
        tarefa.delete()
        messages.success(request, "Tarefa excluída com sucesso!")
        return redirect("projeto_detail", pk=projeto_id)
    
    context = {
        "projeto": projeto,
        "tarefa": tarefa,
    }
    return render(request, "tarefa/confirm_delete_projeto.html", context)


# ========== VIEWS PARA GESTÃO DE CONTRATOS ==========
# Modelo unificado para todos os regimes: Lei 14.133, Lei 13.303 e Privado

# Gestão de Contratos - Listar (todos os regimes)
@group_required("Admin", "Gerente", "Leitor")
def gestao_contratos_list(request):
    """Lista todos os contratos (todos os regimes)"""
    contratos = Contrato.objects.select_related('cliente').order_by('-data_assinatura')
    
    # Filtros
    search_query = request.GET.get('search')
    regime_filter = request.GET.get('regime_legal')
    renovacao_pendente = request.GET.get('renovacao_pendente')
    
    if search_query:
        contratos = contratos.filter(
            Q(numero_contrato__icontains=search_query) |
            Q(cliente__nome_razao_social__icontains=search_query) |
            Q(objeto__icontains=search_query)
        )
    if regime_filter:
        contratos = contratos.filter(regime_legal=regime_filter)
    if renovacao_pendente == 'true':
        # Filtrar contratos com renovação pendente
        contratos_pendentes = ContratoService.listar_contratos_com_renovacao_pendente()
        contratos = contratos.filter(pk__in=[c.pk for c in contratos_pendentes])
    
    # Contadores por regime
    total_contratos = Contrato.objects.count()
    total_lei_14133 = Contrato.objects.filter(regime_legal=RegimeLegal.LEI_14133).count()
    total_lei_13303 = Contrato.objects.filter(regime_legal=RegimeLegal.LEI_13303).count()
    total_privados = Contrato.objects.filter(regime_legal=RegimeLegal.PRIVADO).count()
    total_renovacao_pendente = len(ContratoService.listar_contratos_com_renovacao_pendente())
    
    paginator = Paginator(contratos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        "page_obj": page_obj,
        "search_query": search_query,
        "regime_filter": regime_filter,
        "renovacao_pendente": renovacao_pendente,
        "regime_choices": RegimeLegal.choices,
        "total_contratos": total_contratos,
        "total_lei_14133": total_lei_14133,
        "total_lei_13303": total_lei_13303,
        "total_privados": total_privados,
        "total_renovacao_pendente": total_renovacao_pendente,
    }
    return render(request, "gestao_contratos/list.html", context)


# Gestão de Contratos - Detalhar
@group_required("Admin", "Gerente", "Leitor")
def gestao_contratos_detail(request, pk):
    """Detalhes de um contrato"""
    contrato = get_object_or_404(Contrato.objects.select_related('cliente'), pk=pk)
    
    # Usar service layer para obter resumo completo
    resumo = ContratoService.obter_resumo_contrato(contrato)
    
    # Calcular dados para as barras de progresso
    from datetime import date
    hoje = date.today()
    
    # Progresso da Vigência: dias que já se passaram / total de dias
    dias_totais_vigencia = 0
    dias_passados = 0
    if contrato.data_assinatura and resumo.get('data_fim_atual'):
        dias_totais_vigencia = (resumo['data_fim_atual'] - contrato.data_assinatura).days
        dias_passados = (hoje - contrato.data_assinatura).days
        if dias_passados < 0:
            dias_passados = 0
        if dias_passados > dias_totais_vigencia:
            dias_passados = dias_totais_vigencia
    
    # Limite de Aditivo de Valor: valor consumido / 25% do valor inicial
    limite_aditivo_valor = resumo['valor_inicial'] * Decimal('0.25')
    valor_consumido_aditivos = resumo['valor_atual'] - resumo['valor_inicial']
    if valor_consumido_aditivos < 0:
        valor_consumido_aditivos = Decimal('0.00')
    
    # Consumo dos Itens: valor faturado / valor total dos itens
    valor_total_itens = contrato.get_valor_total_itens()
    valor_faturado_itens = contrato.get_valor_total_faturado()
    if valor_total_itens == 0:
        percentual_consumo_itens = 0
    else:
        percentual_consumo_itens = (valor_faturado_itens / valor_total_itens) * 100
    
    resumo['dias_totais_vigencia'] = dias_totais_vigencia
    resumo['dias_passados'] = dias_passados
    resumo['limite_aditivo_valor'] = limite_aditivo_valor
    resumo['valor_consumido_aditivos'] = valor_consumido_aditivos
    resumo['valor_total_itens'] = valor_total_itens
    resumo['valor_faturado_itens'] = valor_faturado_itens
    resumo['percentual_consumo_itens'] = percentual_consumo_itens
    
    # Termos aditivos ordenados por data
    termos_aditivos = contrato.termos_aditivos.all().order_by('-data_assinatura')
    
    # Itens do contrato
    itens_contrato = contrato.itens.all().order_by('lote', 'numero_item')
    
    # Ordens de Fornecimento do contrato
    ordens_fornecimento = OrdemFornecimento.objects.filter(contrato=contrato).select_related(
        'item_contrato'
    ).prefetch_related('itens_fornecedor__item_fornecedor').order_by('-id')
    
    # Ordens de Serviço do contrato
    ordens_servico = OrdemServico.objects.filter(contrato=contrato).select_related(
        'item_contrato'
    ).prefetch_related('itens_fornecedor__item_fornecedor').order_by('-id')
    
    # Formulário para novo item de contrato
    item_form = ItemContratoForm(initial={'contrato': contrato})
    # Esconder campo de contrato pois já está definido
    item_form.fields['contrato'].widget = forms.HiddenInput()
    item_form.fields['contrato'].initial = contrato.pk
    # Remover campo data_ativacao do formulário
    if 'data_ativacao' in item_form.fields:
        del item_form.fields['data_ativacao']
    
    # Formulário para novo SLA
    from .forms import SLAForm
    sla_form = SLAForm(initial={'contrato': contrato, 'data_inicio': contrato.data_assinatura})
    sla_form.fields['contrato'].widget = forms.HiddenInput()
    sla_form.fields['contrato'].initial = contrato.pk
    
    # SLAs do contrato
    slas = contrato.slas.all().order_by('-criado_em')
    slas_importantes = contrato.slas_importantes.all().order_by('-prioridade', '-criado_em')
    
    # Backlogs do contrato (demandas que ainda não viraram projetos)
    backlogs = contrato.backlogs.filter(status__in=['pendente', 'em_analise']).order_by('-prioridade', '-criado_em')
    
    # Projetos do contrato
    projetos = contrato.projetos.all().select_related('gerente_projeto').prefetch_related('plano_trabalho').order_by('-criado_em')
    
    # Gerentes de projeto para o formulário
    from .models import Colaborador
    gerentes = Colaborador.objects.filter(
        ativo=True,
        cargo__icontains='gerente'
    ).order_by('nome_completo')
    
    # Processar formulário de item
    if request.method == "POST" and 'add_item' in request.POST:
        item_form = ItemContratoForm(request.POST)
        if 'data_ativacao' in item_form.fields:
            del item_form.fields['data_ativacao']
        if item_form.is_valid():
            item = item_form.save(commit=False)
            item.contrato = contrato
            item.save()
            messages.success(request, f"Item {item.numero_item} adicionado com sucesso!")
            return redirect(f"{reverse('gestao_contratos_detail', kwargs={'pk': pk})}?tab=itens")
        else:
            messages.error(request, "Erro ao adicionar item. Verifique os campos.")
    
    # Processar formulário de SLA
    if request.method == "POST" and 'add_sla' in request.POST:
        sla_form = SLAForm(request.POST)
        sla_form.fields['contrato'].widget = forms.HiddenInput()
        if sla_form.is_valid():
            sla = sla_form.save(commit=False)
            sla.contrato = contrato
            sla.save()
            messages.success(request, f"SLA '{sla.titulo}' adicionado com sucesso!")
            return redirect("gestao_contratos_detail", pk=pk)
        else:
            messages.error(request, "Erro ao adicionar SLA. Verifique os campos.")
    
    context = {
        "contrato": contrato,
        "resumo": resumo,
        "termos_aditivos": termos_aditivos,
        "itens_contrato": itens_contrato,
        "ordens_fornecimento": ordens_fornecimento,
        "ordens_servico": ordens_servico,
        "item_form": item_form,
        "sla_form": sla_form,
        "slas": slas,
        "slas_importantes": slas_importantes,
        "backlogs": backlogs,
        "projetos": projetos,
        "gerentes": gerentes,
    }
    return render(request, "gestao_contratos/detail.html", context)


# Gestão de Contratos - Criar
@group_required("Admin", "Gerente")
def gestao_contratos_create(request):
    """Cria um novo contrato"""
    if request.method == "POST":
        form = ContratoPublicoForm(request.POST)
        if form.is_valid():
            contrato = form.save(commit=False)
            
            # Validação de coerência entre regime_legal e origem_contrato
            if contrato.regime_legal and contrato.origem_contrato:
                is_valid, error_msg = ContratoService.origem_aceitavel_para_regime(contrato)
                if not is_valid:
                    messages.warning(request, f"Atenção: {error_msg}")
            
            contrato.save()
            messages.success(request, "Contrato criado com sucesso!")
            return redirect("gestao_contratos_list")
        else:
            messages.error(request, "Erro ao criar o contrato. Verifique os campos.")
    else:
        form = ContratoPublicoForm()
    
    return render(request, "gestao_contratos/form.html", {"form": form})


# Gestão de Contratos - Editar
@group_required("Admin", "Gerente")
def gestao_contratos_update(request, pk):
    """Edita um contrato"""
    contrato = get_object_or_404(Contrato, pk=pk)
    
    if request.method == "POST":
        form = ContratoPublicoForm(request.POST, instance=contrato)
        if form.is_valid():
            contrato = form.save(commit=False)
            
            # Validação de coerência entre regime_legal e origem_contrato
            if contrato.regime_legal and contrato.origem_contrato:
                is_valid, error_msg = ContratoService.origem_aceitavel_para_regime(contrato)
                if not is_valid:
                    messages.warning(request, f"Atenção: {error_msg}")
            
            contrato.save()
            messages.success(request, "Contrato atualizado com sucesso!")
            return redirect("gestao_contratos_detail", pk=contrato.pk)
    else:
        form = ContratoPublicoForm(instance=contrato)
    
    return render(request, "gestao_contratos/form.html", {"form": form, "contrato": contrato})


# Gestão de Contratos - Deletar
@group_required("Admin", "Gerente")
def gestao_contratos_delete(request, pk):
    """Exclui um contrato"""
    contrato = get_object_or_404(Contrato, pk=pk)
    
    if request.method == "POST":
        total_termos = contrato.termos_aditivos.count()
        contrato.delete()
        messages.success(request, f"Contrato excluído com sucesso! ({total_termos} termo(s) aditivo(s) também foram excluídos.)")
        return redirect("gestao_contratos_list")
    
    context = {
        "contrato": contrato,
        "total_termos": contrato.termos_aditivos.count(),
    }
    return render(request, "gestao_contratos/confirm_delete.html", context)


# Termo Aditivo - Criar
@group_required("Admin", "Gerente")
def termo_aditivo_create(request, contrato_id):
    """Cria um novo termo aditivo usando Service Layer"""
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    
    if request.method == "POST":
        form = TermoAditivoForm(request.POST, contrato_id=contrato_id, contrato=contrato)
        if form.is_valid():
            try:
                # Usar Service Layer para criar com validações legais
                termo = ContratoService.criar_termo_aditivo(
                    contrato=contrato,
                    numero_termo=form.cleaned_data['numero_termo'],
                    tipo=form.cleaned_data['tipo'],
                    meses_acrescimo=form.cleaned_data.get('meses_acrescimo') or 0,
                    valor_acrescimo=form.cleaned_data.get('valor_acrescimo') or Decimal('0.00'),
                    data_assinatura=form.cleaned_data['data_assinatura'],
                    justificativa=form.cleaned_data.get('justificativa') or ''
                )
                messages.success(request, "Termo aditivo criado com sucesso!")
                return redirect("gestao_contratos_detail", pk=contrato.pk)
            except Exception as e:
                messages.error(request, f"Erro ao criar termo aditivo: {str(e)}")
        else:
            messages.error(request, "Erro ao criar o termo aditivo. Verifique os campos.")
    else:
        form = TermoAditivoForm(contrato_id=contrato_id, contrato=contrato)
    
    # Obter informações do contrato para validações
    resumo = ContratoService.obter_resumo_contrato(contrato)
    
    context = {
        "form": form,
        "contrato": contrato,
        "resumo": resumo,
    }
    return render(request, "termo_aditivo/form.html", context)


# Termo Aditivo - Editar
@group_required("Admin", "Gerente")
def termo_aditivo_update(request, contrato_id, pk):
    """Edita um termo aditivo"""
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    termo = get_object_or_404(TermoAditivo, pk=pk, contrato=contrato)
    
    if request.method == "POST":
        form = TermoAditivoForm(request.POST, instance=termo, contrato_id=contrato_id, contrato=contrato)
        if form.is_valid():
            # Validar antes de salvar
            tipo = form.cleaned_data['tipo']
            meses_acrescimo = form.cleaned_data.get('meses_acrescimo', 0)
            valor_acrescimo = form.cleaned_data.get('valor_acrescimo', Decimal('0.00'))
            
            if tipo == TipoTermoAditivo.PRORROGACAO:
                # Validar limite de vigência (considerando outros aditivos, exceto este)
                meses_outros_aditivos = sum(
                    ta.meses_acrescimo or 0
                    for ta in contrato.termos_aditivos.exclude(pk=termo.pk).filter(tipo=TipoTermoAditivo.PRORROGACAO)
                )
                meses_totais = contrato.vigencia + meses_outros_aditivos + meses_acrescimo
                
                if contrato.regime_legal == RegimeLegal.LEI_14133:
                    limite = 120
                else:
                    limite = 60
                
                if meses_totais > limite:
                    messages.error(request, f"Vigência total ({meses_totais} meses) ultrapassa o limite legal de {limite} meses.")
                    return render(request, "termo_aditivo/form.html", {"form": form, "contrato": contrato, "termo": termo})
            
            elif tipo in [TipoTermoAditivo.VALOR, TipoTermoAditivo.REEQUILIBRIO]:
                if contrato.regime_legal == RegimeLegal.LEI_14133:
                    # Calcular valor atual sem este aditivo
                    valor_atual_sem_este = ContratoService.calcular_valor_atual(contrato)
                    valor_outros_aditivos = sum(
                        ta.valor_acrescimo or Decimal('0.00')
                        for ta in contrato.termos_aditivos.exclude(pk=termo.pk).filter(
                            tipo__in=[TipoTermoAditivo.VALOR, TipoTermoAditivo.REEQUILIBRIO]
                        )
                    )
                    valor_base = contrato.valor_inicial + valor_outros_aditivos
                    limite = valor_base * Decimal('0.25')
                    
                    if valor_acrescimo > limite:
                        messages.error(request, f"Valor de acréscimo (R$ {valor_acrescimo:,.2f}) ultrapassa o limite legal de 25% (R$ {limite:,.2f}).")
                        return render(request, "termo_aditivo/form.html", {"form": form, "contrato": contrato, "termo": termo})
            
            form.save()
            messages.success(request, "Termo aditivo atualizado com sucesso!")
            return redirect("gestao_contratos_detail", pk=contrato.pk)
    else:
        form = TermoAditivoForm(instance=termo, contrato_id=contrato_id, contrato=contrato)
    
    # Obter informações do contrato para validações
    resumo = ContratoService.obter_resumo_contrato(contrato)
    
    context = {
        "form": form,
        "contrato": contrato,
        "termo": termo,
        "resumo": resumo,
    }
    return render(request, "termo_aditivo/form.html", context)


# Termo Aditivo - Deletar
@group_required("Admin", "Gerente")
def termo_aditivo_delete(request, contrato_id, pk):
    """Exclui um termo aditivo"""
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    termo = get_object_or_404(TermoAditivo, pk=pk, contrato=contrato)
    
    if request.method == "POST":
        termo.delete()
        messages.success(request, "Termo aditivo excluído com sucesso!")
        return redirect("gestao_contratos_detail", pk=contrato.pk)
    
    context = {
        "contrato": contrato,
        "termo": termo,
    }
    return render(request, "termo_aditivo/confirm_delete.html", context)


# ==================== TIMESHEET ====================

@login_required
def timesheet_list(request):
    """Lista de tarefas do colaborador logado para lançamento de horas"""
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Você precisa ter um perfil de colaborador para acessar o Timesheet.")
        return redirect("dashboard")
    
    # Filtros
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    projeto_id = request.GET.get("projeto")
    sprint_id = request.GET.get("sprint")
    
    # Buscar tarefas onde o colaborador é responsável
    tarefas = Tarefa.objects.filter(responsavel=colaborador).select_related(
        "sprint", "sprint__projeto", "ordem_servico"
    ).order_by("-data_inicio_prevista")
    
    # Aplicar filtros
    if projeto_id:
        tarefas = tarefas.filter(sprint__projeto_id=projeto_id)
    if sprint_id:
        tarefas = tarefas.filter(sprint_id=sprint_id)
    
    # Buscar lançamentos do colaborador
    lancamentos = LancamentoHora.objects.filter(colaborador=colaborador).select_related("tarefa")
    
    if data_inicio:
        lancamentos = lancamentos.filter(data__gte=data_inicio)
    if data_fim:
        lancamentos = lancamentos.filter(data__lte=data_fim)
    
    # Calcular totais - somente horas faturáveis
    total_horas_lancadas = lancamentos.filter(faturavel=True).aggregate(
        total=Coalesce(Sum("horas_trabalhadas"), Value(Decimal("0.00")), output_field=DecimalField())
    )["total"]
    
    total_horas_planejadas = tarefas.aggregate(
        total=Coalesce(Sum("horas_planejadas"), Value(Decimal("0.00")), output_field=DecimalField())
    )["total"]
    
    # Horas restantes = Horas Planejadas - Horas Lançadas
    horas_restantes = total_horas_planejadas - total_horas_lancadas
    
    # Projetos do colaborador (para filtro)
    projetos = Projeto.objects.filter(
        sprints__tarefas__responsavel=colaborador
    ).distinct()
    
    # Sprints do colaborador (para filtro)
    sprints = Sprint.objects.filter(
        tarefas__responsavel=colaborador
    ).distinct()
    
    # Agrupar lançamentos por data para a visão semanal
    from collections import defaultdict
    lancamentos_por_data = defaultdict(list)
    for lancamento in lancamentos.order_by("-data"):
        lancamentos_por_data[lancamento.data].append(lancamento)
    
    # Todos os projetos para o dropdown da planilha
    todos_projetos = Projeto.objects.all().order_by("nome")
    
    context = {
        "colaborador": colaborador,
        "tarefas": tarefas,
        "lancamentos": lancamentos.order_by("-data", "-hora_inicio")[:50],
        "lancamentos_por_data": dict(lancamentos_por_data),
        "total_horas_lancadas": total_horas_lancadas,
        "total_horas_planejadas": total_horas_planejadas,
        "horas_restantes": horas_restantes,
        "projetos": projetos,
        "sprints": sprints,
        "todos_projetos": todos_projetos,
        "filtro_data_inicio": data_inicio,
        "filtro_data_fim": data_fim,
        "filtro_projeto": projeto_id,
        "filtro_sprint": sprint_id,
        "today": timezone.now().date(),
    }
    return render(request, "timesheet/list.html", context)


@login_required
def timesheet_lancar_horas(request, tarefa_id):
    """Lançar horas em uma tarefa específica"""
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Você precisa ter um perfil de colaborador.")
        return redirect("timesheet_list")
    
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id, responsavel=colaborador)
    
    if request.method == "POST":
        form = LancamentoHoraForm(request.POST)
        if form.is_valid():
            lancamento = form.save(commit=False)
            lancamento.tarefa = tarefa
            lancamento.colaborador = colaborador
            lancamento.save()
            messages.success(request, f"Horas lançadas com sucesso! ({lancamento.horas_trabalhadas}h)")
            return redirect("timesheet_list")
    else:
        form = LancamentoHoraForm(initial={"data": timezone.now().date()})
    
    # Lançamentos anteriores desta tarefa
    lancamentos_tarefa = tarefa.lancamentos_horas.filter(colaborador=colaborador).order_by("-data")
    total_lancado = lancamentos_tarefa.aggregate(
        total=Coalesce(Sum("horas_trabalhadas"), Value(Decimal("0.00")), output_field=DecimalField())
    )["total"]
    
    context = {
        "form": form,
        "tarefa": tarefa,
        "lancamentos_tarefa": lancamentos_tarefa,
        "total_lancado": total_lancado,
        "horas_restantes": (tarefa.horas_planejadas or Decimal("0.00")) - total_lancado,
        "today": timezone.now().date(),
    }
    return render(request, "timesheet/lancar_horas.html", context)


@login_required
def timesheet_editar_lancamento(request, lancamento_id):
    """Editar um lançamento de horas existente"""
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Você precisa ter um perfil de colaborador.")
        return redirect("timesheet_list")
    
    lancamento = get_object_or_404(LancamentoHora, pk=lancamento_id, colaborador=colaborador)
    
    if request.method == "POST":
        form = LancamentoHoraForm(request.POST, instance=lancamento)
        if form.is_valid():
            form.save()
            messages.success(request, "Lançamento atualizado com sucesso!")
            return redirect("timesheet_list")
    else:
        form = LancamentoHoraForm(instance=lancamento)
    
    context = {
        "form": form,
        "lancamento": lancamento,
        "tarefa": lancamento.tarefa,
    }
    return render(request, "timesheet/editar_lancamento.html", context)


@login_required
def timesheet_excluir_lancamento(request, lancamento_id):
    """Excluir um lançamento de horas"""
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Você precisa ter um perfil de colaborador.")
        return redirect("timesheet_list")
    
    lancamento = get_object_or_404(LancamentoHora, pk=lancamento_id, colaborador=colaborador)
    
    if request.method == "POST":
        lancamento.delete()
        messages.success(request, "Lançamento excluído com sucesso!")
        return redirect("timesheet_list")
    
    context = {
        "lancamento": lancamento,
    }
    return render(request, "timesheet/confirm_delete.html", context)


@login_required
def timesheet_tarefa_detail(request, tarefa_id):
    """Detalhes de uma tarefa no timesheet"""
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Você precisa ter um perfil de colaborador.")
        return redirect("timesheet_list")
    
    tarefa = get_object_or_404(Tarefa, pk=tarefa_id, responsavel=colaborador)
    
    # Lançamentos desta tarefa
    lancamentos = tarefa.lancamentos_horas.filter(colaborador=colaborador).order_by("-data")
    total_lancado = lancamentos.aggregate(
        total=Coalesce(Sum("horas_trabalhadas"), Value(Decimal("0.00")), output_field=DecimalField())
    )["total"]
    
    context = {
        "tarefa": tarefa,
        "lancamentos": lancamentos,
        "total_lancado": total_lancado,
        "horas_restantes": (tarefa.horas_planejadas or Decimal("0.00")) - total_lancado,
        "percentual_concluido": (total_lancado / tarefa.horas_planejadas * 100) if tarefa.horas_planejadas else 0,
    }
    return render(request, "timesheet/tarefa_detail.html", context)


@login_required
def timesheet_planilha(request):
    """Visão de planilha de horas do colaborador"""
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Você precisa ter um perfil de colaborador para acessar o Timesheet.")
        return redirect("dashboard")
    
    # Todos os projetos disponíveis
    todos_projetos = Projeto.objects.all().order_by("nome")
    
    # Lançamentos do colaborador
    lancamentos = LancamentoHora.objects.filter(
        colaborador=colaborador
    ).select_related(
        "tarefa", "tarefa__sprint", "tarefa__sprint__projeto"
    ).order_by("-data", "-criado_em")
    
    # Filtros
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    
    if data_inicio:
        lancamentos = lancamentos.filter(data__gte=data_inicio)
    if data_fim:
        lancamentos = lancamentos.filter(data__lte=data_fim)
    
    # Calcular totais
    total_horas = lancamentos.aggregate(
        total=Coalesce(Sum("horas_trabalhadas"), Value(Decimal("0.00")), output_field=DecimalField())
    )["total"]
    
    context = {
        "colaborador": colaborador,
        "todos_projetos": todos_projetos,
        "lancamentos": lancamentos[:100],
        "total_horas": total_horas,
        "filtro_data_inicio": data_inicio,
        "filtro_data_fim": data_fim,
    }
    return render(request, "timesheet/planilha.html", context)


@login_required
def timesheet_planilha_salvar(request):
    """Salvar lançamento via planilha (AJAX)"""
    import json
    
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)
    
    try:
        colaborador = request.user.colaborador
    except:
        return JsonResponse({"error": "Colaborador não encontrado"}, status=400)
    
    try:
        data = json.loads(request.body)
        
        tarefa_id = data.get("tarefa_id")
        data_lancamento = data.get("data")
        tempo_gasto = data.get("tempo_gasto")  # em formato HH:MM ou decimal
        descricao = data.get("descricao", "")
        faturavel = data.get("faturavel", True)
        lancamento_id = data.get("lancamento_id")  # Para edição
        
        # Validação: se faturável, tarefa é obrigatória
        if faturavel and not tarefa_id:
            return JsonResponse({"error": "Para lançamentos faturáveis, a tarefa é obrigatória"}, status=400)
        
        if not data_lancamento or not tempo_gasto:
            return JsonResponse({"error": "Data e tempo são obrigatórios"}, status=400)
        
        tarefa = None
        if tarefa_id:
            tarefa = get_object_or_404(Tarefa, pk=tarefa_id)
        
        # Converter data_lancamento para date se for string
        from datetime import datetime, time as dt_time, timedelta
        if isinstance(data_lancamento, str):
            data_lancamento = datetime.strptime(data_lancamento, "%Y-%m-%d").date()
        
        # Converter tempo_gasto para horas decimais
        if ":" in str(tempo_gasto):
            partes = tempo_gasto.split(":")
            horas = int(partes[0])
            minutos = int(partes[1]) if len(partes) > 1 else 0
            horas_decimal = Decimal(str(horas)) + Decimal(str(minutos)) / Decimal("60")
        else:
            horas_decimal = Decimal(str(tempo_gasto))
        
        # Calcular hora_inicio e hora_termino baseado no tempo gasto
        hora_inicio = dt_time(9, 0)  # Padrão: 09:00
        
        # Calcular hora de término
        minutos_totais = int(horas_decimal * 60)
        hora_termino_dt = datetime.combine(datetime.today(), hora_inicio) + timedelta(minutes=minutos_totais)
        hora_termino = hora_termino_dt.time()
        
        if lancamento_id:
            # Editar existente
            lancamento = get_object_or_404(LancamentoHora, pk=lancamento_id, colaborador=colaborador)
            lancamento.tarefa = tarefa
            lancamento.data = data_lancamento
            lancamento.hora_inicio = hora_inicio
            lancamento.hora_termino = hora_termino
            lancamento.descricao = descricao
            lancamento.faturavel = faturavel
            lancamento.save()
        else:
            # Criar novo
            lancamento = LancamentoHora.objects.create(
                tarefa=tarefa,
                colaborador=colaborador,
                data=data_lancamento,
                hora_inicio=hora_inicio,
                hora_termino=hora_termino,
                descricao=descricao,
                faturavel=faturavel
            )
        
        return JsonResponse({
            "success": True,
            "lancamento_id": lancamento.id,
            "horas_trabalhadas": str(lancamento.horas_trabalhadas)
        })
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
def timesheet_planilha_excluir(request, lancamento_id):
    """Excluir lançamento via planilha (AJAX)"""
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)
    
    try:
        colaborador = request.user.colaborador
    except:
        return JsonResponse({"error": "Colaborador não encontrado"}, status=400)
    
    lancamento = get_object_or_404(LancamentoHora, pk=lancamento_id, colaborador=colaborador)
    lancamento.delete()
    
    return JsonResponse({"success": True})


@login_required
def timesheet_exportar(request):
    """Exportar lançamentos do colaborador para XLSX"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from io import BytesIO
    
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Colaborador não encontrado")
        return redirect("timesheet_planilha")
    
    lancamentos = LancamentoHora.objects.filter(
        colaborador=colaborador
    ).select_related(
        "tarefa", "tarefa__sprint", "tarefa__sprint__projeto"
    ).order_by("-data")
    
    # Filtros
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    
    if data_inicio:
        lancamentos = lancamentos.filter(data__gte=data_inicio)
    if data_fim:
        lancamentos = lancamentos.filter(data__lte=data_fim)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos
    headers = ["Data", "Projeto", "Sprint", "Tarefa", "Descrição", "Tempo Gasto (h)", "Faturável"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Dados
    for row, lancamento in enumerate(lancamentos, 2):
        if lancamento.tarefa and lancamento.tarefa.sprint:
            projeto = lancamento.tarefa.sprint.projeto.nome
            sprint = lancamento.tarefa.sprint.nome
        else:
            projeto = "-"
            sprint = "-"
        
        ws.cell(row=row, column=1, value=lancamento.data.strftime("%d/%m/%Y")).border = thin_border
        ws.cell(row=row, column=2, value=projeto).border = thin_border
        ws.cell(row=row, column=3, value=sprint).border = thin_border
        ws.cell(row=row, column=4, value=lancamento.tarefa.titulo if lancamento.tarefa else "-").border = thin_border
        ws.cell(row=row, column=5, value=lancamento.descricao or "").border = thin_border
        ws.cell(row=row, column=6, value=float(lancamento.horas_trabalhadas)).border = thin_border
        ws.cell(row=row, column=7, value="Sim" if lancamento.faturavel else "Não").border = thin_border
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="timesheet_{colaborador.nome_completo}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


@login_required
def timesheet_importar(request):
    """Importar lançamentos de XLSX"""
    from openpyxl import load_workbook
    from datetime import datetime, time as dt_time, timedelta
    
    if request.method != "POST":
        return redirect("timesheet_planilha")
    
    try:
        colaborador = request.user.colaborador
    except:
        messages.error(request, "Colaborador não encontrado")
        return redirect("timesheet_planilha")
    
    arquivo = request.FILES.get("arquivo")
    if not arquivo:
        messages.error(request, "Nenhum arquivo selecionado")
        return redirect("timesheet_planilha")
    
    try:
        # Ler XLSX
        wb = load_workbook(arquivo)
        ws = wb.active
        
        importados = 0
        erros = []
        
        # Pular cabeçalho (linha 1)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Linha vazia
                    continue
                
                # Parsear data
                data_valor = row[0]
                if isinstance(data_valor, datetime):
                    data = data_valor.date()
                elif isinstance(data_valor, str):
                    data_str = data_valor.strip()
                    if "/" in data_str:
                        data = datetime.strptime(data_str, "%d/%m/%Y").date()
                    else:
                        data = datetime.strptime(data_str, "%Y-%m-%d").date()
                else:
                    data = data_valor
                
                # Buscar tarefa pelo nome
                tarefa_nome = str(row[3] or "").strip()  # Coluna D - Tarefa
                projeto_nome = str(row[1] or "").strip()  # Coluna B - Projeto
                
                if tarefa_nome:
                    tarefa = Tarefa.objects.filter(
                        titulo__icontains=tarefa_nome,
                        sprint__projeto__nome__icontains=projeto_nome
                    ).first()
                    
                    if not tarefa:
                        erros.append(f"Linha {i}: Tarefa '{tarefa_nome}' não encontrada")
                        continue
                else:
                    erros.append(f"Linha {i}: Tarefa não informada")
                    continue
                
                # Parsear tempo gasto
                tempo_valor = row[5]  # Coluna F - Tempo Gasto
                if isinstance(tempo_valor, (int, float)):
                    horas_decimal = Decimal(str(tempo_valor))
                else:
                    tempo_str = str(tempo_valor or "0").strip().replace(",", ".")
                    horas_decimal = Decimal(tempo_str)
                
                # Calcular hora_inicio e hora_termino
                hora_inicio = dt_time(9, 0)
                minutos_totais = int(horas_decimal * 60)
                hora_termino_dt = datetime.combine(datetime.today(), hora_inicio) + timedelta(minutes=minutos_totais)
                hora_termino = hora_termino_dt.time()
                
                # Descrição
                descricao = str(row[4] or "").strip()  # Coluna E - Descrição
                
                # Criar lançamento
                LancamentoHora.objects.create(
                    tarefa=tarefa,
                    colaborador=colaborador,
                    data=data,
                    hora_inicio=hora_inicio,
                    hora_termino=hora_termino,
                    descricao=descricao
                )
                importados += 1
                
            except Exception as e:
                erros.append(f"Linha {i}: {str(e)}")
        
        if importados > 0:
            messages.success(request, f"{importados} lançamento(s) importado(s) com sucesso!")
        if erros:
            messages.warning(request, f"Erros: {'; '.join(erros[:5])}")
            
    except Exception as e:
        messages.error(request, f"Erro ao processar arquivo: {str(e)}")
    
    return redirect("timesheet_planilha")


# APIs para carregar dados dinâmicos na planilha
@login_required
@require_GET
def api_sprints_por_projeto(request):
    """Retorna sprints de um projeto"""
    projeto_id = request.GET.get("projeto_id")
    if not projeto_id:
        return JsonResponse({"sprints": []}, safe=False)
    
    sprints = Sprint.objects.filter(projeto_id=projeto_id).values("id", "nome")
    return JsonResponse({"sprints": list(sprints)}, safe=False)


@require_GET
def api_projetos_por_contrato(request):
    """Retorna projetos de um contrato"""
    contrato_id = request.GET.get("contrato_id")
    if not contrato_id:
        return JsonResponse({"projetos": []}, safe=False)
    
    projetos = Projeto.objects.filter(contrato_id=contrato_id).values("id", "nome")
    return JsonResponse({"projetos": list(projetos)}, safe=False)


@require_GET
def api_ordens_servico_por_contrato(request):
    """Retorna ordens de serviço de um contrato"""
    contrato_id = request.GET.get("contrato_id")
    projeto_id = request.GET.get("projeto_id")
    
    if not contrato_id:
        return JsonResponse({"ordens": []}, safe=False)
    
    os_queryset = OrdemServico.objects.filter(contrato_id=contrato_id)
    
    # Se projeto_id for fornecido, filtrar também por projeto (via Projeto vinculado à OS)
    if projeto_id:
        # Projeto está relacionado à OS via Projeto.ordem_servico (OneToOneField, related_name="projeto")
        os_queryset = os_queryset.filter(projeto__id=projeto_id)
    
    ordens = os_queryset.values("id", "numero_os")
    return JsonResponse({"ordens": list(ordens)}, safe=False)


@login_required
def api_tarefas_por_sprint(request):
    """Retorna tarefas de uma sprint"""
    sprint_id = request.GET.get("sprint_id")
    if not sprint_id:
        return JsonResponse([], safe=False)
    
    try:
        colaborador = request.user.colaborador
        # Retorna apenas tarefas do colaborador
        tarefas = Tarefa.objects.filter(
            sprint_id=sprint_id,
            responsavel=colaborador
        ).values("id", "titulo")
    except:
        tarefas = Tarefa.objects.filter(sprint_id=sprint_id).values("id", "titulo")
    
    return JsonResponse(list(tarefas), safe=False)


# =============================================================================
# MÓDULO DE ANÁLISE DE CONTRATOS COM IA
# =============================================================================

@group_required("Admin", "Gerente")
def documento_contrato_list(request):
    """Lista análises de contrato"""
    analises = AnaliseContrato.objects.all().order_by('-criado_em')
    
    # Filtros
    status = request.GET.get('status')
    
    if status:
        analises = analises.filter(status=status)
    
    # Paginação
    paginator = Paginator(analises, 20)
    page = request.GET.get('page', 1)
    analises = paginator.get_page(page)
    
    context = {
        'analises': analises,
        'status_choices': AnaliseContrato.STATUS_CHOICES,
        'filtro_status': status,
    }
    return render(request, 'ia_contratos/list.html', context)


@group_required("Admin", "Gerente")
def documento_contrato_upload(request):
    """Upload múltiplo de documentos para análise"""
    if request.method == 'POST':
        # Validação manual de arquivos múltiplos ANTES de validar o formulário
        arquivos = request.FILES.getlist('arquivos')
        erro_arquivos = None
        
        if not arquivos:
            erro_arquivos = 'Selecione pelo menos um arquivo.'
        else:
            # Valida cada arquivo
            for arquivo in arquivos:
                ext = arquivo.name.split('.')[-1].lower()
                if ext not in ['pdf', 'doc', 'docx']:
                    erro_arquivos = f'Formato não suportado: {arquivo.name}. Use PDF ou Word (.doc, .docx)'
                    break
                if arquivo.size > 50 * 1024 * 1024:
                    erro_arquivos = f'Arquivo muito grande: {arquivo.name}. Limite: 50MB'
                    break
        
        # Cria o formulário - sempre passa FILES mesmo se vazio para evitar erro de validação
        form = AnaliseContratoForm(request.POST, request.FILES)
        
        # Remove a validação do campo arquivos do formulário, pois já validamos manualmente
        # O Django valida o campo arquivos mesmo quando required=False se não houver FILES
        if 'arquivos' in form.errors:
            # Remove o erro de "Nenhum arquivo enviado" se tivermos arquivos ou se for erro de validação do Django
            if arquivos or 'Nenhum arquivo enviado' in str(form.errors.get('arquivos', [])):
                form.errors.pop('arquivos', None)
        
        # Se houver erro nos arquivos, adiciona ao formulário
        if erro_arquivos:
            form.add_error('arquivos', erro_arquivos)
        
        # Valida o formulário (nome e número de contrato)
        # Importante: validar mesmo se houver erro_arquivos para mostrar todos os erros
        form_valid = form.is_valid()
        
        if form_valid and not erro_arquivos:
            nome_analise = form.cleaned_data['nome_analise']
            numero_contrato_busca = form.cleaned_data.get('numero_contrato_busca', '').strip()
            
            # Verifica se o contrato já existe
            contrato_existente = None
            cliente_existente = None
            if numero_contrato_busca:
                from .services.contract_ai_service import ContractAIService
                contrato_existente, cliente_existente = ContractAIService.verificar_contrato_existente(
                    numero_contrato=numero_contrato_busca
                )
            
            # Cria a análise
            analise = AnaliseContrato.objects.create(
                nome=nome_analise,
                criado_por=request.user,
                contrato_gerado=contrato_existente,
                cliente_gerado=cliente_existente,
            )
            
            # Cria os documentos
            tipos_por_nome = {
                'contrato': 'contrato',
                'edital': 'edital',
                'etp': 'etp',
                'tr': 'tr',
                'arp': 'arp',
                'proposta': 'proposta',
                'termo': 'termo_aditivo',
                'aditivo': 'termo_aditivo',
            }
            
            for arquivo in arquivos:
                nome_arquivo = arquivo.name.lower()
                tipo_doc = 'outro'
                
                # Tenta identificar o tipo pelo nome do arquivo
                for palavra_chave, tipo in tipos_por_nome.items():
                    if palavra_chave in nome_arquivo:
                        tipo_doc = tipo
                        break
                
                DocumentoContrato.objects.create(
                    analise=analise,
                    nome=arquivo.name,
                    tipo_documento=tipo_doc,
                    arquivo=arquivo,
                )
            
            messages.success(request, f'Análise "{nome_analise}" criada com {len(arquivos)} documento(s)!')
            if contrato_existente:
                messages.info(request, f'Vinculada ao contrato existente: {contrato_existente.numero_contrato}')
            
            # Executa análise automaticamente após criar
            try:
                import logging
                import os
                from django.conf import settings
                from decouple import config
                logger = logging.getLogger(__name__)
                from .services.contract_ai_service import ContractAIService
                
                # Verifica se a chave da API está configurada
                # Tenta ler do settings primeiro, depois do .env usando decouple
                api_key = getattr(settings, 'OPENAI_API_KEY', None)
                if not api_key:
                    api_key = config('OPENAI_API_KEY', default=None)
                if not api_key:
                    api_key = os.environ.get('OPENAI_API_KEY')
                
                if not api_key:
                    messages.warning(
                        request, 
                        'Análise criada com sucesso! Configure a chave da API (OPENAI_API_KEY) no arquivo .env para executar a análise automaticamente. '
                        'Você pode executar a análise manualmente na página de detalhes.'
                    )
                else:
                    service = ContractAIService()
                    dados = service.process_multiple_documents(analise)
                    messages.success(request, f'Análise concluída! {len(analise.documentos.all())} documento(s) processado(s).')
            except ValueError as e:
                # Erro de configuração (chave não configurada)
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Chave da API não configurada: {str(e)}")
                messages.warning(
                    request, 
                    f'Análise criada com sucesso! {str(e)}. '
                    'Configure a chave da API (OPENAI_API_KEY) no arquivo .env e execute a análise manualmente na página de detalhes.'
                )
            except Exception as e:
                import traceback
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Erro na análise automática: {traceback.format_exc()}")
                messages.warning(request, f'Análise criada, mas houve erro ao processar: {str(e)}. Você pode tentar novamente na página de detalhes.')
            
            return redirect('documento_contrato_detail', pk=analise.pk)
        else:
            # Se o formulário não for válido, adiciona mensagem de erro
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Formulário inválido. Erros: {form.errors}, erro_arquivos: {erro_arquivos}")
            
            if not form.is_valid() or erro_arquivos:
                # Adiciona mensagem de erro genérica
                if erro_arquivos:
                    messages.error(request, erro_arquivos)
                elif form.errors:
                    # Adiciona mensagem de erro específica para cada campo
                    for field, errors in form.errors.items():
                        for error in errors:
                            if field in form.fields:
                                messages.error(request, f'{form.fields[field].label}: {error}')
                            else:
                                messages.error(request, f'{error}')
                else:
                    messages.error(request, 'Erro ao criar análise. Verifique os campos.')
    else:
        form = AnaliseContratoForm()
    
    context = {'form': form}
    return render(request, 'ia_contratos/upload.html', context)


@group_required("Admin", "Gerente")
def documento_contrato_detail(request, pk):
    """Detalhes da análise e resultados"""
    analise = get_object_or_404(AnaliseContrato, pk=pk)
    
    context = {
        'analise': analise,
        'documentos': analise.documentos.all(),
        'dados': analise.dados_extraidos or {},
    }
    return render(request, 'ia_contratos/detail.html', context)


@group_required("Admin", "Gerente")
def documento_contrato_analisar(request, pk):
    """Executa análise de todos os documentos com IA"""
    analise = get_object_or_404(AnaliseContrato, pk=pk)
    
    try:
        from .services.contract_ai_service import ContractAIService
        
        service = ContractAIService()
        dados = service.process_multiple_documents(analise)
        
        messages.success(request, f'Análise concluída! {len(analise.documentos.all())} documento(s) processado(s).')
        
    except Exception as e:
        import traceback
        print(f"Erro na análise: {traceback.format_exc()}")
        messages.error(request, f'Erro na análise: {str(e)}')
    
    return redirect('documento_contrato_detail', pk=pk)


@group_required("Admin", "Gerente")
def documento_contrato_criar_registros(request, pk):
    """Cria registros (Cliente, Contrato, Itens) a partir dos dados extraídos"""
    analise = get_object_or_404(AnaliseContrato, pk=pk)
    
    if not analise.dados_extraidos:
        messages.error(request, 'Análise ainda não foi concluída.')
        return redirect('documento_contrato_detail', pk=pk)
    
    try:
        from .services.contract_ai_service import ContractAIService
        
        dados = analise.dados_extraidos
        
        # Usa cliente/contrato existente se já vinculados
        cliente_existente = analise.cliente_gerado
        contrato_existente = analise.contrato_gerado
        
        # Criar ou usar Cliente existente
        cliente = ContractAIService.create_cliente_from_data(dados, request.user, cliente_existente)
        if cliente:
            analise.cliente_gerado = cliente
            if cliente_existente:
                messages.info(request, f'Usando cliente existente: "{cliente.nome_razao_social}"')
            else:
                messages.success(request, f'Cliente "{cliente.nome_razao_social}" criado.')
        
        # Criar ou usar Contrato existente
        if cliente:
            contrato = ContractAIService.create_contrato_from_data(dados, cliente, request.user, contrato_existente)
            if contrato:
                analise.contrato_gerado = contrato
                if contrato_existente:
                    messages.info(request, f'Usando contrato existente: "{contrato.numero_contrato}"')
                else:
                    messages.success(request, f'Contrato "{contrato.numero_contrato}" criado.')
                
                # Criar Itens (apenas se for contrato novo)
                if not contrato_existente:
                    itens = ContractAIService.create_itens_from_data(dados, contrato)
                    if itens:
                        messages.success(request, f'{len(itens)} item(ns) de contrato criado(s).')
                    
                    # Criar SLAs (apenas se for contrato novo)
                    slas = ContractAIService.create_slas_from_data(dados, contrato)
                    if slas:
                        messages.success(request, f'{len(slas)} SLA(s) criado(s).')
                    
                    # Criar primeiro registro do Diário de Bordo (apenas se for contrato novo)
                    diario = ContractAIService.create_diario_bordo_from_data(dados, contrato, request.user)
                    if diario:
                        messages.success(request, 'Registro inicial do Diário de Bordo criado.')
                else:
                    messages.info(request, 'Itens e SLAs não foram criados pois o contrato já existe.')
        
        analise.status = 'revisado'
        analise.save()
        
    except Exception as e:
        import traceback
        print(f"Erro ao criar registros: {traceback.format_exc()}")
        messages.error(request, f'Erro ao criar registros: {str(e)}')
    
    return redirect('documento_contrato_detail', pk=pk)


@group_required("Admin", "Gerente")
def documento_contrato_delete(request, pk):
    """Exclui análise e todos os documentos"""
    analise = get_object_or_404(AnaliseContrato, pk=pk)
    
    if request.method == 'POST':
        nome = analise.nome
        analise.delete()  # Isso também deleta os documentos relacionados (CASCADE)
        messages.success(request, f'Análise "{nome}" excluída.')
        return redirect('documento_contrato_list')
    
    return render(request, 'ia_contratos/confirm_delete.html', {'analise': analise})


@group_required("Admin", "Gerente", "Leitor")
def documento_contrato_download(request, pk):
    """Faz download do arquivo do documento"""
    from django.http import FileResponse, Http404
    from django.conf import settings
    from django.utils.text import get_valid_filename
    import os
    
    documento = get_object_or_404(DocumentoContrato, pk=pk)
    
    if not documento.arquivo:
        messages.error(request, 'Arquivo não encontrado.')
        if documento.analise:
            return redirect('documento_contrato_detail', pk=documento.analise.pk)
        return redirect('documento_contrato_list')
    
    try:
        # Verifica se o arquivo existe
        file_path = documento.arquivo.path
        if not os.path.exists(file_path):
            messages.error(request, 'Arquivo não encontrado no sistema de arquivos.')
            if documento.analise:
                return redirect('documento_contrato_detail', pk=documento.analise.pk)
            return redirect('documento_contrato_list')
        
        # Obtém o nome do arquivo original
        original_filename = os.path.basename(documento.arquivo.name)
        # Se não tiver extensão, tenta usar o nome do documento
        if not original_filename or '.' not in original_filename:
            filename = documento.nome or 'documento'
            # Adiciona a extensão do arquivo se não tiver
            ext = os.path.splitext(file_path)[1]
            if ext:
                filename = f"{filename}{ext}"
        else:
            filename = original_filename
        
        # Sanitiza o nome do arquivo para evitar problemas
        filename = get_valid_filename(filename)
        
        # Determina o content-type baseado na extensão
        ext = os.path.splitext(filename)[1].lower()
        content_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        }
        content_type = content_types.get(ext, 'application/octet-stream')
        
        # Retorna o arquivo para download
        file_handle = open(file_path, 'rb')
        response = FileResponse(file_handle, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Http404:
        raise
    except Exception as e:
        import traceback
        print(f"Erro ao fazer download do arquivo: {traceback.format_exc()}")
        messages.error(request, f'Erro ao fazer download: {str(e)}')
        if documento.analise:
            return redirect('documento_contrato_detail', pk=documento.analise.pk)
        return redirect('documento_contrato_list')


@group_required("Admin", "Gerente")
def plano_trabalho_gerar(request, projeto_id):
    """Gera plano de trabalho completo usando IA para um projeto"""
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    
    # Verifica se já existe plano
    if hasattr(projeto, 'plano_trabalho'):
        messages.info(request, 'Este projeto já possui um plano de trabalho.')
        return redirect('plano_trabalho_detail', pk=projeto.plano_trabalho.pk)
    
    # Busca análise relacionada ao contrato
    contrato = projeto.contrato
    analise = contrato.analises_origem.first()
    if not analise or not analise.texto_consolidado:
        messages.error(request, 'Análise não encontrada ou não foi concluída. É necessário ter uma análise de contrato concluída.')
        return redirect('projeto_detail', pk=projeto_id)
    
    try:
        from .services.contract_ai_service import ContractAIService
        
        dados_plano = ContractAIService.gerar_plano_trabalho_completo(
            projeto, 
            analise.texto_consolidado
        )
        
        plano = ContractAIService.criar_plano_trabalho(projeto, dados_plano, request.user)
        
        messages.success(request, 'Plano de trabalho gerado com sucesso! Revise e aprove para aplicar ao projeto.')
        return redirect('plano_trabalho_detail', pk=plano.pk)
        
    except Exception as e:
        import traceback
        print(f"Erro ao gerar plano: {traceback.format_exc()}")
        messages.error(request, f'Erro ao gerar plano de trabalho: {str(e)}')
        return redirect('projeto_detail', pk=projeto_id)


@group_required("Admin", "Gerente")
def plano_trabalho_detail(request, pk):
    """Detalhes do plano de trabalho"""
    plano = get_object_or_404(PlanoTrabalho, pk=pk)
    
    # Buscar dados do contrato através do projeto
    contrato = plano.projeto.contrato if plano.projeto else None
    
    context = {
        'plano': plano,
        'projeto': plano.projeto,
        'contrato': contrato,
        'slas_importantes': contrato.slas_importantes.all() if contrato else [],
        'clausulas_criticas': contrato.clausulas_criticas.all() if contrato else [],
        'matriz_raci': contrato.matriz_raci.all() if contrato else [],
    }
    return render(request, 'ia_contratos/plano_detail.html', context)


@group_required("Admin", "Gerente")
def plano_trabalho_aprovar(request, pk):
    """Aprova o plano de trabalho - o projeto já deve existir"""
    plano = get_object_or_404(PlanoTrabalho, pk=pk)
    
    if request.method == 'POST':
        observacoes = request.POST.get('observacoes', '')
        
        try:
            import logging
            
            logger = logging.getLogger(__name__)
            
            # Verifica se o projeto existe
            if not plano.projeto:
                messages.error(request, 'O plano de trabalho deve estar vinculado a um projeto antes de ser aprovado.')
                return redirect('plano_trabalho_detail', pk=pk)
            
            # Aprova o plano
            plano.aprovar(request.user, observacoes)
            logger.info(f"Plano {plano.pk} aprovado por {request.user}")
            
            projeto = plano.projeto
            
            # Criar sprints e tarefas automaticamente após aprovação
            try:
                from .services.contract_ai_service import ContractAIService
                ContractAIService.criar_projeto_sprints_tarefas(plano, request.user)
                logger.info(f"Sprints e tarefas criadas automaticamente para o projeto {projeto.pk}")
                messages.success(
                    request, 
                    f'Plano aprovado! Sprints e tarefas foram criadas automaticamente para o projeto "{projeto.nome}".'
                )
            except Exception as e:
                logger.error(f"Erro ao criar sprints e tarefas: {e}")
                import traceback
                logger.error(traceback.format_exc())
                messages.warning(
                    request,
                    f'Plano aprovado, mas houve um erro ao criar sprints e tarefas automaticamente: {str(e)}. '
                    f'Você pode criá-las manualmente.'
                )
            
            return redirect('projeto_detail', pk=projeto.pk)
            
        except Exception as e:
            import traceback
            logger = logging.getLogger(__name__)
            error_trace = traceback.format_exc()
            logger.error(f"Erro ao aprovar plano {plano.pk}: {error_trace}")
            print(f"Erro ao aprovar plano: {error_trace}")
            messages.error(request, f'Erro ao aprovar plano: {str(e)}. Verifique os logs para mais detalhes.')
    
    return redirect('plano_trabalho_detail', pk=pk)


@group_required("Admin", "Gerente")
def plano_trabalho_rejeitar(request, pk):
    """Rejeita o plano de trabalho"""
    plano = get_object_or_404(PlanoTrabalho, pk=pk)
    
    if request.method == 'POST':
        observacoes = request.POST.get('observacoes', '')
        plano.status = 'rejeitado'
        plano.observacoes_aprovacao = observacoes
        plano.save()
        
        messages.info(request, 'Plano de trabalho rejeitado.')
    
    return redirect('plano_trabalho_detail', pk=pk)


@group_required("Admin", "Gerente")
def plano_trabalho_update(request, pk):
    """Edita o plano de trabalho"""
    import json
    plano = get_object_or_404(PlanoTrabalho, pk=pk)
    
    if request.method == 'POST':
        form = PlanoTrabalhoForm(request.POST, instance=plano)
        
        # Processar dados JSON dos campos dinâmicos
        if 'pontos_atencao_json' in request.POST:
            try:
                pontos_atencao = json.loads(request.POST['pontos_atencao_json'])
                plano.pontos_atencao = pontos_atencao
            except json.JSONDecodeError:
                pass
        
        if 'processo_execucao_json' in request.POST:
            try:
                processo_execucao = json.loads(request.POST['processo_execucao_json'])
                plano.processo_execucao = processo_execucao
            except json.JSONDecodeError:
                pass
        
        if 'cronograma_detalhado_json' in request.POST:
            try:
                cronograma_detalhado = json.loads(request.POST['cronograma_detalhado_json'])
                plano.cronograma_detalhado = cronograma_detalhado
            except json.JSONDecodeError:
                pass
        
        if 'plano_comunicacao_json' in request.POST:
            try:
                plano_comunicacao = json.loads(request.POST['plano_comunicacao_json'])
                plano.plano_comunicacao = plano_comunicacao
            except json.JSONDecodeError:
                pass
        
        if form.is_valid():
            # Salvar campos do formulário
            plano.resumo_contrato = form.cleaned_data.get('resumo_contrato', plano.resumo_contrato)
            plano.data_inicio_prevista = form.cleaned_data.get('data_inicio_prevista', plano.data_inicio_prevista)
            plano.data_fim_prevista = form.cleaned_data.get('data_fim_prevista', plano.data_fim_prevista)
            plano.template_status_report = form.cleaned_data.get('template_status_report', plano.template_status_report)
            plano.frequencia_status_report = form.cleaned_data.get('frequencia_status_report', plano.frequencia_status_report)
            
            # Detecta se o processo_execucao foi alterado
            processo_execucao_alterado = 'processo_execucao_json' in request.POST
            
            plano.save()
            
            # Se o processo de execução foi alterado e o plano tem projeto, sincroniza sprints e tarefas
            # Funciona para planos aprovados ou não (se já tiver projeto criado)
            if processo_execucao_alterado:
                if plano.projeto:
                    try:
                        from .services.contract_ai_service import ContractAIService
                        import logging
                        
                        logger = logging.getLogger(__name__)
                        logger.info(f"Sincronizando projeto após alteração do processo de execução do plano {plano.pk}")
                        
                        # Força sincronização mesmo se o plano não estiver aprovado (se já tiver projeto)
                        projeto = ContractAIService.sincronizar_projeto_com_plano(plano, request.user, forcar_sincronizacao=True)
                        
                        if projeto:
                            sprints_count = projeto.sprints.count()
                            tarefas_count = sum(sprint.tarefas.count() for sprint in projeto.sprints.all())
                            messages.success(
                                request, 
                                f'Plano de trabalho atualizado e sincronizado com o projeto! '
                                f'Projeto possui {sprints_count} sprint(s) e {tarefas_count} tarefa(s).'
                            )
                        else:
                            messages.success(request, 'Plano de trabalho atualizado com sucesso!')
                    except Exception as e:
                        import traceback
                        logger = logging.getLogger(__name__)
                        error_trace = traceback.format_exc()
                        logger.error(f"Erro ao sincronizar projeto após edição do plano {plano.pk}: {error_trace}")
                        messages.warning(
                            request, 
                            f'Plano atualizado, mas houve erro ao sincronizar com o projeto: {str(e)}. '
                            f'Verifique os logs para mais detalhes.'
                        )
                else:
                    messages.success(
                        request, 
                        'Plano de trabalho atualizado com sucesso! '
                        'As alterações no processo de execução serão aplicadas quando o plano for aprovado.'
                    )
            else:
                messages.success(request, 'Plano de trabalho atualizado com sucesso!')
            
            return redirect('plano_trabalho_detail', pk=plano.pk)
        else:
            messages.error(request, 'Erro ao atualizar plano. Verifique os campos.')
    else:
        # Ao editar, o ModelForm já carrega automaticamente as datas salvas via instance=plano
        # Se não houver datas salvas, preenche com as do contrato como padrão
        form = PlanoTrabalhoForm(instance=plano)
        
        # Se o plano não tiver datas salvas, preenche com as do contrato através do projeto
        if not plano.data_inicio_prevista and plano.projeto:
            form.fields['data_inicio_prevista'].initial = plano.projeto.contrato.data_assinatura
        if not plano.data_fim_prevista and plano.projeto:
            form.fields['data_fim_prevista'].initial = plano.projeto.contrato.data_fim_atual or plano.projeto.contrato.data_fim
    
    context = {
        'form': form,
        'plano': plano,
    }
    return render(request, 'ia_contratos/plano_form.html', context)


@group_required("Admin", "Gerente", "Leitor")
def plano_trabalho_exportar_pdf(request, pk):
    """Exporta o plano de trabalho em PDF usando reportlab"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from io import BytesIO
    import json
    
    plano = get_object_or_404(PlanoTrabalho.objects.select_related('projeto__contrato__cliente'), pk=pk)
    
    # Obter contrato através do projeto
    if not plano.projeto:
        messages.error(request, 'O plano de trabalho deve estar vinculado a um projeto.')
        return redirect('plano_trabalho_detail', pk=pk)
    
    contrato = plano.projeto.contrato
    projeto = plano.projeto
    
    # Criar resposta PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"plano_trabalho_{projeto.nome.replace(' ', '_').replace('/', '_')}_{contrato.numero_contrato.replace('/', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    try:
        # Criar documento PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        # Cores Alltech (valores hex para uso em strings HTML)
        cor_laranja_hex = '#FF6B35'
        cor_azul_escuro_hex = '#1e3a5f'
        cor_cinza_hex = '#9ca3af'
        cor_branco_hex = '#FFFFFF'
        
        # Cores Alltech (objetos Color para uso em estilos)
        cor_laranja = colors.HexColor(cor_laranja_hex)
        cor_azul_escuro = colors.HexColor(cor_azul_escuro_hex)
        cor_cinza = colors.HexColor(cor_cinza_hex)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=cor_azul_escuro,
            spaceAfter=12,
            borderWidth=0,
            borderColor=cor_laranja,
            borderPadding=(0, 0, 5, 0),
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=cor_azul_escuro,
            spaceAfter=10,
        )
        
        # Estilo para células de tabela com quebra de linha
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            spaceBefore=0,
            spaceAfter=0,
        )
        
        # Estilo para cabeçalhos de tabela
        header_cell_style = ParagraphStyle(
            'HeaderCellStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            alignment=TA_LEFT,
            textColor=colors.white,
            spaceBefore=0,
            spaceAfter=0,
        )
        normal_style = styles['Normal']
        normal_style.fontSize = 10
        normal_style.leading = 14
        
        # Conteúdo do PDF
        story = []
        
        # Cabeçalho com logo Alltech
        header_data = [
            [Paragraph(f'<font color="{cor_laranja_hex}"><b>ALL</b></font><font color="{cor_azul_escuro_hex}"><b>TECH</b></font><br/><font size="9" color="{cor_cinza_hex}">Soluções em Tecnologia</font>', 
                      ParagraphStyle('Header', parent=styles['Normal'], 
                                   fontSize=28, textColor=colors.black,
                                   alignment=TA_CENTER, spaceAfter=5))],
            [Paragraph('<b>PLANO DE TRABALHO</b>', 
                      ParagraphStyle('HeaderTitle', parent=styles['Normal'],
                                   fontSize=18, textColor=colors.white,
                                   alignment=TA_CENTER, backColor=cor_azul_escuro,
                                   spaceBefore=10, spaceAfter=10))],
            [Paragraph(f'<b>Projeto:</b> {projeto.nome}<br/>'
                      f'<b>Contrato:</b> {contrato.numero_contrato}<br/>'
                      f'<b>Cliente:</b> {contrato.cliente.nome_razao_social}',
                      ParagraphStyle('HeaderInfo', parent=styles['Normal'],
                                   fontSize=10, textColor=colors.white,
                                   alignment=TA_CENTER, spaceAfter=5))],
        ]
        header_table = Table(header_data, colWidths=[16*cm])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 1), (0, 1), cor_azul_escuro),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Informações do Contrato
        info_data = [
            [
                Paragraph('<b>Projeto</b>', header_cell_style),
                Paragraph(str(projeto.nome), cell_style),
                Paragraph('<b>Data de Início</b>', header_cell_style),
                Paragraph(plano.data_inicio_prevista.strftime('%d/%m/%Y'), cell_style)
            ],
            [
                Paragraph('<b>Número do Contrato</b>', header_cell_style),
                Paragraph(str(contrato.numero_contrato), cell_style),
                Paragraph('<b>Data de Término</b>', header_cell_style),
                Paragraph(plano.data_fim_prevista.strftime('%d/%m/%Y'), cell_style)
            ],
            [
                Paragraph('<b>Cliente</b>', header_cell_style),
                Paragraph(contrato.cliente.nome_razao_social, cell_style),
                Paragraph('<b>Status</b>', header_cell_style),
                Paragraph(plano.get_status_display(), cell_style)
            ],
            [
                Paragraph('<b>CNPJ/CPF</b>', header_cell_style),
                Paragraph(str(contrato.cliente.cnpj_cpf or ''), cell_style),
                Paragraph('<b>Fornecedor</b>', header_cell_style),
                Paragraph(str(plano.fornecedor or 'Não definido'), cell_style)
            ],
        ]
        info_table = Table(info_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), cor_azul_escuro),
            ('BACKGROUND', (2, 0), (2, 0), cor_azul_escuro),
            ('BACKGROUND', (0, 1), (0, 1), cor_azul_escuro),
            ('BACKGROUND', (2, 1), (2, 1), cor_azul_escuro),
            ('BACKGROUND', (0, 2), (0, 2), cor_azul_escuro),
            ('BACKGROUND', (2, 2), (2, 2), cor_azul_escuro),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(Paragraph('<b>Informações do Contrato</b>', heading_style))
        story.append(info_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Resumo do Contrato
        story.append(Paragraph('<b>Resumo do Contrato</b>', title_style))
        story.append(Paragraph(plano.resumo_contrato.replace('\n', '<br/>'), normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # Pontos de Atenção
        if plano.pontos_atencao:
            story.append(PageBreak())
            story.append(Paragraph('<b>Pontos de Atenção</b>', title_style))
            
            pontos_data = [[
                Paragraph('<b>Título</b>', header_cell_style),
                Paragraph('<b>Descrição</b>', header_cell_style),
                Paragraph('<b>Prioridade</b>', header_cell_style),
                Paragraph('<b>Ação Recomendada</b>', header_cell_style)
            ]]
            for ponto in plano.pontos_atencao:
                pontos_data.append([
                    Paragraph(str(ponto.get('titulo', '-')), cell_style),
                    Paragraph(str(ponto.get('descricao', '-')), cell_style),
                    Paragraph(str(ponto.get('prioridade', 'media').upper()), cell_style),
                    Paragraph(str(ponto.get('acao_recomendada', '-')), cell_style),
                ])
            
            pontos_table = Table(pontos_data, colWidths=[3.5*cm, 5*cm, 2.5*cm, 5*cm], repeatRows=1)
            pontos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), cor_azul_escuro),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(pontos_table)
            story.append(Spacer(1, 0.3*cm))
        
        # Cláusulas Críticas
        clausulas_criticas = contrato.clausulas_criticas.all()
        if clausulas_criticas:
            story.append(PageBreak())
            story.append(Paragraph('<b>Cláusulas Críticas</b>', title_style))
            
            clausulas_data = [[
                Paragraph('<b>Título</b>', header_cell_style),
                Paragraph('<b>Nº Cláusula</b>', header_cell_style),
                Paragraph('<b>Descrição</b>', header_cell_style),
                Paragraph('<b>Impacto</b>', header_cell_style),
                Paragraph('<b>Ação Necessária</b>', header_cell_style)
            ]]
            for clausula in clausulas_criticas:
                clausulas_data.append([
                    Paragraph(str(clausula.titulo), cell_style),
                    Paragraph(str(clausula.numero_clausula or '-'), cell_style),
                    Paragraph(str(clausula.descricao), cell_style),
                    Paragraph(str(clausula.get_impacto_display().upper()), cell_style),
                    Paragraph(str(clausula.acao_necessaria), cell_style),
                ])
            
            clausulas_table = Table(clausulas_data, colWidths=[3*cm, 2*cm, 4*cm, 2*cm, 5*cm], repeatRows=1)
            clausulas_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), cor_azul_escuro),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(clausulas_table)
            story.append(Spacer(1, 0.3*cm))
        
        # Processo de Execução
        if plano.processo_execucao:
            story.append(PageBreak())
            story.append(Paragraph('<b>Processo de Execução</b>', title_style))
            
            for etapa in plano.processo_execucao:
                etapa_nome = etapa.get('nome') or etapa.get('etapa') or 'Etapa'
                fase = etapa.get('fase', '')
                story.append(Paragraph(f'<b>{etapa_nome}</b> - {fase}', heading_style))
                story.append(Paragraph(f'<b>Descrição:</b> {etapa.get("descricao", "-")}', normal_style))
                story.append(Paragraph(f'<b>Duração:</b> {etapa.get("duracao_dias", "-")} dias', normal_style))
                
                if etapa.get('entregaveis'):
                    story.append(Paragraph('<b>Entregáveis:</b>', normal_style))
                    for entregavel in etapa.get('entregaveis', []):
                        story.append(Paragraph(f'• {entregavel}', normal_style))
                
                story.append(Spacer(1, 0.3*cm))
        
        # SLAs Importantes
        slas_importantes = contrato.slas_importantes.all()
        if slas_importantes:
            story.append(PageBreak())
            story.append(Paragraph('<b>SLAs Importantes</b>', title_style))
            
            slas_data = [[
                Paragraph('<b>Nome</b>', header_cell_style),
                Paragraph('<b>Descrição</b>', header_cell_style),
                Paragraph('<b>Resposta</b>', header_cell_style),
                Paragraph('<b>Solução</b>', header_cell_style),
                Paragraph('<b>Prioridade</b>', header_cell_style)
            ]]
            for sla in slas_importantes:
                slas_data.append([
                    Paragraph(str(sla.nome), cell_style),
                    Paragraph(str(sla.descricao), cell_style),
                    Paragraph(f'{sla.tempo_resposta_horas}h', cell_style),
                    Paragraph(f'{sla.tempo_solucao_horas}h', cell_style),
                    Paragraph(str(sla.get_prioridade_display().upper()), cell_style),
                ])
            
            slas_table = Table(slas_data, colWidths=[3.5*cm, 5*cm, 2*cm, 2*cm, 3.5*cm], repeatRows=1)
            slas_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), cor_azul_escuro),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(slas_table)
            story.append(Spacer(1, 0.3*cm))
        
        # Matriz RACI
        matriz_raci = contrato.matriz_raci.all()
        if matriz_raci:
            story.append(PageBreak())
            story.append(Paragraph('<b>Matriz RACI</b>', title_style))
            
            raci_data = [[
                Paragraph('<b>Atividade</b>', header_cell_style),
                Paragraph('<b>Fase</b>', header_cell_style),
                Paragraph('<b>R</b>', header_cell_style),
                Paragraph('<b>A</b>', header_cell_style),
                Paragraph('<b>C</b>', header_cell_style),
                Paragraph('<b>I</b>', header_cell_style)
            ]]
            for raci in matriz_raci:
                fase_display = {
                    'planejamento': 'Planejamento',
                    'implantacao': 'Implantação',
                    'execucao': 'Execução',
                    'suporte': 'Suporte'
                }.get(raci.fase, raci.fase)
                
                raci_data.append([
                    Paragraph(str(raci.atividade), cell_style),
                    Paragraph(str(fase_display), cell_style),
                    Paragraph(str(raci.responsavel or '-'), cell_style),
                    Paragraph(str(raci.aprovador or '-'), cell_style),
                    Paragraph(str(raci.consultado or '-'), cell_style),
                    Paragraph(str(raci.informado or '-'), cell_style),
                ])
            
            raci_table = Table(raci_data, colWidths=[4*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm], repeatRows=1)
            raci_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), cor_azul_escuro),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(raci_table)
            story.append(Spacer(1, 0.3*cm))
        
        # Cronograma Detalhado
        if plano.cronograma_detalhado:
            story.append(PageBreak())
            story.append(Paragraph('<b>Cronograma Detalhado</b>', title_style))
            
            crono_data = [[
                Paragraph('<b>Nome do Marco</b>', header_cell_style),
                Paragraph('<b>Data</b>', header_cell_style),
                Paragraph('<b>Descrição</b>', header_cell_style)
            ]]
            for marco in plano.cronograma_detalhado:
                from datetime import datetime
                try:
                    if isinstance(marco.get('data'), str):
                        data_obj = datetime.strptime(marco['data'], '%Y-%m-%d')
                        data_str = data_obj.strftime('%d/%m/%Y')
                    else:
                        data_str = marco.get('data', '-')
                except:
                    data_str = str(marco.get('data', '-'))
                
                crono_data.append([
                    Paragraph(str(marco.get('nome', '-')), cell_style),
                    Paragraph(str(data_str), cell_style),
                    Paragraph(str(marco.get('descricao', '-')), cell_style),
                ])
            
            crono_table = Table(crono_data, colWidths=[5*cm, 3*cm, 8*cm], repeatRows=1)
            crono_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), cor_azul_escuro),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(crono_table)
            story.append(Spacer(1, 0.3*cm))
        
        # Plano de Comunicação
        if plano.plano_comunicacao:
            story.append(PageBreak())
            story.append(Paragraph('<b>Plano de Comunicação</b>', title_style))
            
            if isinstance(plano.plano_comunicacao, str):
                try:
                    plano_com = json.loads(plano.plano_comunicacao)
                except:
                    plano_com = {}
            else:
                plano_com = plano.plano_comunicacao
            
            if plano_com.get('stakeholders'):
                story.append(Paragraph('<b>Stakeholders</b>', heading_style))
                stakeholders_data = [[
                    Paragraph('<b>Nome</b>', header_cell_style),
                    Paragraph('<b>Papel</b>', header_cell_style),
                    Paragraph('<b>Frequência</b>', header_cell_style),
                    Paragraph('<b>Canais</b>', header_cell_style)
                ]]
                for stakeholder in plano_com.get('stakeholders', []):
                    canais = stakeholder.get('canais', [])
                    canais_str = ', '.join(canais) if isinstance(canais, list) else str(canais)
                    stakeholders_data.append([
                        Paragraph(str(stakeholder.get('nome', '-')), cell_style),
                        Paragraph(str(stakeholder.get('papel', '-')), cell_style),
                        Paragraph(str(stakeholder.get('frequencia_comunicacao', '-')), cell_style),
                        Paragraph(str(canais_str), cell_style),
                    ])
                
                stakeholders_table = Table(stakeholders_data, colWidths=[4*cm, 4*cm, 3*cm, 5*cm], repeatRows=1)
                stakeholders_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), cor_azul_escuro),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(stakeholders_table)
                story.append(Spacer(1, 0.3*cm))
            
            if plano_com.get('reunioes'):
                story.append(Paragraph('<b>Reuniões</b>', heading_style))
                reunioes_data = [[
                    Paragraph('<b>Tipo</b>', header_cell_style),
                    Paragraph('<b>Frequência</b>', header_cell_style),
                    Paragraph('<b>Participantes</b>', header_cell_style),
                    Paragraph('<b>Objetivo</b>', header_cell_style)
                ]]
                for reuniao in plano_com.get('reunioes', []):
                    participantes = reuniao.get('participantes', [])
                    participantes_str = ', '.join(participantes) if isinstance(participantes, list) else str(participantes)
                    reunioes_data.append([
                        Paragraph(str(reuniao.get('tipo', '-')), cell_style),
                        Paragraph(str(reuniao.get('frequencia', '-')), cell_style),
                        Paragraph(str(participantes_str), cell_style),
                        Paragraph(str(reuniao.get('objetivo', '-')), cell_style),
                    ])
                
                reunioes_table = Table(reunioes_data, colWidths=[3*cm, 3*cm, 4*cm, 6*cm], repeatRows=1)
                reunioes_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), cor_azul_escuro),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(reunioes_table)
                story.append(Spacer(1, 0.3*cm))
        
        # Template de Status Report
        if plano.template_status_report:
            story.append(PageBreak())
            story.append(Paragraph('<b>Template de Status Report</b>', title_style))
            story.append(Paragraph(plano.template_status_report.replace('\n', '<br/>'), normal_style))
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(f'<b>Frequência:</b> {plano.frequencia_status_report.title()}', normal_style))
        
        # Rodapé
        story.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER,
        )
        story.append(Paragraph('<b>ALLTECH - Soluções em Tecnologia</b>', footer_style))
        story.append(Paragraph('www.alltechsolucoes.com.br', footer_style))
        from django.utils import timezone
        story.append(Paragraph(f'Documento gerado em {timezone.now().strftime("%d/%m/%Y %H:%M")}', footer_style))
        story.append(Paragraph('Este documento foi gerado automaticamente pelo sistema de gestão de contratos.', footer_style))
        
        # Construir PDF
        doc.build(story)
        
        # Retornar PDF
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response
        
    except Exception as e:
        import traceback
        print(f"Erro ao gerar PDF: {traceback.format_exc()}")
        messages.error(request, f'Erro ao gerar PDF: {str(e)}')
        return redirect('plano_trabalho_detail', pk=pk)


# ==================== CUSTOMER SUCCESS - FEEDBACK SPRINT/OS ====================

@login_required
@group_required("Admin", "Gerente", "Customer Success")
def customer_success_list(request):
    """Lista feedbacks pendentes de Sprint/OS faturadas"""
    from django.db.models import Q
    from django.utils import timezone
    
    # Filtros
    status_filter = request.GET.get('status', '')
    
    # Query base
    feedbacks = FeedbackSprintOS.objects.select_related(
        'sprint', 'ordem_servico', 'cliente', 'contrato', 'projeto', 'gerente_sucessos'
    ).all()
    
    # Aplicar filtros
    if status_filter:
        feedbacks = feedbacks.filter(status=status_filter)
    
    # Ordenar por data de criação (mais recentes primeiro)
    feedbacks = feedbacks.order_by('-criado_em')
    
    # Estatísticas
    total_pendentes = FeedbackSprintOS.objects.filter(status='pendente').count()
    total_em_contato = FeedbackSprintOS.objects.filter(status='em_contato').count()
    total_respondidos = FeedbackSprintOS.objects.filter(status='respondido').count()
    total_concluidos = FeedbackSprintOS.objects.filter(status='concluido').count()
    
    context = {
        'feedbacks': feedbacks,
        'status_filter': status_filter,
        'total_pendentes': total_pendentes,
        'total_em_contato': total_em_contato,
        'total_respondidos': total_respondidos,
        'total_concluidos': total_concluidos,
    }
    
    return render(request, 'customer_success/list.html', context)


@login_required
@group_required("Admin", "Gerente", "Customer Success")
def customer_success_detail(request, pk):
    """Visualiza e edita um feedback específico"""
    feedback = get_object_or_404(
        FeedbackSprintOS.objects.select_related(
            'sprint', 'ordem_servico', 'cliente', 'contrato', 'projeto', 'gerente_sucessos'
        ),
        pk=pk
    )
    
    if request.method == 'POST':
        form = FeedbackSprintOSForm(request.POST, instance=feedback)
        if form.is_valid():
            # Atualizar data de contato se status mudou para 'em_contato'
            if form.cleaned_data['status'] == 'em_contato' and not feedback.data_contato:
                from django.utils import timezone
                feedback.data_contato = timezone.now()
            
            # Atualizar data de resposta se status mudou para 'respondido'
            if form.cleaned_data['status'] == 'respondido' and not feedback.data_resposta:
                from django.utils import timezone
                feedback.data_resposta = timezone.now()
            
            # Atribuir gerente de sucessos se não tiver
            if not feedback.gerente_sucessos and form.cleaned_data.get('gerente_sucessos'):
                feedback.gerente_sucessos = form.cleaned_data['gerente_sucessos']
            
            form.save()
            messages.success(request, 'Feedback atualizado com sucesso!')
            return redirect('customer_success_detail', pk=feedback.pk)
        else:
            messages.error(request, 'Erro ao atualizar feedback. Verifique os campos.')
    else:
        form = FeedbackSprintOSForm(instance=feedback)
    
    context = {
        'feedback': feedback,
        'form': form,
    }
    
    return render(request, 'customer_success/detail.html', context)


@login_required
@group_required("Admin", "Gerente", "Customer Success")
def customer_success_criar_ticket(request):
    """Cria um ticket de contato manualmente"""
    if request.method == 'POST':
        form = CriarTicketContatoForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            # Se não tiver sprint ou OS, definir cliente e contrato como obrigatórios
            if not ticket.sprint and not ticket.ordem_servico:
                if not ticket.cliente or not ticket.contrato:
                    messages.error(request, 'Cliente e Contrato são obrigatórios quando não há Sprint ou OS vinculada.')
                    return render(request, 'customer_success/criar_ticket.html', {'form': form})
            
            # Se tiver sprint, preencher automaticamente cliente, contrato e projeto
            if ticket.sprint:
                ticket.cliente = ticket.sprint.projeto.contrato.cliente
                ticket.contrato = ticket.sprint.projeto.contrato
                ticket.projeto = ticket.sprint.projeto
            
            # Se tiver OS, preencher automaticamente cliente e contrato
            if ticket.ordem_servico:
                ticket.cliente = ticket.ordem_servico.cliente
                ticket.contrato = ticket.ordem_servico.contrato
                # OrdemServico tem relação OneToOne com Sprint (via related_name 'sprint')
                try:
                    if ticket.ordem_servico.sprint:
                        ticket.projeto = ticket.ordem_servico.sprint.projeto
                except AttributeError:
                    pass
                # Se não tiver sprint, buscar projeto via contrato
                if not ticket.projeto and ticket.contrato and ticket.contrato.projetos.exists():
                    ticket.projeto = ticket.contrato.projetos.first()
            
            ticket.status = 'pendente'
            ticket.save()
            messages.success(request, 'Ticket de contato criado com sucesso!')
            return redirect('customer_success_detail', pk=ticket.pk)
        else:
            messages.error(request, 'Erro ao criar ticket. Verifique os campos.')
    else:
        # Pré-selecionar cliente se passado via GET
        cliente_id = request.GET.get('cliente')
        if cliente_id:
            try:
                cliente = Cliente.objects.get(pk=cliente_id)
                form = CriarTicketContatoForm(initial={'cliente': cliente})
            except Cliente.DoesNotExist:
                form = CriarTicketContatoForm()
        else:
            form = CriarTicketContatoForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'customer_success/criar_ticket.html', context)


@login_required
@group_required("Admin", "Gerente", "Customer Success")
def customer_success_delete(request, pk):
    """Exclui um ticket de Customer Success"""
    try:
        ticket = FeedbackSprintOS.objects.get(pk=pk)
    except FeedbackSprintOS.DoesNotExist:
        messages.error(request, 'Ticket não encontrado.')
        return redirect('customer_success_list')
    
    if request.method == 'POST':
        numero_ticket = ticket.numero_ticket
        ticket.delete()
        messages.success(request, f'Ticket {numero_ticket} excluído com sucesso!')
        return redirect('customer_success_list')
    
    context = {
        'ticket': ticket,
    }
    return render(request, 'customer_success/confirm_delete.html', context)
