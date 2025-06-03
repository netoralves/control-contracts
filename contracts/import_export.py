# contracts/import_export.py
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import (
    Cliente,
    Contrato,
    ItemContrato,
    ItemFornecedor,
    OrdemFornecimento,
    OrdemServico,
)


def export_all_data():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(title, queryset, fields):
        ws = wb.create_sheet(title)
        ws.append(fields)
        for obj in queryset:
            row = [
                (
                    getattr(obj, field)
                    if not callable(getattr(obj, field))
                    else getattr(obj, field)()
                )
                for field in fields
            ]
            ws.append(row)

    # Exporta dados de todas as entidades
    add_sheet(
        "Clientes",
        Cliente.objects.all(),
        [
            "id",
            "nome_razao_social",
            "nome_fantasia",
            "tipo_cliente",
            "tipo_pessoa",
            "cnpj_cpf",
            "cidade",
            "estado",
            "email_contato",
            "ativo",
        ],
    )

    add_sheet(
        "Contratos",
        Contrato.objects.all(),
        [
            "id",
            "numero_contrato",
            "cliente_id",
            "vigencia",
            "data_assinatura",
            "data_fim",
            "valor_global",
            "situacao",
            "fornecedores",
        ],
    )

    add_sheet(
        "ItensContrato",
        ItemContrato.objects.all(),
        [
            "id",
            "contrato_id",
            "lote",
            "numero_item",
            "descricao",
            "tipo",
            "unidade",
            "quantidade",
            "valor_unitario",
            "valor_total",
            "vigencia_produto",
        ],
    )

    add_sheet(
        "ItensFornecedor",
        ItemFornecedor.objects.all(),
        [
            "id",
            "fornecedor",
            "outro_fornecedor",
            "tipo",
            "sku",
            "descricao",
            "unidade",
            "valor_unitario",
        ],
    )

    add_sheet(
        "OrdensFornecimento",
        OrdemFornecimento.objects.all(),
        [
            "id",
            "numero_of",
            "cliente_id",
            "contrato_id",
            "item_contrato_id",
            "item_fornecedor_id",
            "unidade",
            "quantidade",
            "valor_unitario",
            "valor_total",
            "status",
            "data_ativacao",
            "data_faturamento",
        ],
    )

    add_sheet(
        "OrdensServico",
        OrdemServico.objects.all(),
        [
            "id",
            "numero_os",
            "cliente_id",
            "contrato_id",
            "item_contrato_id",
            "item_fornecedor_id",
            "gerente_projetos",
            "consultor_tecnico",
            "unidade",
            "quantidade",
            "valor_unitario",
            "valor_total",
            "data_inicio",
            "hora_inicio",
            "data_termino",
            "hora_termino",
            "status",
            "data_emissao_trd",
            "data_faturamento",
            "horas_consultor",
            "horas_gerente",
            "horas_totais",
        ],
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="dados_contratos.xlsx"'
    wb.save(response)
    return response
